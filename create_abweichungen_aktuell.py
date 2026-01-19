"""
Erstellt aktualisierte Excel-Liste mit allen OFFENEN Abweichungen
OHNE die als IGNORIEREN markierten Eintraege
Stand: 2026-01-18
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Workbook erstellen
wb = Workbook()
ws = wb.active
ws.title = "Offene Abweichungen"

# Styles definieren
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
kritisch_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
hoch_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
mittel_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
niedrig_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header
headers = [
    "Nr", "Formular", "Bereich", "Problem", "Schweregrad",
    "Datei", "Loesung", "Status"
]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Alle OFFENEN Abweichungen (ohne IGNORIEREN und ERLEDIGT)
abweichungen = [
    # === KRITISCH ===
    (1, "frm_N_Dienstplanuebersicht", "HTML fehlt",
     "HTML-Datei existiert NICHT in forms3 Ordner - Formular komplett erstellen",
     "Kritisch", "forms3/", "Access-Export + HTML erstellen", "OFFEN"),

    (2, "frm_VA_Planungsuebersicht", "Bridge.query",
     "Bridge.query() Methode existiert nicht - Formular laedt keine Daten",
     "Kritisch", "frm_VA_Planungsuebersicht.html", "fetch('/api/...') statt Bridge.query()", "OFFEN"),

    (3, "frm_KD_Verrechnungssaetze", "Falsche Tabelle",
     "API verwendet tbl_KD_Kundenpreise - Tabelle existiert NICHT",
     "Kritisch", "api_server.py", "Korrekte Tabelle: tbl_KD_Standardpreise", "OFFEN"),

    (4, "frm_va_Auftragstamm", "Anfragen-Panel",
     "Panel verwendet REST-API statt VBA Bridge - kann keine echten E-Mails senden",
     "Kritisch", "frm_va_Auftragstamm.html", "Auf VBA Bridge Port 5002 umstellen", "OFFEN"),

    # === HOCH ===
    (5, "frm_va_Auftragstamm", "Bedingte Format.",
     "Soll>Ist Rot-Markierung fehlt: Wenn MA_Anzahl_Ist < MA_Anzahl soll Zeile rot sein",
     "Hoch", "frm_va_Auftragstamm.html", "CSS-Klasse bei Soll>Ist + rot einfaerben", "OFFEN"),

    (6, "frm_va_Auftragstamm", "Button fehlt",
     "btnAuftrBerech (Berechnung) fehlt im HTML",
     "Hoch", "frm_va_Auftragstamm.html", "Button + VBA Bridge Call hinzufuegen", "OFFEN"),

    (7, "frm_va_Auftragstamm", "Tab Zusatzdateien",
     "Upload-Funktion fehlt - nur Download implementiert",
     "Hoch", "frm_va_Auftragstamm.html", "POST /api/auftraege/<id>/dateien", "OFFEN"),

    (8, "frm_va_Auftragstamm", "Tab Rechnung",
     "API-Endpoint /api/auftraege/<id>/rechnungen fehlt",
     "Hoch", "api_server.py", "SELECT * FROM tbl_Rch_Kopf WHERE VA_ID=?", "OFFEN"),

    (9, "frm_va_Auftragstamm", "Tab Kosten",
     "API-Endpoint /api/auftraege/<id>/kosten fehlt",
     "Hoch", "api_server.py", "SELECT * FROM tbl_VA_Kosten WHERE VA_ID=?", "OFFEN"),

    (10, "frm_va_Auftragstamm", "Subform",
     "sub_VA_Tag: Ausgewaehlter Tag wird nicht an Parent gemeldet",
     "Hoch", "sub_VA_Tag.html", "postMessage({type:'TAG_SELECTED', datum})", "OFFEN"),

    (11, "frm_MA_Mitarbeiterstamm", "API fehlt",
     "Endpoint /api/mitarbeiter/<id>/zuordnungen fehlt (Einsatzuebersicht + Dienstplan Tab)",
     "Hoch", "api_server.py", "SELECT * FROM tbl_MA_VA_Planung WHERE MA_ID=?", "OFFEN"),

    (12, "frm_MA_Mitarbeiterstamm", "API fehlt",
     "Endpoint /api/mitarbeiter/<id>/nverfueg fehlt (Nicht Verfuegbar Tab)",
     "Hoch", "api_server.py", "SELECT * FROM tbl_MA_NVerfuegZeiten WHERE MA_ID=?", "OFFEN"),

    (13, "frm_MA_Mitarbeiterstamm", "API fehlt",
     "Endpoint /api/mitarbeiter/<id>/zeitkonto fehlt",
     "Hoch", "api_server.py", "Zeitkonto-Daten Endpoint", "OFFEN"),

    (14, "frm_N_Dienstplanuebersicht", "VBA Bridge",
     "VBA-Funktionen DP_Anzeigen, DP_Drucken, DP_Senden fehlen in Whitelist",
     "Hoch", "vba_bridge.py", "Zu allowed_functions hinzufuegen", "OFFEN"),

    (15, "frm_N_Dienstplanuebersicht", "API fehlt",
     "Endpoint /api/dienstplan/uebersicht fehlt",
     "Hoch", "api_server.py", "Aggregierte Dienstplan-Daten", "OFFEN"),

    (16, "frm_KD_Verrechnungssaetze", "Felder",
     "Feldnamen im HTML stimmen nicht mit DB ueberein",
     "Hoch", "frm_KD_Verrechnungssaetze.html", "data-field aus tbl_KD_Standardpreise", "OFFEN"),

    (17, "sub_Rch_Kopf (MA)", "Neues Formular",
     "Page subrch muss als separates HTML-Formular erstellt werden",
     "Hoch", "NEUES FORMULAR", "frm_MA_Subrch.html erstellen", "OFFEN"),

    (18, "sub_Rch_Kopf (MA)", "API Endpoints",
     "7 neue API-Endpoints fuer MA-Rechnungs-Subform benoetigt",
     "Hoch", "api_server.py", "/api/ma/<id>/rechnungen etc.", "OFFEN"),

    (19, "sub_Rch_Kopf (MA)", "VBA Events",
     "~15 VBA Events muessen nach JavaScript portiert werden",
     "Hoch", "Form_sub_Rch_Kopf.bas", "AfterUpdate, OnClick in JS", "OFFEN"),

    (20, "Anfragen-Panel", "VBA Funktion",
     "zmd_Mail.Anfragen() muss ueber VBA Bridge aufgerufen werden",
     "Hoch", "frm_va_Auftragstamm.logic.js", "fetch Port 5002 wie Schnellauswahl", "OFFEN"),

    (21, "Anfragen-Panel", "Parameter",
     "ma_id, va_id, vadatum_id, vastart_id muessen korrekt uebergeben werden",
     "Hoch", "frm_va_Auftragstamm.logic.js", "Alle 4 IDs aus state extrahieren", "OFFEN"),

    # === MITTEL ===
    (22, "frm_va_Auftragstamm", "Bedingte Format.",
     "Status-Farben in Auftragsliste fehlen: In Planung=gelb, Beendet=gruen",
     "Mittel", "frm_va_Auftragstamm.html", "Status-abhaengige Hintergrundfarben", "OFFEN"),

    (23, "frm_va_Auftragstamm", "Feldabhaengigkeit",
     "Objekt-Auswahl aktualisiert nicht automatisch Ansprechpartner-Dropdown",
     "Mittel", "frm_va_Auftragstamm.logic.js", "onchange Objekt → Ansprechpartner laden", "OFFEN"),

    (24, "frm_va_Auftragstamm", "Filter",
     "Datumsbereich-Filter (Von/Bis oben) filtert Auftragsliste nicht",
     "Mittel", "frm_va_Auftragstamm.logic.js", "applyDateFilter() implementieren", "OFFEN"),

    (25, "frm_va_Auftragstamm", "Subform",
     "sub_MA_VA_Zuordnung: Doppelklick auf MA oeffnet nicht Mitarbeiterstamm",
     "Mittel", "sub_MA_VA_Zuordnung.html", "dblclick → Shell-Navigation mit MA_ID", "OFFEN"),

    (26, "frm_MA_Mitarbeiterstamm", "Feldname",
     "Hat_Fahrerausweis: HTML hat data-field, Logic.js setzt 'Fahrerlaubnis'",
     "Mittel", "frm_MA_Mitarbeiterstamm.logic.js", "setCheckbox('Hat_Fahrerausweis',...)", "OFFEN"),

    (27, "frm_KD_Kundenstamm", "Button",
     "btnUmsAuswert oeffnet falsches Formular?",
     "Mittel", "frm_KD_Kundenstamm.html", "Navigation-Ziel mit VBA abgleichen", "PRUEFEN"),

    (28, "frm_KD_Kundenstamm", "API",
     "Endpoint /api/adressen moeglicherweise nicht implementiert",
     "Mittel", "api_server.py", "Endpoint implementieren falls benoetigt", "PRUEFEN"),

    (29, "frm_VA_Planungsuebersicht", "Filter",
     "Mitarbeiter-Filter laedt alle MA statt nur aktive",
     "Mittel", "frm_VA_Planungsuebersicht.logic.js", "?aktiv=true Parameter", "OFFEN"),

    (30, "frm_KD_Verrechnungssaetze", "CRUD",
     "Nur Lesen implementiert - Bearbeiten/Loeschen fehlt",
     "Mittel", "frm_KD_Verrechnungssaetze.html", "PUT/DELETE + onclick Handler", "OFFEN"),

    (31, "frm_KD_Verrechnungssaetze", "Kundenfilter",
     "Dropdown 'Kunde' filtert Liste nicht",
     "Mittel", "frm_KD_Verrechnungssaetze.logic.js", "onchange → Liste mit KD_ID filtern", "OFFEN"),

    (32, "sub_Rch_Kopf (MA)", "Einbettung",
     "frm_MA_Mitarbeiterstamm muss iframe fuer sub_Rch_Kopf einbetten",
     "Mittel", "frm_MA_Mitarbeiterstamm.html", "Tab Rechnungen iframe", "OFFEN"),

    (33, "sub_Rch_Kopf (MA)", "PostMessage",
     "Kommunikation Parent<->Subform via postMessage definieren",
     "Mittel", "sub_Rch_Kopf.html", "MA_SELECTED, REFRESH_DATA Events", "OFFEN"),

    (34, "Anfragen-Panel", "Feedback",
     "Keine Rueckmeldung ob E-Mail erfolgreich gesendet",
     "Mittel", "frm_va_Auftragstamm.html", "response.success → Status + Toast", "OFFEN"),

    (35, "Anfragen-Panel", "Fehler-Handling",
     "Bei VBA Bridge Fehler keine Fehlermeldung an User",
     "Mittel", "frm_va_Auftragstamm.logic.js", "try/catch mit Fehlermeldung", "OFFEN"),

    (36, "frm_va_Auftragstamm", "Eventdaten",
     "loadEventdaten() verwendet veraltete Variable 'currentRecord' statt 'state.currentAuftrag'",
     "Mittel", "frm_va_Auftragstamm.html", "Variable vereinheitlichen", "OFFEN"),

    (37, "frm_VA_Planungsuebersicht", "Duplikat",
     "2 Formulare existieren: forms3/ UND forms/ - eines loeschen",
     "Mittel", "forms3/ und forms/", "Kanonisches Formular definieren", "PRUEFEN"),

    # === NIEDRIG ===
    (38, "frm_MA_Mitarbeiterstamm", "Dropdown",
     "Anstellungsart Dropdown unvollstaendig (nur 3 statt 6+ Optionen)",
     "Niedrig", "frm_MA_Mitarbeiterstamm.html", "Optionen erweitern", "OFFEN"),

    (39, "frm_MA_VA_Schnellauswahl", "Zeiten",
     "Beginn/Ende-Zeiten (von/bis) fehlen in der Anzeige",
     "Niedrig", "frm_MA_VA_Schnellauswahl.logic.js", "API erweitert - Logic pruefen", "PRUEFEN"),
]

# Daten einfuegen
for row_idx, row_data in enumerate(abweichungen, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Schweregrad farbig markieren
        if col_idx == 5:  # Schweregrad
            if value == "Kritisch":
                cell.fill = kritisch_fill
                cell.font = Font(bold=True)
            elif value == "Hoch":
                cell.fill = hoch_fill
            elif value == "Mittel":
                cell.fill = mittel_fill
            elif value == "Niedrig":
                cell.fill = niedrig_fill

        # Status farbig markieren
        if col_idx == 8:  # Status
            if value == "OFFEN":
                cell.fill = kritisch_fill
            elif value == "PRUEFEN":
                cell.fill = mittel_fill

# Spaltenbreiten anpassen
column_widths = [5, 28, 18, 60, 12, 35, 45, 10]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Zeile 1 fixieren (Header)
ws.freeze_panes = "A2"

# Autofilter aktivieren
ws.auto_filter.ref = f"A1:H{len(abweichungen)+1}"

# Zusammenfassung Sheet
ws2 = wb.create_sheet("Zusammenfassung")
ws2["A1"] = "ZUSAMMENFASSUNG - Offene Abweichungen"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws2["A4"] = "Ohne IGNORIEREN und ERLEDIGT Eintraege"

ws2["A6"] = "Schweregrad"
ws2["B6"] = "Anzahl"
ws2["C6"] = "Prozent"

for col in range(1, 4):
    ws2.cell(row=6, column=col).font = Font(bold=True)
    ws2.cell(row=6, column=col).fill = header_fill
    ws2.cell(row=6, column=col).font = Font(bold=True, color="FFFFFF")

summary_data = [
    ("Kritisch", 4, "10%"),
    ("Hoch", 17, "44%"),
    ("Mittel", 16, "41%"),
    ("Niedrig", 2, "5%"),
    ("GESAMT", 39, "100%"),
]

for row_idx, data in enumerate(summary_data, 7):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 11:  # Gesamt-Zeile
            cell.font = Font(bold=True)
        if col_idx == 1:
            if value == "Kritisch":
                cell.fill = kritisch_fill
            elif value == "Hoch":
                cell.fill = hoch_fill
            elif value == "Mittel":
                cell.fill = mittel_fill
            elif value == "Niedrig":
                cell.fill = niedrig_fill

# Bereich-Aufschluesselung
ws2["A14"] = "Nach Formular/Bereich"
ws2["A14"].font = Font(bold=True)

ws2["A15"] = "Bereich"
ws2["B15"] = "Anzahl"
for col in range(1, 3):
    ws2.cell(row=15, column=col).fill = header_fill
    ws2.cell(row=15, column=col).font = Font(bold=True, color="FFFFFF")

bereich_data = [
    ("frm_va_Auftragstamm", 10),
    ("frm_MA_Mitarbeiterstamm", 4),
    ("frm_N_Dienstplanuebersicht", 3),
    ("frm_VA_Planungsuebersicht", 3),
    ("frm_KD_Verrechnungssaetze", 4),
    ("frm_KD_Kundenstamm", 2),
    ("sub_Rch_Kopf (MA)", 5),
    ("Anfragen-Panel", 6),
    ("frm_MA_VA_Schnellauswahl", 1),
    ("Sonstige", 1),
]

for row_idx, data in enumerate(bereich_data, 16):
    for col_idx, value in enumerate(data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12

# Prioritaeten Sheet
ws3 = wb.create_sheet("Prioritaeten")
ws3["A1"] = "PRIORITAETEN - Reihenfolge der Bearbeitung"
ws3["A1"].font = Font(bold=True, size=14)

ws3["A3"] = "PHASE 1: KRITISCH (Blocker - System funktioniert nicht)"
ws3["A3"].font = Font(bold=True)
ws3["A3"].fill = kritisch_fill

phase1 = [
    "1. #1: Dienstplanuebersicht HTML erstellen",
    "2. #2: Planungsuebersicht Bridge.query → fetch() ersetzen",
    "3. #3: Verrechnungssaetze korrekte Tabelle verwenden",
    "4. #4: Anfragen-Panel auf VBA Bridge umstellen",
]
for i, item in enumerate(phase1, 4):
    ws3[f"A{i}"] = item

ws3["A9"] = "PHASE 2: HOCH - API Endpoints (Backend)"
ws3["A9"].font = Font(bold=True)
ws3["A9"].fill = hoch_fill

phase2 = [
    "5. #8-9: /api/auftraege/<id>/rechnungen + /kosten",
    "6. #11-13: /api/mitarbeiter/<id>/zuordnungen + nverfueg + zeitkonto",
    "7. #15: /api/dienstplan/uebersicht",
    "8. #18: MA-Rechnungs-Subform Endpoints",
]
for i, item in enumerate(phase2, 10):
    ws3[f"A{i}"] = item

ws3["A15"] = "PHASE 3: HOCH - Frontend Funktionen"
ws3["A15"].font = Font(bold=True)
ws3["A15"].fill = hoch_fill

phase3 = [
    "9. #5: Soll>Ist Rot-Markierung",
    "10. #10: sub_VA_Tag postMessage Kommunikation",
    "11. #17-19: sub_Rch_Kopf neues Formular + VBA Events",
    "12. #20-21: Anfragen-Panel VBA Integration",
]
for i, item in enumerate(phase3, 16):
    ws3[f"A{i}"] = item

ws3["A21"] = "PHASE 4: MITTEL - Verbesserungen"
ws3["A21"].font = Font(bold=True)
ws3["A21"].fill = mittel_fill

phase4 = [
    "13. Status-Farben, Feldabhaengigkeiten, Filter",
    "14. Doppelklick-Navigation, Feedback, Fehler-Handling",
    "15. CRUD fuer Verrechnungssaetze",
]
for i, item in enumerate(phase4, 22):
    ws3[f"A{i}"] = item

ws3.column_dimensions["A"].width = 70

# Speichern
output_path = r"C:\Users\guenther.siegert\Desktop\ABWEICHUNGEN_AKTUELL_18012026.xlsx"
wb.save(output_path)
print(f"Excel-Datei erstellt: {output_path}")
print(f"\nGesamt: {len(abweichungen)} offene Abweichungen")
print(f"- Kritisch: 4")
print(f"- Hoch: 17")
print(f"- Mittel: 16")
print(f"- Niedrig: 2")
