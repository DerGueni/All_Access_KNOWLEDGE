"""
Erstellt Excel-Liste mit allen Abweichungen aus der Senior Master Agent Prüfung
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Workbook erstellen
wb = Workbook()
ws = wb.active
ws.title = "Abweichungen HTML-Forms"

# Styles definieren
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
gray_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header
headers = [
    "Nr", "Formular", "Bereich/Tab", "Problem", "Schweregrad",
    "Datei", "Zeile", "Lösungsvorschlag", "Aktion", "Status"
]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Daten - Alle gefundenen Abweichungen
abweichungen = [
    # Auftragstamm
    (1, "frm_va_Auftragstamm", "Eventdaten Tab", "loadEventdaten() verwendet veraltete Variable 'currentRecord' statt 'state.currentAuftrag'", "Mittel",
     "frm_va_Auftragstamm.html", "5277, 5300", "Variable zu state.currentAuftrag vereinheitlichen", "BEHEBEN", "Offen"),

    (2, "frm_va_Auftragstamm", "Zusatzdateien Tab", "Attachment-Download funktioniert nur in WebView2, nicht im Browser", "Niedrig",
     "frm_va_Auftragstamm.html", "4880-4885", "REST-API Fallback mit /zusatzdateien/{id}/download implementieren", "IGNORIEREN", "Browser-only Feature"),

    (3, "frm_va_Auftragstamm", "Layout", "Auftragsliste war zu schmal - Lücke zur Einsatzliste", "Hoch",
     "frm_va_Auftragstamm.html", "882-885", "width: 580px → 700px (BEREITS BEHOBEN)", "ERLEDIGT", "Behoben 18.01.26"),

    # Kundenstamm
    (4, "frm_KD_Kundenstamm", "Zusatzdateien Tab", "tbody ID falsch: JS sucht 'zusatzdateienBody', HTML hat 'dateienBody'", "Hoch",
     "frm_KD_Kundenstamm.html", "3197", "ID korrigiert zu 'dateienBody' (BEREITS BEHOBEN)", "ERLEDIGT", "Behoben 18.01.26"),

    (5, "frm_KD_Kundenstamm", "Buttons", "btnUmsAuswert öffnet falsches Formular? (frm_KD_Umsatzauswertung vs frm_Auswertung_Kunde_Jahr)", "Mittel",
     "frm_KD_Kundenstamm.html", "899", "Navigation-Ziel mit VBA-Referenz abgleichen", "PRÜFEN", "Unklar"),

    (6, "frm_KD_Kundenstamm", "Feldnamen", "Casing-Inkonsistenz: kun_PLZ vs kun_plz, kun_email vs kun_eMail", "Niedrig",
     "frm_KD_Kundenstamm.html", "diverse", "Feldnamen mit tbl_KD_Kundenstamm abgleichen", "IGNORIEREN", "Funktioniert trotzdem"),

    (7, "frm_KD_Kundenstamm", "Filter", "cboSuchSuchF (Sortfeld) Dropdown ist statisch, sollte dynamisch sein", "Niedrig",
     "frm_KD_Kundenstamm.html", "1570-1572", "Dropdown dynamisch aus DB befüllen", "IGNORIEREN", "Nice-to-have"),

    (8, "frm_KD_Kundenstamm", "API", "Endpoint /api/adressen möglicherweise nicht implementiert", "Mittel",
     "api_server.py", "N/A", "Endpoint implementieren oder Alternative nutzen", "PRÜFEN", "Unklar ob benötigt"),

    # Mitarbeiterstamm
    (9, "frm_MA_Mitarbeiterstamm", "Einsatzübersicht Tab", "API-Endpoint /api/mitarbeiter/<id>/zuordnungen fehlt", "Hoch",
     "api_server.py", "N/A", "Endpoint hinzufügen: SELECT * FROM tbl_MA_VA_Planung WHERE MA_ID=?", "BEHEBEN", "Offen"),

    (10, "frm_MA_Mitarbeiterstamm", "Dienstplan Tab", "API-Endpoint /api/mitarbeiter/<id>/zuordnungen fehlt (gleich wie #9)", "Hoch",
     "api_server.py", "N/A", "Gleicher Endpoint wie #9", "BEHEBEN", "Offen"),

    (11, "frm_MA_Mitarbeiterstamm", "Nicht Verfügbar Tab", "API-Endpoint /api/mitarbeiter/<id>/nverfueg fehlt", "Hoch",
     "api_server.py", "N/A", "Endpoint hinzufügen: SELECT * FROM tbl_MA_NVerfuegZeiten WHERE MA_ID=?", "BEHEBEN", "Offen"),

    (12, "frm_MA_Mitarbeiterstamm", "Zeitkonto Tab", "API-Endpoint /api/mitarbeiter/<id>/zeitkonto fehlt", "Hoch",
     "api_server.py", "N/A", "Endpoint hinzufügen für Zeitkonto-Daten", "BEHEBEN", "Offen"),

    (13, "frm_MA_Mitarbeiterstamm", "Feldnamen", "Hat_Fahrerausweis: HTML hat data-field, Logic.js setzt 'Fahrerlaubnis'", "Mittel",
     "frm_MA_Mitarbeiterstamm.logic.js", "503", "setCheckbox('Hat_Fahrerausweis', rec.Hat_Fahrerausweis)", "BEHEBEN", "Offen"),

    (14, "frm_MA_Mitarbeiterstamm", "Buttons", "12 Buttons nutzen nur Bridge.sendEvent(), kein REST-API Fallback", "Mittel",
     "frm_MA_Mitarbeiterstamm.html", "diverse", "REST-API Fallback wenn Bridge nicht verfügbar", "IGNORIEREN", "Nur für Browser-Modus relevant"),

    (15, "frm_MA_Mitarbeiterstamm", "Dropdown", "Anstellungsart Dropdown unvollständig (nur 3 statt 6+ Optionen)", "Niedrig",
     "frm_MA_Mitarbeiterstamm.html", "1170-1175", "Optionen erweitern: Mini, Inaktiv, etc.", "BEHEBEN", "Offen"),

    # Schnellauswahl
    (16, "frm_MA_VA_Schnellauswahl", "MA-Liste", "Beginn/Ende-Zeiten (von/bis) fehlen in der Anzeige", "Mittel",
     "frm_MA_VA_Schnellauswahl.logic.js", "500-506", "API bereits erweitert (17.01.26) - Logic.js prüfen ob Felder genutzt werden", "PRÜFEN", "Teilweise behoben"),

    (17, "frm_MA_VA_Schnellauswahl", "Button", "btnDelSelected verwendet Bridge statt REST-API", "Niedrig",
     "frm_MA_VA_Schnellauswahl.html", "1765-1770", "Auf REST-API umstellen wie addMAToPlanung()", "IGNORIEREN", "DblClick funktioniert"),
]

# Daten einfügen
for row_idx, row_data in enumerate(abweichungen, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Schweregrad farbig markieren
        if col_idx == 5:  # Schweregrad
            if value == "Hoch":
                cell.fill = red_fill
            elif value == "Mittel":
                cell.fill = yellow_fill
            elif value == "Niedrig":
                cell.fill = green_fill

        # Aktion farbig markieren
        if col_idx == 9:  # Aktion
            if value == "BEHEBEN":
                cell.fill = red_fill
                cell.font = Font(bold=True)
            elif value == "PRÜFEN":
                cell.fill = yellow_fill
            elif value == "IGNORIEREN":
                cell.fill = gray_fill
            elif value == "ERLEDIGT":
                cell.fill = green_fill

# Spaltenbreiten anpassen
column_widths = [5, 25, 20, 50, 12, 35, 15, 50, 12, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Zeile 1 fixieren (Header)
ws.freeze_panes = "A2"

# Autofilter aktivieren
ws.auto_filter.ref = f"A1:J{len(abweichungen)+1}"

# Zusammenfassung Sheet
ws2 = wb.create_sheet("Zusammenfassung")
ws2["A1"] = "ZUSAMMENFASSUNG - Senior Master Agent Prüfung"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws2["A5"] = "Formular"
ws2["B5"] = "Gesamt"
ws2["C5"] = "BEHEBEN"
ws2["D5"] = "PRÜFEN"
ws2["E5"] = "IGNORIEREN"
ws2["F5"] = "ERLEDIGT"

for col in range(1, 7):
    ws2.cell(row=5, column=col).font = Font(bold=True)
    ws2.cell(row=5, column=col).fill = header_fill
    ws2.cell(row=5, column=col).font = Font(bold=True, color="FFFFFF")

summary_data = [
    ("frm_va_Auftragstamm", 3, 1, 0, 1, 1),
    ("frm_KD_Kundenstamm", 5, 0, 2, 2, 1),
    ("frm_MA_Mitarbeiterstamm", 7, 5, 0, 2, 0),
    ("frm_MA_VA_Schnellauswahl", 2, 0, 1, 1, 0),
    ("GESAMT", 17, 6, 3, 6, 2),
]

for row_idx, data in enumerate(summary_data, 6):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 10:  # Gesamt-Zeile
            cell.font = Font(bold=True)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12

# Speichern
output_path = r"C:\Users\guenther.siegert\Desktop\ABWEICHUNGEN_HTML_FORMS_18012026.xlsx"
wb.save(output_path)
print(f"Excel-Datei erstellt: {output_path}")
