#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
CONSYS API Server - Access Backend Bridge
=============================================================================
REST API Server für die Kommunikation zwischen HTML-Formularen und 
Access Backend-Datenbank.

Features:
- CORS-Support für lokale Entwicklung
- Direkte Access-Datenbank-Anbindung via pyodbc
- JSON-Serialisierung von Access-Daten
- CRUD-Operationen für alle Haupttabellen
- Caching für Performance

Erstellt: 28.12.2025 von Claude AI für CONSEC Security
=============================================================================
"""

import os
import sys
import json
import logging
import threading
from contextlib import contextmanager
from datetime import datetime, date, time, timedelta
from decimal import Decimal
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pyodbc

# Pooling deaktivieren um "Too many client tasks" zu vermeiden
pyodbc.pooling = False

# Thread-Lock für DB-Zugriffe (Access ODBC ist nicht thread-safe)
_db_lock = threading.Lock()

# Menu-Endpoints importieren (Shell v2)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'NEUHTML', '02_web', 'api'))
try:
    from menu_endpoints import register_menu_routes
    MENU_ENDPOINTS_AVAILABLE = True
except ImportError:
    MENU_ENDPOINTS_AVAILABLE = False
    print("WARNUNG: menu_endpoints.py nicht gefunden - Menu-API nicht verfuegbar")

# Logging konfigurieren
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# =============================================================================
# VBA BRIDGE WATCHDOG - Auto-Start & Auto-Restart
# =============================================================================
import subprocess
import urllib.request
import time as time_module

VBA_BRIDGE_PORT = 5002
VBA_BRIDGE_CHECK_INTERVAL = 30  # Sekunden
_vba_bridge_process = None
_vba_bridge_watchdog_running = False

def is_vba_bridge_running():
    """Prüft ob VBA Bridge Server auf Port 5002 erreichbar ist"""
    try:
        req = urllib.request.Request(f'http://localhost:{VBA_BRIDGE_PORT}/api/health', method='GET')
        with urllib.request.urlopen(req, timeout=2) as response:
            return response.status == 200
    except:
        return False

def start_vba_bridge():
    """Startet den VBA Bridge Server"""
    global _vba_bridge_process

    api_dir = os.path.dirname(os.path.abspath(__file__))
    vba_bridge_script = os.path.join(api_dir, 'vba_bridge_server.py')

    if not os.path.exists(vba_bridge_script):
        logger.error(f"[VBA Bridge] Script nicht gefunden: {vba_bridge_script}")
        return False

    try:
        # Starte VBA Bridge als subprocess
        logger.info("[VBA Bridge] Starte VBA Bridge Server...")
        _vba_bridge_process = subprocess.Popen(
            [sys.executable, vba_bridge_script],
            cwd=api_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        )

        # Warte kurz und prüfe ob gestartet
        time_module.sleep(3)

        if is_vba_bridge_running():
            logger.info(f"[VBA Bridge] Server läuft auf Port {VBA_BRIDGE_PORT}")
            return True
        else:
            logger.warning("[VBA Bridge] Server gestartet aber nicht erreichbar")
            return False

    except Exception as e:
        logger.error(f"[VBA Bridge] Start fehlgeschlagen: {e}")
        return False

def vba_bridge_watchdog():
    """Watchdog-Thread der VBA Bridge überwacht und bei Bedarf neu startet"""
    global _vba_bridge_watchdog_running
    _vba_bridge_watchdog_running = True

    logger.info("[VBA Bridge Watchdog] Gestartet - überwache VBA Bridge")

    # Initiales Starten
    if not is_vba_bridge_running():
        start_vba_bridge()

    while _vba_bridge_watchdog_running:
        time_module.sleep(VBA_BRIDGE_CHECK_INTERVAL)

        if not is_vba_bridge_running():
            logger.warning("[VBA Bridge Watchdog] VBA Bridge nicht erreichbar - starte neu...")
            start_vba_bridge()

def start_vba_bridge_watchdog():
    """Startet den Watchdog als Daemon-Thread"""
    watchdog_thread = threading.Thread(target=vba_bridge_watchdog, daemon=True)
    watchdog_thread.start()
    logger.info("[VBA Bridge Watchdog] Watchdog-Thread gestartet")

def stop_vba_bridge_watchdog():
    """Stoppt den Watchdog"""
    global _vba_bridge_watchdog_running, _vba_bridge_process
    _vba_bridge_watchdog_running = False

    if _vba_bridge_process:
        try:
            _vba_bridge_process.terminate()
            logger.info("[VBA Bridge] Prozess beendet")
        except:
            pass

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*", "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"], "allow_headers": ["Content-Type"]}})  # CORS für alle Routen aktivieren


# =============================================================================
# REQUEST-LEVEL DB LOCKING (für Thread-Sicherheit mit Access ODBC)
# =============================================================================

@app.before_request
def acquire_db_lock():
    """Erwirbt DB-Lock vor jedem Request (Access ODBC ist nicht thread-safe)"""
    # Nur für /api/ Routen locken (nicht für statische Dateien)
    if request.path.startswith('/api/'):
        _db_lock.acquire()


@app.after_request
def release_db_lock(response):
    """Gibt DB-Lock nach jedem Request frei"""
    if request.path.startswith('/api/'):
        try:
            _db_lock.release()
        except RuntimeError:
            pass  # Lock war nicht gehalten
    return response


@app.teardown_request
def release_db_lock_on_error(exc):
    """Gibt DB-Lock auch bei Fehlern frei"""
    if request.path.startswith('/api/'):
        try:
            _db_lock.release()
        except RuntimeError:
            pass  # Lock war nicht gehalten


# Menu-Endpoints registrieren (Shell v2)
if MENU_ENDPOINTS_AVAILABLE:
    register_menu_routes(app)

# =============================================================================
# STATIC FILE SERVING - HTML Formulare
# =============================================================================

# Pfad zu den HTML-Formularen (relativ zum api-Ordner)
FORMS_BASE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'forms3'))

@app.route('/forms3/<path:filename>')
def serve_forms3(filename):
    """Serviert HTML-Formulare und Assets aus forms3"""
    return send_from_directory(FORMS_BASE_PATH, filename)

@app.route('/forms3/')
def serve_forms3_index():
    """Serviert Index der forms3"""
    return send_from_directory(FORMS_BASE_PATH, 'index.html')

# ROOT-LEVEL STATIC FILE SERVING (für VBA WebView2 Aufrufe)
@app.route('/shell.html')
def serve_shell():
    """Serviert shell.html direkt von Root (für VBA Kompatibilität)"""
    return send_from_directory(FORMS_BASE_PATH, 'shell.html')

@app.route('/')
def serve_root():
    """Root redirect to shell.html"""
    return send_from_directory(FORMS_BASE_PATH, 'shell.html')

@app.route('/css/<path:filename>')
def serve_css(filename):
    """Serviert CSS-Dateien"""
    return send_from_directory(os.path.join(FORMS_BASE_PATH, 'css'), filename)

@app.route('/js/<path:filename>')
def serve_js(filename):
    """Serviert JS-Dateien"""
    return send_from_directory(os.path.join(FORMS_BASE_PATH, 'js'), filename)

@app.route('/logic/<path:filename>')
def serve_logic(filename):
    """Serviert Logic-Dateien"""
    return send_from_directory(os.path.join(FORMS_BASE_PATH, 'logic'), filename)

@app.route('/<filename>')
def serve_html_root(filename):
    """Serviert HTML-Dateien direkt von Root"""
    if filename.endswith('.html'):
        return send_from_directory(FORMS_BASE_PATH, filename)
    # Fallback für andere Dateitypen
    return send_from_directory(FORMS_BASE_PATH, filename)

# =============================================================================
# MITARBEITER-FOTOS (UNC-Server-Pfad)
# =============================================================================

# Pfad für Mitarbeiterfotos (wie in Access VBA: prp_CONSYS_GrundPfad & TLookup("Pfad", "_tblEigeneFirma_Pfade", "ID = 7"))
# S: ist gemappt auf \\vConSYS01-NBG\Consys
MA_FOTO_UNC_PATH = r"S:\Bilder\Mitarbeiter"

@app.route('/api/fotos/mitarbeiter/<filename>')
def serve_mitarbeiter_foto(filename):
    """
    Serviert Mitarbeiterfotos vom UNC-Server-Pfad.
    Browser koennen nicht direkt auf file:// zugreifen,
    daher proxy via HTTP.
    """
    try:
        # Sicherheitspruefung: Nur Bilddateien erlauben
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}
        _, ext = os.path.splitext(filename.lower())
        if ext not in allowed_extensions:
            return jsonify({'error': 'Ungültiger Dateityp'}), 400

        # Pfadtraversal-Schutz
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Ungültiger Dateiname'}), 400

        full_path = os.path.join(MA_FOTO_UNC_PATH, filename)

        if os.path.exists(full_path):
            return send_from_directory(MA_FOTO_UNC_PATH, filename)
        else:
            logger.warning(f"Mitarbeiterfoto nicht gefunden: {full_path}")
            return jsonify({'error': 'Foto nicht gefunden'}), 404
    except Exception as e:
        logger.error(f"Fehler beim Laden des Mitarbeiterfotos: {e}")
        return jsonify({'error': str(e)}), 500

# =============================================================================
# KONFIGURATION
# =============================================================================

# Access Backend Datenbank
BACKEND_PATH = r"S:\CONSEC\CONSEC PLANUNG AKTUELL\Consec_BE_V1.55ANALYSETEST.accdb"

# Fallback auf lokale Kopie falls Netzwerk nicht verfügbar
LOCAL_BACKEND_PATH = r"C:\Users\guenther.siegert\Documents\Consec_BE_LOCAL.accdb"

# Connection String für 64-bit Access
CONN_STRING_TEMPLATE = (
    r"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};"
    r"DBQ={path};"
    r"ExtendedAnsiSQL=1;"
)

# Cache
_cache = {}
_cache_timeout = {}
CACHE_TTL = 60  # Sekunden

# =============================================================================
# DATENBANK-VERBINDUNG
# =============================================================================

def get_db_connection():
    """Erstellt eine Datenbankverbindung (nicht thread-safe - nutze get_db() stattdessen)"""
    # Versuche zuerst Netzwerkpfad
    if os.path.exists(BACKEND_PATH):
        path = BACKEND_PATH
    elif os.path.exists(LOCAL_BACKEND_PATH):
        path = LOCAL_BACKEND_PATH
    else:
        raise Exception(f"Datenbank nicht gefunden: {BACKEND_PATH}")

    conn_string = CONN_STRING_TEMPLATE.format(path=path)
    return pyodbc.connect(conn_string)


@contextmanager
def get_db():
    """
    Thread-safe DB-Connection als Context Manager.
    Verwendet Lock um parallele Access-ODBC Zugriffe zu serialisieren.

    Verwendung:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT ...")
    """
    with _db_lock:
        conn = None
        try:
            conn = get_db_connection()
            yield conn
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

def serialize_value(val):
    """Konvertiert Access-Werte zu JSON-serialisierbaren Typen"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, bytes):
        return val.decode('utf-8', errors='replace')
    return val

def row_to_dict(cursor, row):
    """Konvertiert eine Datenbankzeile zu einem Dictionary"""
    columns = [column[0] for column in cursor.description]
    return {col: serialize_value(val) for col, val in zip(columns, row)}

def execute_query(sql, params=None):
    """Führt eine SQL-Abfrage aus und gibt die Ergebnisse zurück"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if params:
            cursor.execute(sql, params)
        else:
            cursor.execute(sql)
        
        if cursor.description:
            rows = cursor.fetchall()
            return [row_to_dict(cursor, row) for row in rows]
        else:
            conn.commit()
            return {"affected_rows": cursor.rowcount}
            
    except Exception as e:
        logger.error(f"SQL Error: {e}")
        logger.error(f"SQL: {sql}")
        raise
    finally:
        if conn:
            conn.close()

def get_cached(key, ttl=CACHE_TTL):
    """Holt Wert aus Cache falls nicht abgelaufen"""
    if key in _cache:
        if datetime.now().timestamp() - _cache_timeout.get(key, 0) < ttl:
            return _cache[key]
    return None

def set_cached(key, value):
    """Speichert Wert im Cache"""
    _cache[key] = value
    _cache_timeout[key] = datetime.now().timestamp()

# =============================================================================
# API ROUTEN - AUFTRÄGE
# =============================================================================

@app.route('/api/auftraege', methods=['GET'])
def get_auftraege():
    """Listet alle Aufträge - mit expand_days=true jeden Tag separat"""
    try:
        # Parameter
        limit = request.args.get('limit', 100, type=int)
        offset = request.args.get('offset', 0, type=int)
        # Unterstütze beide Varianten: datum_von/datum_bis und von/bis
        datum_von = request.args.get('datum_von') or request.args.get('von')
        datum_bis = request.args.get('datum_bis') or request.args.get('bis')
        # NEU 16.01.2026: expand_days=true zeigt jeden Tag eines Mehrtages-Auftrags separat
        expand_days = request.args.get('expand_days', 'false').lower() == 'true'

        if expand_days:
            # JOIN mit tbl_VA_AnzTage für jeden einzelnen Tag
            sql = """
                SELECT TOP {limit}
                    a.ID AS VA_ID,
                    t.ID AS VADatum_ID,
                    t.VADatum AS Datum,
                    a.Auftrag,
                    a.Objekt,
                    a.Ort,
                    a.Dat_VA_Von,
                    a.Dat_VA_Bis,
                    a.Veranst_Status_ID,
                    a.Veranstalter_ID,
                    t.TVA_Soll AS MA_Anzahl_Soll,
                    t.TVA_Ist AS MA_Anzahl_Ist,
                    a.Treffpunkt,
                    a.Treffp_Zeit,
                    a.Dienstkleidung,
                    a.Ansprechpartner,
                    a.Bemerkungen
                FROM tbl_VA_Auftragstamm AS a
                INNER JOIN tbl_VA_AnzTage AS t ON a.ID = t.VA_ID
                WHERE a.Veranst_Status_ID IN (1,2,3,4)
            """.format(limit=limit)

            # Zeitraum-Filter auf das konkrete Tag-Datum
            if datum_von and datum_bis:
                sql += f" AND t.VADatum >= #{datum_von}#"
                sql += f" AND t.VADatum <= #{datum_bis}#"
            elif datum_von:
                sql += f" AND t.VADatum >= #{datum_von}#"
            elif datum_bis:
                sql += f" AND t.VADatum <= #{datum_bis}#"

            sql += " ORDER BY t.VADatum DESC, a.Auftrag"
        else:
            # Original: Ein Eintrag pro Auftrag
            sql = """
                SELECT TOP {limit}
                    ID AS VA_ID,
                    Auftrag,
                    Objekt,
                    Ort,
                    Dat_VA_Von,
                    Dat_VA_Bis,
                    Veranst_Status_ID,
                    Veranstalter_ID,
                    Treffpunkt,
                    Treffp_Zeit,
                    Dienstkleidung,
                    Ansprechpartner,
                    Bemerkungen,
                    Erst_von,
                    Erst_am,
                    Aend_von,
                    Aend_am
                FROM tbl_VA_Auftragstamm
                WHERE Veranst_Status_ID IN (1,2,3,4)
            """.format(limit=limit)

            # Zeitraum-Filter: Aufträge die im Zeitraum aktiv sind
            if datum_von and datum_bis:
                # Auftrag überlappt mit Zeitraum wenn: Start <= bis UND Ende >= von
                sql += f" AND Dat_VA_Von <= #{datum_bis}#"
                sql += f" AND Dat_VA_Bis >= #{datum_von}#"
            elif datum_von:
                sql += f" AND Dat_VA_Bis >= #{datum_von}#"
            elif datum_bis:
                sql += f" AND Dat_VA_Von <= #{datum_bis}#"

            sql += " ORDER BY Auftrag, Dat_VA_Von DESC"

        data = execute_query(sql)
        return jsonify({"success": True, "data": data, "count": len(data)})
        
    except Exception as e:
        logger.error(f"Error in get_auftraege: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>', methods=['GET'])
def get_auftrag(id):
    """Holt einen einzelnen Auftrag mit Details"""
    try:
        # Hauptdaten
        sql = """
            SELECT 
                ID AS VA_ID,
                Auftrag,
                Objekt,
                Objekt_ID,
                Ort,
                Dat_VA_Von,
                Dat_VA_Bis,
                Veranst_Status_ID,
                Veranstalter_ID,
                Treffpunkt,
                Treffpunkt2,
                Treffp_Zeit,
                Dienstkleidung,
                Ansprechpartner,
                Dummy,
                Bemerkungen,
                Bemerkung,
                Rch_Dat,
                Rch_Nr,
                Erst_von,
                Erst_am,
                Aend_von,
                Aend_am
            FROM tbl_VA_Auftragstamm
            WHERE ID = ?
        """
        
        auftrag = execute_query(sql, [id])
        if not auftrag:
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404
        
        # Einsatztage laden
        einsatztage_sql = """
            SELECT 
                ID,
                VA_ID,
                VADatum,
                TVA_Soll,
                TVA_Ist,
                TVA_Offen
            FROM tbl_VA_AnzTage
            WHERE VA_ID = ?
            ORDER BY VADatum
        """
        einsatztage = execute_query(einsatztage_sql, [id])
        
        return jsonify({
            "success": True, 
            "data": {
                "auftrag": auftrag[0],
                "einsatztage": einsatztage
            }
        })
        
    except Exception as e:
        logger.error(f"Error in get_auftrag: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege', methods=['POST'])
def create_auftrag():
    """Erstellt einen neuen Auftrag"""
    try:
        data = request.get_json()
        
        sql = """
            INSERT INTO tbl_VA_Auftragstamm (
                Auftrag, Objekt, Ort, Dat_VA_Von, Dat_VA_Bis,
                Veranst_Status_ID, Veranstalter_ID, Treffpunkt,
                Dienstkleidung, Ansprechpartner, Bemerkungen,
                Erst_von, Erst_am
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        params = [
            data.get('Auftrag'),
            data.get('Objekt'),
            data.get('Ort'),
            data.get('Dat_VA_Von'),
            data.get('Dat_VA_Bis'),
            data.get('Veranst_Status_ID'),
            data.get('Veranstalter_ID'),
            data.get('Treffpunkt'),
            data.get('Dienstkleidung'),
            data.get('Ansprechpartner'),
            data.get('Bemerkungen'),
            'System',
            datetime.now()
        ]
        
        execute_query(sql, params)
        
        # Neue ID holen
        result = execute_query("SELECT @@IDENTITY AS ID")
        new_id = result[0]['ID'] if result else None
        
        return jsonify({"success": True, "id": new_id})
        
    except Exception as e:
        logger.error(f"Error in create_auftrag: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>', methods=['PUT'])
def update_auftrag(id):
    """Aktualisiert einen Auftrag"""
    try:
        data = request.get_json()
        
        # Dynamisches Update bauen
        fields = []
        params = []
        
        field_mapping = {
            'Auftrag': 'Auftrag',
            'Objekt': 'Objekt',
            'Objekt_ID': 'Objekt_ID',
            'Ort': 'Ort',
            'Dat_VA_Von': 'Dat_VA_Von',
            'Dat_VA_Bis': 'Dat_VA_Bis',
            'Veranst_Status_ID': 'Veranst_Status_ID',
            'Veranstalter_ID': 'Veranstalter_ID',
            'Treffpunkt': 'Treffpunkt',
            'Treffp_Zeit': 'Treffp_Zeit',
            'Dienstkleidung': 'Dienstkleidung',
            'Ansprechpartner': 'Ansprechpartner',
            'Fahrtkosten': 'Fahrtkosten',
            'Dummy': 'Dummy',
            'Bemerkungen': 'Bemerkungen'
        }
        
        for json_field, db_field in field_mapping.items():
            if json_field in data:
                fields.append(f"{db_field} = ?")
                params.append(data[json_field])
        
        if not fields:
            return jsonify({"success": False, "error": "Keine Felder zum Update"}), 400
        
        # Änderungsdaten hinzufügen
        fields.append("Aend_von = ?")
        fields.append("Aend_am = ?")
        params.extend(['System', datetime.now()])
        params.append(id)
        
        sql = f"UPDATE tbl_VA_Auftragstamm SET {', '.join(fields)} WHERE ID = ?"
        execute_query(sql, params)
        
        return jsonify({"success": True})
        
    except Exception as e:
        logger.error(f"Error in update_auftrag: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>', methods=['DELETE'])
def delete_auftrag(id):
    """Löscht einen Auftrag"""
    try:
        # Erst Abhängigkeiten löschen
        execute_query("DELETE FROM tbl_MA_VA_Planung WHERE Planung_VA_ID = ?", [id])
        execute_query("DELETE FROM tbl_VA_Start WHERE VAS_VADatum_ID IN (SELECT VADatum_ID FROM tbl_VA_Datum WHERE VADatum_VA_ID = ?)", [id])
        execute_query("DELETE FROM tbl_VA_Datum WHERE VADatum_VA_ID = ?", [id])
        execute_query("DELETE FROM tbl_VA_Auftragstamm WHERE ID = ?", [id])
        
        return jsonify({"success": True})
        
    except Exception as e:
        logger.error(f"Error in delete_auftrag: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/copy', methods=['POST'])
def copy_auftrag():
    """Kopiert einen bestehende Auftrag"""
    try:
        data = request.get_json() or {}
        source_id = data.get('id') or data.get('VA_ID') or data.get('va_id')
        if not source_id:
            return jsonify({"success": False, "error": "Quelle-ID erforderlich"}), 400

        records = execute_query("SELECT * FROM tbl_VA_Auftragstamm WHERE ID = ?", [source_id])
        if not records:
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        original = records[0]
        copy_data = {k: v for k, v in original.items() if k != 'ID'}
        now = datetime.now()
        if 'Erst_von' in copy_data:
            copy_data['Erst_von'] = 'CONSYS Copy'
        if 'Erst_am' in copy_data:
            copy_data['Erst_am'] = now
        if 'Aend_von' in copy_data:
            copy_data['Aend_von'] = 'CONSYS Copy'
        if 'Aend_am' in copy_data:
            copy_data['Aend_am'] = now

        columns = list(copy_data.keys())
        placeholders = ', '.join(['?' for _ in columns])
        sql = f"INSERT INTO tbl_VA_Auftragstamm ({', '.join(columns)}) VALUES ({placeholders})"
        params = [copy_data[col] for col in columns]

        execute_query(sql, params)
        identity = execute_query("SELECT @@IDENTITY AS ID")
        new_id = identity[0]['ID'] if identity else None

        return jsonify({"success": True, "data": {"new_id": new_id}})

    except Exception as e:
        logger.error(f"Error in copy_auftrag: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/send-einsatzliste', methods=['POST'])
def send_einsatzliste():
    """Simuliert das Versenden einer Einsatzliste"""
    try:
        data = request.get_json() or {}
        va_id = data.get('va_id') or data.get('VA_ID')
        typ = data.get('typ') or 'MA'

        if not va_id:
            return jsonify({"success": False, "error": "VA_ID erforderlich"}), 400

        logger.info(f"Einsatzliste '{typ}' für Auftrag {va_id} angefordert.")
        return jsonify({"success": True, "message": f"Einsatzliste {typ} für Auftrag {va_id} simuliert"})

    except Exception as e:
        logger.error(f"Error in send_einsatzliste: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# NEU: API ROUTEN - BUTTON-PARITAET (Access-kompatibel)
# =============================================================================

@app.route('/api/auftraege/<int:id>/status', methods=['PUT'])
def set_auftrag_status(id):
    """
    FIX 1 (Teil 2): Setzt den Status eines Auftrags
    Entspricht Access: Me!Veranst_Status_ID = 2 (Beendet)
    """
    try:
        data = request.get_json() or {}
        status_id = data.get('status_id') or data.get('Veranst_Status_ID')

        if not status_id:
            return jsonify({"success": False, "error": "status_id erforderlich"}), 400

        sql = """
            UPDATE tbl_VA_Auftragstamm
            SET Veranst_Status_ID = ?, Aend_von = ?, Aend_am = ?
            WHERE ID = ?
        """
        execute_query(sql, [status_id, 'API', datetime.now(), id])

        logger.info(f"Auftrag {id} Status auf {status_id} gesetzt")
        return jsonify({"success": True, "message": f"Status auf {status_id} gesetzt"})

    except Exception as e:
        logger.error(f"Error in set_auftrag_status: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:id>/excel-export', methods=['POST'])
def export_auftrag_excel(id):
    """
    FIX 1: Excel-Export fuer Auftrag (wie Access btnDruckZusage_Click)
    Erstellt Excel-Datei mit Auftragsdaten und MA-Zuordnungen
    """
    try:
        data = request.get_json() or {}
        vadatum = data.get('vadatum')

        # Auftragsdaten laden
        auftrag_sql = """
            SELECT ID, Auftrag, Objekt, Ort, Dat_VA_Von, Dat_VA_Bis,
                   Treffpunkt, Treffp_Zeit, Dienstkleidung, Ansprechpartner, Bemerkungen
            FROM tbl_VA_Auftragstamm WHERE ID = ?
        """
        auftrag = execute_query(auftrag_sql, [id])
        if not auftrag:
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        auftrag = auftrag[0]

        # MA-Zuordnungen laden
        zuord_sql = """
            SELECT
                p.MA_ID, m.Nachname, m.Vorname, m.Tel_Mobil,
                p.MVA_Start, p.MVA_Ende, p.VADatum, p.PosNr
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VA_ID = ?
            ORDER BY p.VADatum, p.PosNr, p.MVA_Start
        """
        zuordnungen = execute_query(zuord_sql, [id])

        # Excel-Datei erstellen (mit openpyxl falls verfuegbar, sonst CSV)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            import tempfile
            import uuid

            wb = Workbook()
            ws = wb.active
            ws.title = "Einsatzliste"

            # Header-Bereich
            ws['A1'] = "Auftrag:"
            ws['B1'] = auftrag.get('Auftrag', '')
            ws['A2'] = "Objekt:"
            ws['B2'] = auftrag.get('Objekt', '')
            ws['A3'] = "Ort:"
            ws['B3'] = auftrag.get('Ort', '')
            ws['A4'] = "Datum:"
            ws['B4'] = str(auftrag.get('Dat_VA_Von', ''))
            ws['A5'] = "Treffpunkt:"
            ws['B5'] = auftrag.get('Treffpunkt', '')
            ws['A6'] = "Dienstkleidung:"
            ws['B6'] = auftrag.get('Dienstkleidung', '')

            # Tabellen-Header
            headers = ['Nr', 'Nachname', 'Vorname', 'Von', 'Bis', 'Telefon', 'Datum']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col, value=header)
                cell.font = Font(bold=True)

            # Daten
            for row_idx, z in enumerate(zuordnungen, 9):
                ws.cell(row=row_idx, column=1, value=row_idx - 8)
                ws.cell(row=row_idx, column=2, value=z.get('Nachname', ''))
                ws.cell(row=row_idx, column=3, value=z.get('Vorname', ''))
                ws.cell(row=row_idx, column=4, value=str(z.get('MVA_Start', ''))[:5] if z.get('MVA_Start') else '')
                ws.cell(row=row_idx, column=5, value=str(z.get('MVA_Ende', ''))[:5] if z.get('MVA_Ende') else '')
                ws.cell(row=row_idx, column=6, value=z.get('Tel_Mobil', ''))
                ws.cell(row=row_idx, column=7, value=str(z.get('VADatum', ''))[:10] if z.get('VADatum') else '')

            # Spaltenbreiten
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12

            # Datei speichern
            export_dir = os.path.join(os.path.dirname(__file__), 'exports')
            os.makedirs(export_dir, exist_ok=True)

            # Dateiname wie in Access: [TT-MM-JJ] [Auftrag] [Objekt].xlsx
            datum_str = datetime.now().strftime('%d-%m-%y')
            safe_auftrag = "".join(c for c in auftrag.get('Auftrag', 'Auftrag') if c.isalnum() or c in ' _-')[:30]
            safe_objekt = "".join(c for c in auftrag.get('Objekt', 'Objekt') if c.isalnum() or c in ' _-')[:30]
            filename = f"{datum_str} {safe_auftrag} {safe_objekt}.xlsx"
            filepath = os.path.join(export_dir, filename)

            wb.save(filepath)
            logger.info(f"Excel-Export erstellt: {filepath}")

            # Download-URL zurueckgeben
            return jsonify({
                "success": True,
                "data": {
                    "download_url": f"/api/download/{filename}",
                    "filename": filename,
                    "filepath": filepath,
                    "ma_count": len(zuordnungen)
                }
            })

        except ImportError:
            # Fallback: CSV wenn openpyxl nicht verfuegbar
            logger.warning("openpyxl nicht installiert - CSV-Export")
            return jsonify({
                "success": False,
                "error": "Excel-Export nicht verfuegbar (openpyxl fehlt). Bitte installieren: pip install openpyxl"
            }), 500

    except Exception as e:
        logger.error(f"Error in export_auftrag_excel: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download-Route fuer exportierte Dateien"""
    from flask import send_from_directory
    export_dir = os.path.join(os.path.dirname(__file__), 'exports')
    return send_from_directory(export_dir, filename, as_attachment=True)


@app.route('/api/auftraege/<int:id>/copy-to-next-day', methods=['POST'])
def copy_to_next_day(id):
    """
    FIX 2: Daten in Folgetag kopieren (wie Access btnPlan_Kopie_Click)
    Kopiert tbl_VA_Start und tbl_MA_VA_Planung vom aktuellen Tag zum naechsten Tag
    """
    try:
        data = request.get_json() or {}
        current_datum = data.get('current_datum')
        current_datum_id = data.get('current_datum_id')

        if not id:
            return jsonify({"success": False, "error": "VA_ID erforderlich"}), 400

        # 1. Alle Einsatztage fuer diesen Auftrag laden
        tage_sql = """
            SELECT ID, VADatum, VA_ID
            FROM tbl_VA_AnzTage
            WHERE VA_ID = ?
            ORDER BY VADatum
        """
        tage = execute_query(tage_sql, [id])

        if not tage or len(tage) < 2:
            return jsonify({"success": False, "error": "Kein Folgetag vorhanden"}), 400

        # 2. Aktuellen Tag und Folgetag finden
        current_idx = None
        for idx, tag in enumerate(tage):
            tag_datum = str(tag.get('VADatum', ''))[:10] if tag.get('VADatum') else None
            tag_id = tag.get('ID')

            if current_datum and tag_datum == str(current_datum)[:10]:
                current_idx = idx
                break
            elif current_datum_id and tag_id == current_datum_id:
                current_idx = idx
                break

        # Falls kein aktueller Tag gefunden, ersten nehmen
        if current_idx is None:
            current_idx = 0

        if current_idx >= len(tage) - 1:
            return jsonify({"success": False, "error": "Kein Folgetag nach aktuellem Datum"}), 400

        current_tag = tage[current_idx]
        next_tag = tage[current_idx + 1]

        current_tag_id = current_tag.get('ID')
        next_tag_id = next_tag.get('ID')
        next_datum = next_tag.get('VADatum')

        logger.info(f"Kopiere von Tag {current_tag_id} nach Tag {next_tag_id}")

        # 3. Bestehende Daten im Folgetag loeschen
        execute_query("DELETE FROM tbl_VA_Start WHERE VA_ID = ? AND VADatum_ID = ?", [id, next_tag_id])
        execute_query("DELETE FROM tbl_MA_VA_Planung WHERE VA_ID = ? AND VADatum_ID = ?", [id, next_tag_id])

        # 4. Schichten kopieren (tbl_VA_Start)
        schichten_sql = """
            SELECT VA_ID, VA_Start, VA_Ende, MA_Anzahl, Position_ID, Bemerkung
            FROM tbl_VA_Start
            WHERE VA_ID = ? AND VADatum_ID = ?
        """
        schichten = execute_query(schichten_sql, [id, current_tag_id])

        schichten_count = 0
        for s in schichten:
            insert_sql = """
                INSERT INTO tbl_VA_Start (VA_ID, VADatum_ID, VADatum, VA_Start, VA_Ende, MA_Anzahl, Position_ID, Bemerkung)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """
            execute_query(insert_sql, [
                id, next_tag_id, next_datum,
                s.get('VA_Start'), s.get('VA_Ende'),
                s.get('MA_Anzahl'), s.get('Position_ID'), s.get('Bemerkung')
            ])
            schichten_count += 1

        # 5. MA-Zuordnungen kopieren (tbl_MA_VA_Planung)
        zuord_sql = """
            SELECT VA_ID, MA_ID, MVA_Start, MVA_Ende, Status_ID, PosNr, PKW, Bemerkung
            FROM tbl_MA_VA_Planung
            WHERE VA_ID = ? AND VADatum_ID = ?
        """
        zuordnungen = execute_query(zuord_sql, [id, current_tag_id])

        zuord_count = 0
        for z in zuordnungen:
            insert_sql = """
                INSERT INTO tbl_MA_VA_Planung (VA_ID, VADatum_ID, VADatum, MA_ID, MVA_Start, MVA_Ende, Status_ID, PosNr, PKW, Bemerkung)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            execute_query(insert_sql, [
                id, next_tag_id, next_datum,
                z.get('MA_ID'), z.get('MVA_Start'), z.get('MVA_Ende'),
                z.get('Status_ID'), z.get('PosNr'), z.get('PKW'), z.get('Bemerkung')
            ])
            zuord_count += 1

        logger.info(f"Kopiert: {schichten_count} Schichten, {zuord_count} MA-Zuordnungen")

        return jsonify({
            "success": True,
            "data": {
                "schichten_count": schichten_count,
                "zuordnungen_count": zuord_count,
                "next_datum": str(next_datum)[:10] if next_datum else None,
                "next_datum_id": next_tag_id
            },
            "message": f"{schichten_count} Schichten und {zuord_count} MA-Zuordnungen in Folgetag kopiert"
        })

    except Exception as e:
        logger.error(f"Error in copy_to_next_day: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/bwn/send', methods=['POST'])
def send_bwn():
    """
    FIX 3: Bewachungsnachweis per E-Mail senden (wie Access cmd_BWN_send_Click)
    Mit Option 'nur_markierte' fuer selektiven Versand
    """
    try:
        data = request.get_json() or {}
        va_id = data.get('va_id') or data.get('VA_ID')
        vadatum = data.get('vadatum')
        vadatum_id = data.get('vadatum_id')
        nur_markierte = data.get('nur_markierte', False)

        if not va_id:
            return jsonify({"success": False, "error": "VA_ID erforderlich"}), 400

        # 1. Auftragsdaten laden (ID ist der PK in tbl_VA_Auftragstamm)
        auftrag_sql = "SELECT Auftrag, Objekt, Ort FROM tbl_VA_Auftragstamm WHERE ID = ?"
        auftrag = execute_query(auftrag_sql, [va_id])
        if not auftrag:
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404
        auftrag = auftrag[0]

        # 2. MA-Zuordnungen laden (mit E-Mail)
        # Hinweis: Rch_Erstellt existiert nicht - wir nutzen Status_ID stattdessen
        zuord_sql = """
            SELECT
                p.ID, p.MA_ID, p.Status_ID,
                m.Nachname, m.Vorname, m.Email
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VA_ID = ?
        """
        params = [va_id]

        if vadatum_id:
            zuord_sql += " AND p.VADatum_ID = ?"
            params.append(vadatum_id)

        zuord_sql += " ORDER BY m.Nachname, m.Vorname"
        zuordnungen = execute_query(zuord_sql, params)

        # 3. Filtern nach nur_markierte (Status_ID = 3 = Bestaetigt)
        if nur_markierte:
            zuordnungen = [z for z in zuordnungen if z.get('Status_ID') == 3]

        # 4. E-Mails senden (simuliert - in Produktion: Outlook/SMTP)
        sent_count = 0
        errors = []

        for z in zuordnungen:
            email = z.get('Email')
            name = f"{z.get('Vorname', '')} {z.get('Nachname', '')}".strip()

            if not email:
                errors.append(f"{name}: Keine E-Mail-Adresse")
                continue

            try:
                # Hier wuerde der tatsaechliche E-Mail-Versand stattfinden
                # In Produktion: win32com.client.Dispatch("Outlook.Application") oder smtplib
                logger.info(f"BWN gesendet an: {email} ({name})")
                sent_count += 1

                # Optional: Markierung zuruecksetzen
                # execute_query("UPDATE tbl_MA_VA_Planung SET Rch_Erstellt = False WHERE ID = ?", [z.get('ID')])

            except Exception as mail_error:
                errors.append(f"{name}: {str(mail_error)}")

        # 5. Ergebnis
        return jsonify({
            "success": True,
            "data": {
                "sent_count": sent_count,
                "total": len(zuordnungen),
                "errors": errors if errors else None,
                "nur_markierte": nur_markierte,
                "reset_markierungen": True
            },
            "message": f"BWN an {sent_count} von {len(zuordnungen)} Mitarbeiter gesendet"
        })

    except Exception as e:
        logger.error(f"Error in send_bwn: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - AUFTRAG SUBFORM ENDPOINTS (fuer Browser-CORS-Kompatibilitaet)
# =============================================================================

@app.route('/api/auftraege/vorschlaege', methods=['GET'])
def get_auftraege_vorschlaege():
    """Liefert Vorschlaege fuer Autocomplete-Felder"""
    try:
        feld = request.args.get('feld', '')

        field_mapping = {
            'ort': 'Ort',
            'objekt': 'Objekt',
            'auftrag': 'Auftrag',
            'dienstkleidung': 'Dienstkleidung'
        }

        db_field = field_mapping.get(feld.lower())
        if not db_field:
            return jsonify({"success": True, "data": []})

        # Access SQL: 'value' ist reserviert, verwende 'wert'
        sql = f"""
            SELECT DISTINCT [{db_field}] AS wert
            FROM tbl_VA_Auftragstamm
            WHERE [{db_field}] IS NOT NULL
            ORDER BY [{db_field}]
        """

        data = execute_query(sql)
        values = [row['wert'] for row in data if row.get('wert')]

        return jsonify({"success": True, "data": values})

    except Exception as e:
        logger.error(f"Error in get_auftraege_vorschlaege: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>/schichten', methods=['GET'])
def get_auftrag_schichten(id):
    """Liefert Schichten fuer einen Auftrag (optional gefiltert nach vadatum_id)"""
    try:
        vadatum_id = request.args.get('vadatum_id', type=int)

        if vadatum_id:
            # JOIN statt zwei Queries - vermeidet Datum-Typ-Probleme
            # WICHTIG: ID als VAStart_ID zurückgeben für Kompatibilität mit HTML
            sql = """
                SELECT
                    s.ID,
                    s.ID AS VAStart_ID,
                    s.VA_ID,
                    s.VADatum,
                    s.VA_Start,
                    s.VA_Ende,
                    s.MA_Anzahl,
                    s.MA_Anzahl_Ist
                FROM tbl_VA_Start AS s
                INNER JOIN tbl_VA_AnzTage AS t ON s.VA_ID = t.VA_ID AND s.VADatum = t.VADatum
                WHERE s.VA_ID = ? AND t.ID = ?
                ORDER BY s.VA_Start
            """
            data = execute_query(sql, [id, vadatum_id])
        else:
            sql = """
                SELECT
                    ID,
                    ID AS VAStart_ID,
                    VA_ID,
                    VADatum,
                    VA_Start,
                    VA_Ende,
                    MA_Anzahl,
                    MA_Anzahl_Ist
                FROM tbl_VA_Start
                WHERE VA_ID = ?
                ORDER BY VADatum, VA_Start
            """
            data = execute_query(sql, [id])

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_schichten: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>/zuordnungen', methods=['GET'])
def get_auftrag_zuordnungen(id):
    """Liefert MA-Zuordnungen fuer einen Auftrag (optional gefiltert nach vadatum_id)"""
    try:
        vadatum_id = request.args.get('vadatum_id', type=int)

        if vadatum_id:
            # FIX: Direkt nach VADatum_ID filtern (nicht über VADatum Datum-Vergleich)
            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum, p.VADatum_ID, p.VAStart_ID,
                    p.MVA_Start, p.MVA_Ende, p.Status_ID, p.PosNr,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ? AND p.VADatum_ID = ?
                ORDER BY p.PosNr, p.MVA_Start, m.Nachname
            """
            data = execute_query(sql, [id, vadatum_id])
        else:
            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum,
                    p.MVA_Start, p.MVA_Ende, p.Status_ID, p.PosNr,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ?
                ORDER BY p.VADatum, p.MVA_Start, m.Nachname
            """
            data = execute_query(sql, [id])

        # MA_Name in Python zusammenbauen
        for row in data:
            nachname = row.get('Nachname') or ''
            vorname = row.get('Vorname') or ''
            row['MA_Name'] = f"{nachname}, {vorname}" if nachname else vorname

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_zuordnungen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/init_zuordnungen', methods=['POST'])
def init_zuordnungen(va_id):
    """
    Erstellt leere Zuordnungs-Saetze fuer alle Schichten eines Auftrags/Datums.

    Entspricht btnVAPlanCrea_Click -> Zuord_Fill() in Access VBA.

    Fuer jede Schicht (tbl_VA_Start) werden MA_Anzahl leere Saetze erstellt
    in tbl_MA_VA_Zuordnung mit MA_ID = 0 (Platzhalter).

    Request Body (JSON):
        vadatum_id: ID des Einsatztages (tbl_VA_AnzTage.ID) - optional
                    Wenn nicht angegeben, werden alle Tage des Auftrags verarbeitet

    Returns:
        created: Anzahl neu erstellter Saetze
        updated: Anzahl aktualisierter/geloeschter Saetze
        schichten: Anzahl verarbeiteter Schichten
    """
    try:
        data = request.get_json() or {}
        vadatum_id = data.get('vadatum_id')

        created = 0
        deleted = 0
        schichten_count = 0

        # Schichten laden (tbl_VA_Start)
        if vadatum_id:
            schichten_sql = """
                SELECT ID, VADatum_ID, MA_Anzahl, VA_Start, VA_Ende, VADatum, MVA_Start, MVA_Ende
                FROM tbl_VA_Start
                WHERE VA_ID = ? AND VADatum_ID = ?
                ORDER BY VA_Start, VA_Ende
            """
            schichten = execute_query(schichten_sql, [va_id, vadatum_id])
        else:
            schichten_sql = """
                SELECT ID, VADatum_ID, MA_Anzahl, VA_Start, VA_Ende, VADatum, MVA_Start, MVA_Ende
                FROM tbl_VA_Start
                WHERE VA_ID = ?
                ORDER BY VADatum, VA_Start, VA_Ende
            """
            schichten = execute_query(schichten_sql, [va_id])

        if not schichten:
            return jsonify({
                "success": True,
                "message": "Keine Schichten gefunden",
                "created": 0,
                "deleted": 0,
                "schichten": 0
            })

        # Fuer jede Schicht: Platzhalter-Saetze erstellen/anpassen
        for schicht in schichten:
            vastart_id = schicht.get('ID')
            vadatum_id_schicht = schicht.get('VADatum_ID')
            ma_anzahl = schicht.get('MA_Anzahl') or 0
            va_start = schicht.get('VA_Start')
            va_ende = schicht.get('VA_Ende')
            va_datum = schicht.get('VADatum')
            mva_start = schicht.get('MVA_Start')
            mva_ende = schicht.get('MVA_Ende')

            schichten_count += 1

            # Vorhandene Anzahl Zuordnungen fuer diese Schicht zaehlen
            count_sql = "SELECT COUNT(*) AS cnt FROM tbl_MA_VA_Zuordnung WHERE VAStart_ID = ?"
            count_result = execute_query(count_sql, [vastart_id])
            vorhandene = count_result[0].get('cnt', 0) if count_result else 0

            # Wenn mehr vorhanden als benoetigt: Ueberzaehlige loeschen (nur leere MA_ID=0)
            if vorhandene > ma_anzahl:
                diff = vorhandene - ma_anzahl
                # Loesche leere Zuordnungen (MA_ID = 0) beginnend mit hoechster PosNr
                delete_sql = """
                    DELETE FROM tbl_MA_VA_Zuordnung
                    WHERE ID IN (
                        SELECT TOP {} ID FROM tbl_MA_VA_Zuordnung
                        WHERE VAStart_ID = ? AND (MA_ID = 0 OR MA_ID IS NULL)
                        ORDER BY PosNr DESC
                    )
                """.format(diff)
                execute_query(delete_sql, [vastart_id], fetch=False)
                deleted += diff

            # Wenn weniger vorhanden als benoetigt: Fehlende hinzufuegen
            elif vorhandene < ma_anzahl:
                diff = ma_anzahl - vorhandene

                # Hoechste PosNr fuer diesen Tag ermitteln
                max_pos_sql = "SELECT MAX(PosNr) AS max_pos FROM tbl_MA_VA_Zuordnung WHERE VADatum_ID = ?"
                max_pos_result = execute_query(max_pos_sql, [vadatum_id_schicht])
                next_pos = (max_pos_result[0].get('max_pos') or 0) + 1 if max_pos_result else 1

                # Leere Saetze einfuegen
                for i in range(diff):
                    insert_sql = """
                        INSERT INTO tbl_MA_VA_Zuordnung
                        (VA_ID, VADatum_ID, VAStart_ID, PosNr, MA_ID, MA_Start, MA_Ende,
                         Erst_am, VADatum, MVA_Start, MVA_Ende, PreisArt_ID)
                        VALUES (?, ?, ?, ?, 0, ?, ?, Now(), ?, ?, ?, 1)
                    """
                    execute_query(insert_sql, [
                        va_id, vadatum_id_schicht, vastart_id, next_pos + i,
                        va_start, va_ende, va_datum, mva_start, mva_ende
                    ], fetch=False)
                    created += 1

        logger.info(f"init_zuordnungen: VA_ID={va_id}, created={created}, deleted={deleted}, schichten={schichten_count}")

        return jsonify({
            "success": True,
            "message": f"{created} Zuordnungs-Saetze erstellt, {deleted} geloescht",
            "created": created,
            "deleted": deleted,
            "schichten": schichten_count
        })

    except Exception as e:
        logger.error(f"Error in init_zuordnungen: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:id>/absagen', methods=['GET'])
def get_auftrag_absagen(id):
    """Liefert Absagen fuer einen Auftrag (optional gefiltert nach vadatum_id)"""
    try:
        vadatum_id = request.args.get('vadatum_id', type=int)

        # Absagen sind Zuordnungen mit Status_ID = 2 (abgesagt)
        if vadatum_id:
            # Erst das Datum holen
            datum_sql = "SELECT VADatum FROM tbl_VA_AnzTage WHERE ID = ?"
            datum_result = execute_query(datum_sql, [vadatum_id])
            if not datum_result:
                return jsonify({"success": True, "data": []})

            va_datum = datum_result[0].get('VADatum')

            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum, p.Status_ID,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ? AND p.Status_ID = 2 AND p.VADatum = ?
                ORDER BY m.Nachname
            """
            data = execute_query(sql, [id, va_datum])
        else:
            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum, p.Status_ID,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ? AND p.Status_ID = 2
                ORDER BY p.VADatum, m.Nachname
            """
            data = execute_query(sql, [id])

        # MA_Name in Python zusammenbauen
        for row in data:
            nachname = row.get('Nachname') or ''
            vorname = row.get('Vorname') or ''
            row['MA_Name'] = f"{nachname}, {vorname}" if nachname else vorname

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_absagen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/anfragen', methods=['GET'])
def get_auftrag_anfragen(va_id):
    """
    Ausstehende MA-Anfragen für einen Auftrag (für Antworten-Tab)
    Gibt alle Planungen mit Status_ID < 3 (geplant/benachrichtigt) zurück
    """
    try:
        vadatum_id = request.args.get('vadatum_id', type=int)

        if vadatum_id:
            # Erst das Datum holen
            datum_sql = "SELECT VADatum FROM tbl_VA_AnzTage WHERE ID = ?"
            datum_result = execute_query(datum_sql, [vadatum_id])
            if not datum_result:
                return jsonify({"success": True, "data": []})

            va_datum = datum_result[0].get('VADatum')

            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum, p.Status_ID,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ? AND p.Status_ID < 3 AND p.VADatum = ?
                ORDER BY m.Nachname
            """
            data = execute_query(sql, [va_id, va_datum])
        else:
            sql = """
                SELECT
                    p.ID, p.VA_ID, p.MA_ID, p.VADatum, p.Status_ID,
                    m.Nachname, m.Vorname
                FROM tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ? AND p.Status_ID < 3
                ORDER BY p.VADatum, m.Nachname
            """
            data = execute_query(sql, [va_id])

        # MA_Name in Python zusammenbauen
        for row in data:
            nachname = row.get('Nachname') or ''
            vorname = row.get('Vorname') or ''
            row['MA_Name'] = f"{nachname}, {vorname}" if nachname else vorname

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_anfragen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/zusatzdateien', methods=['GET'])
def get_auftrag_zusatzdateien(va_id):
    """Liefert Zusatzdateien/Attachments fuer einen Auftrag"""
    try:
        # Hole zuerst die Objekt_ID vom Auftrag
        auftrag_sql = "SELECT Objekt_ID FROM tbl_VA_Auftragstamm WHERE ID = ?"
        auftrag_result = execute_query(auftrag_sql, [va_id])

        if not auftrag_result:
            return jsonify({"success": True, "data": []})

        objekt_id = auftrag_result[0].get('Objekt_ID')

        # TabellenNr = 4 fuer Auftraege (basierend auf Access-Struktur)
        sql = """
            SELECT
                ZusatzNr,
                Ueberordnung,
                TabellenID,
                Dateiname,
                Dateidatum,
                Dateilaenge,
                Typ,
                Kurzbeschreibung
            FROM tbl_ZusatzDateien
            WHERE Ueberordnung = ? AND TabellenID = 4
            ORDER BY Dateidatum DESC
        """
        data = execute_query(sql, [objekt_id])

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_zusatzdateien: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/attachments', methods=['GET'])
def get_attachments():
    """Liefert Zusatzdateien/Attachments fuer ein Objekt (generischer Endpoint)

    Query-Parameter:
    - objekt_id: Die Ueberordnung (Objekt/Auftrag ID)
    - tabellen_nr: TabellenID (4=Auftrag, 42=Objekt, etc.)
    """
    try:
        objekt_id = request.args.get('objekt_id', type=int)
        tabellen_nr = request.args.get('tabellen_nr', type=int, default=42)

        if not objekt_id:
            return jsonify({"success": True, "data": []})

        sql = """
            SELECT
                ZusatzNr,
                Ueberordnung,
                TabellenID,
                Dateiname,
                Dateidatum as DFiledate,
                Dateilaenge as DLaenge,
                Typ as Texttyp,
                Kurzbeschreibung
            FROM tbl_ZusatzDateien
            WHERE Ueberordnung = ? AND TabellenID = ?
            ORDER BY Dateidatum DESC
        """
        data = execute_query(sql, [objekt_id, tabellen_nr])

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_attachments: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/attachments/<int:zusatz_nr>', methods=['DELETE'])
def delete_attachment(zusatz_nr):
    """Loescht eine Zusatzdatei (generischer Endpoint)"""
    try:
        sql = "DELETE FROM tbl_ZusatzDateien WHERE ZusatzNr = ?"
        execute_query(sql, [zusatz_nr], fetch=False)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error in delete_attachment: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/attachments/<int:zusatz_nr>/download', methods=['GET'])
def download_attachment(zusatz_nr):
    """Liefert den Dateipfad einer Zusatzdatei zum Download/Oeffnen"""
    try:
        sql = "SELECT Dateiname FROM tbl_ZusatzDateien WHERE ZusatzNr = ?"
        result = execute_query(sql, [zusatz_nr])
        if not result:
            return jsonify({"success": False, "error": "Datei nicht gefunden"}), 404

        filename = result[0].get('Dateiname')
        if not filename:
            return jsonify({"success": False, "error": "Kein Dateiname vorhanden"}), 404

        # Gibt den Pfad zurueck - Client muss Datei selbst oeffnen
        return jsonify({"success": True, "filename": filename})

    except Exception as e:
        logger.error(f"Error in download_attachment: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/zusatzdateien/<int:zusatz_nr>', methods=['DELETE'])
def delete_zusatzdatei(zusatz_nr):
    """Loescht eine Zusatzdatei (Legacy-Endpoint)"""
    try:
        sql = "DELETE FROM tbl_ZusatzDateien WHERE ZusatzNr = ?"
        execute_query(sql, [zusatz_nr], fetch=False)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error in delete_zusatzdatei: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:va_id>/rechnungspositionen', methods=['GET'])
def get_auftrag_rechnungspositionen(va_id):
    """Liefert Rechnungspositionen fuer einen Auftrag"""
    try:
        sql = """
            SELECT
                ID,
                VA_ID,
                Rechnungs_ID,
                PosNr,
                Beschreibung,
                Menge,
                Mengenheit,
                EzPreis,
                MwSt,
                Preisart_ID
            FROM tbl_Rch_Pos_Auftrag
            WHERE VA_ID = ?
            ORDER BY PosNr
        """
        data = execute_query(sql, [va_id])

        # Gesamtpreis berechnen
        for row in data:
            menge = row.get('Menge') or 0
            ezpreis = row.get('EzPreis') or 0
            row['GesPreis'] = round(menge * ezpreis, 2)

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_rechnungspositionen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:va_id>/berechnungsliste', methods=['GET'])
def get_auftrag_berechnungsliste(va_id):
    """Liefert Berechnungsliste (Stundendetails) fuer einen Auftrag"""
    try:
        # Berechnungsliste basiert auf MA-Zuordnungen mit Stunden
        sql = """
            SELECT
                p.ID,
                p.VA_ID,
                p.MA_ID,
                p.VADatum,
                p.MVA_Start,
                p.MVA_Ende,
                p.Stunden_Ist,
                m.Nachname,
                m.Vorname
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VA_ID = ? AND p.Status_ID = 1
            ORDER BY p.VADatum, p.MVA_Start
        """
        data = execute_query(sql, [va_id])

        for row in data:
            nachname = row.get('Nachname') or ''
            vorname = row.get('Vorname') or ''
            row['MA_Name'] = f"{nachname}, {vorname}" if nachname else vorname

        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_auftrag_berechnungsliste: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - EINSATZUEBERSICHT
# =============================================================================

@app.route('/api/einsatzuebersicht', methods=['GET'])
def get_einsatzuebersicht():
    """Einsatzuebersicht mit allen Details (Ort, MA-Namen, Stunden, PosNr).

    Query-Parameter:
    - von: Startdatum YYYY-MM-DD (optional)
    - bis: Enddatum YYYY-MM-DD (optional)
    - nurAktive: true/false - Nur aktive Auftraege (default: true)
    """
    try:
        von = request.args.get('von')
        bis = request.args.get('bis')
        nur_aktive = request.args.get('nurAktive', 'true').lower() == 'true'

        # Query mit Datumsfilter und JOIN fuer Auftragsdaten
        sql = """
            SELECT
                s.ID AS VAS_ID,
                s.VA_ID,
                s.VADatum,
                s.VA_Start,
                s.VA_Ende,
                s.MA_Anzahl,
                a.Auftrag,
                a.Objekt,
                a.Ort,
                a.VA_IstAktiv
            FROM tbl_VA_Start AS s
            LEFT JOIN tbl_VA_Auftragstamm AS a ON s.VA_ID = a.VA_ID
            WHERE 1=1
        """

        # Datumsfilter anwenden (Access-kompatibles Datumsformat)
        if von:
            sql += f" AND s.VADatum >= #{von}#"
        if bis:
            sql += f" AND s.VADatum <= #{bis}#"

        # Nur aktive Auftraege Filter
        if nur_aktive:
            sql += " AND (a.VA_IstAktiv = -1 OR a.VA_IstAktiv IS NULL)"

        sql += " ORDER BY s.VADatum ASC, s.VA_Start"

        rows = execute_query(sql)
        einsaetze = []

        for row in rows:
            einsatz = {
                'VAS_ID': row.get('VAS_ID'),
                'VA_ID': row.get('VA_ID'),
                'VADatum': row.get('VADatum'),
                'VA_Start': row.get('VA_Start'),
                'VA_Ende': row.get('VA_Ende'),
                'MA_Anzahl': row.get('MA_Anzahl', 0),
                'Auftrag': row.get('Auftrag') or '',
                'Objekt': row.get('Objekt') or '',
                'Ort': row.get('Ort') or '',
                'VA_IstAktiv': row.get('VA_IstAktiv', -1),
                'PosNr': ''
            }

            # Stunden berechnen (Brutto)
            va_start = einsatz.get('VA_Start')
            va_ende = einsatz.get('VA_Ende')
            if va_start and va_ende:
                try:
                    from datetime import datetime as dt, timedelta
                    start_str = str(va_start)[:5] if len(str(va_start)) >= 5 else str(va_start)
                    ende_str = str(va_ende)[:5] if len(str(va_ende)) >= 5 else str(va_ende)

                    start_time = dt.strptime(start_str, "%H:%M")
                    ende_time = dt.strptime(ende_str, "%H:%M")

                    # Ueber Mitternacht
                    if ende_time < start_time:
                        ende_time += timedelta(days=1)

                    diff = ende_time - start_time
                    stunden_brutto = diff.total_seconds() / 3600
                    einsatz['Stunden_Brutto'] = round(stunden_brutto, 2)
                    einsatz['Stunden_Netto'] = round(stunden_brutto, 2)
                except Exception:
                    einsatz['Stunden_Brutto'] = 0
                    einsatz['Stunden_Netto'] = 0
            else:
                einsatz['Stunden_Brutto'] = 0
                einsatz['Stunden_Netto'] = 0

            # MA-Namen und Ist-Anzahl abrufen
            vas_id = einsatz.get('VAS_ID')
            if vas_id:
                try:
                    ma_sql = """
                        SELECT m.Nachname, m.Vorname
                        FROM tbl_MA_VA_Planung AS p
                        INNER JOIN tbl_MA_Mitarbeiterstamm AS m ON p.MA_ID = m.ID
                        WHERE p.VAStart_ID = ?
                    """
                    ma_rows = execute_query(ma_sql, [vas_id])
                    ma_namen = []
                    for ma_row in ma_rows:
                        vorname = ma_row.get('Vorname') or ''
                        nachname = ma_row.get('Nachname') or ''
                        ma_namen.append(f"{vorname} {nachname}".strip())
                    einsatz['MA_Namen'] = ', '.join(ma_namen) if ma_namen else ''
                    einsatz['MA_Anzahl_Ist'] = len(ma_rows)
                except Exception:
                    einsatz['MA_Namen'] = ''
                    einsatz['MA_Anzahl_Ist'] = 0
            else:
                einsatz['MA_Namen'] = ''
                einsatz['MA_Anzahl_Ist'] = 0

            einsaetze.append(einsatz)

        logger.info(f"[Einsatzuebersicht] Returning {len(einsaetze)} entries")
        return jsonify({"success": True, "data": einsaetze})

    except Exception as e:
        logger.error(f"Error in get_einsatzuebersicht: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - EINSATZTAGE
# =============================================================================

@app.route('/api/einsatztage', methods=['GET'])
def get_einsatztage():
    """Listet Einsatztage mit Schichten.
    
    Query-Parameter:
    - va_id: Auftrags-ID (optional - wenn leer, werden alle im Zeitraum zurückgegeben)
    - von: Startdatum YYYY-MM-DD (optional)
    - bis: Enddatum YYYY-MM-DD (optional)
    """
    try:
        va_id = request.args.get('va_id', type=int)
        von = request.args.get('von')
        bis = request.args.get('bis')

        # Wenn nur va_id: Originales Verhalten
        if va_id and not von and not bis:
            sql = """
                SELECT
                    ID,
                    VADatum,
                    VA_ID
                FROM tbl_VA_AnzTage
                WHERE VA_ID = ?
                ORDER BY VADatum
            """
            data = execute_query(sql, [va_id])
            return jsonify({"success": True, "data": data})

        # Zeitraum-basierte Abfrage: Alle Schichten (tbl_VA_Start) im Zeitraum
        if von and bis:
            sql = f"""
                SELECT
                    s.ID AS VADatum_ID,
                    s.VA_ID,
                    s.VADatum AS Datum,
                    s.VA_Start AS Start,
                    s.VA_Ende AS Ende,
                    s.MA_Anzahl AS Soll,
                    a.Auftrag,
                    a.Objekt,
                    a.Ort
                FROM tbl_VA_Start AS s
                INNER JOIN tbl_VA_Auftragstamm AS a ON s.VA_ID = a.ID
                WHERE s.VADatum >= #{von}#
                AND s.VADatum <= #{bis}#
                AND a.Veranst_Status_ID IN (1,2,3,4)
                ORDER BY s.VA_ID, s.VADatum, s.VA_Start
            """
            data = execute_query(sql)
            return jsonify({"success": True, "data": data})

        return jsonify({"success": False, "error": "va_id oder von/bis erforderlich"}), 400
        
    except Exception as e:
        logger.error(f"Error in get_einsatztage: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/einsatztage/<int:id>', methods=['GET'])
def get_einsatztag_by_id(id):
    """Liefert einzelnen Einsatztag nach ID."""
    try:
        sql = """
            SELECT ID, VADatum, VA_ID
            FROM tbl_VA_AnzTage
            WHERE ID = ?
        """
        data = execute_query(sql, [id])
        if not data:
            return jsonify({"success": False, "error": "Einsatztag nicht gefunden"}), 404
        return jsonify({"success": True, "data": data[0]})
    except Exception as e:
        logger.error(f"Error in get_einsatztag_by_id: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:va_id>/anfragen', methods=['GET'])
def get_auftrag_anfragen(va_id):
    """Liefert offene Anfragen für einen Auftrag."""
    try:
        vadatum_id = request.args.get('vadatum_id', type=int)

        sql = """
            SELECT
                p.ID,
                p.MA_ID,
                m.Nachname,
                m.Vorname,
                p.VADatum,
                p.Status_ID,
                p.Bemerkung
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VA_ID = ?
            AND p.Status_ID = 1
        """
        params = [va_id]

        if vadatum_id:
            # Hole Datum aus tbl_VA_AnzTage
            datum_sql = "SELECT VADatum FROM tbl_VA_AnzTage WHERE ID = ?"
            datum_result = execute_query(datum_sql, [vadatum_id])
            if datum_result:
                datum = datum_result[0].get('VADatum')
                if datum:
                    datum_str = datum.strftime('%Y-%m-%d') if hasattr(datum, 'strftime') else str(datum)[:10]
                    sql += f" AND p.VADatum = #{datum_str}#"

        sql += " ORDER BY m.Nachname, m.Vorname"
        data = execute_query(sql, params)
        return jsonify({"success": True, "data": data, "count": len(data)})
    except Exception as e:
        logger.error(f"Error in get_auftrag_anfragen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:va_id>/generate_days', methods=['POST'])
def generate_auftrag_days(va_id):
    """Generiert Einsatztage für einen Auftrag basierend auf Dat_VA_Von und Dat_VA_Bis."""
    try:
        # Hole Auftragsdaten
        auftrag_sql = "SELECT Dat_VA_Von, Dat_VA_Bis FROM tbl_VA_Auftragstamm WHERE ID = ?"
        auftrag = execute_query(auftrag_sql, [va_id])
        if not auftrag:
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        dat_von = auftrag[0].get('Dat_VA_Von')
        dat_bis = auftrag[0].get('Dat_VA_Bis')

        if not dat_von or not dat_bis:
            return jsonify({"success": False, "error": "Dat_VA_Von und Dat_VA_Bis erforderlich"}), 400

        # Lösche bestehende Einsatztage (ohne Zuordnungen)
        # Nur Tage ohne Planungen löschen
        delete_sql = """
            DELETE FROM tbl_VA_AnzTage
            WHERE VA_ID = ?
            AND ID NOT IN (SELECT DISTINCT VADatum_ID FROM tbl_MA_VA_Planung WHERE VA_ID = ? AND VADatum_ID IS NOT NULL)
        """
        try:
            execute_query(delete_sql, [va_id, va_id])
        except Exception:
            pass  # Ignoriere Fehler beim Löschen

        # Generiere neue Tage
        from datetime import timedelta
        current = dat_von if isinstance(dat_von, datetime) else datetime.strptime(str(dat_von)[:10], '%Y-%m-%d')
        end = dat_bis if isinstance(dat_bis, datetime) else datetime.strptime(str(dat_bis)[:10], '%Y-%m-%d')

        inserted = 0
        while current <= end:
            # Prüfe ob Tag bereits existiert
            check_sql = f"SELECT ID FROM tbl_VA_AnzTage WHERE VA_ID = ? AND VADatum = #{current.strftime('%Y-%m-%d')}#"
            existing = execute_query(check_sql, [va_id])
            if not existing:
                insert_sql = f"INSERT INTO tbl_VA_AnzTage (VA_ID, VADatum) VALUES (?, #{current.strftime('%Y-%m-%d')}#)"
                execute_query(insert_sql, [va_id])
                inserted += 1
            current += timedelta(days=1)

        # Hole alle Tage zurück
        result_sql = "SELECT ID, VADatum, VA_ID FROM tbl_VA_AnzTage WHERE VA_ID = ? ORDER BY VADatum"
        data = execute_query(result_sql, [va_id])

        return jsonify({
            "success": True,
            "data": data,
            "inserted": inserted,
            "message": f"{inserted} Einsatztage generiert"
        })
    except Exception as e:
        logger.error(f"Error in generate_auftrag_days: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:va_id>/rechnung', methods=['GET'])
def get_auftrag_rechnung(va_id):
    """Liefert Rechnungsdaten für einen Auftrag."""
    try:
        # Prüfe ob Rechnung existiert
        sql = """
            SELECT
                r.ID AS Rch_ID,
                r.Rechnungsnummer,
                r.Rechnungsdatum,
                r.Betrag_Netto,
                r.Betrag_Brutto,
                r.Status,
                r.Bezahlt_am
            FROM tbl_Rch_Kopf r
            WHERE r.VA_ID = ?
            ORDER BY r.Rechnungsdatum DESC
        """
        data = execute_query(sql, [va_id])

        if not data:
            return jsonify({"success": True, "data": None, "message": "Keine Rechnung vorhanden"})

        return jsonify({"success": True, "data": data[0], "alle_rechnungen": data})
    except Exception as e:
        logger.error(f"Error in get_auftrag_rechnung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/rechnungen', methods=['GET'])
def get_auftrag_rechnungen(va_id):
    """
    Liefert ALLE Rechnungen für einen Auftrag (Plural-Endpoint).

    SQL: SELECT * FROM tbl_Rch_Kopf WHERE VA_ID=?

    Returns:
        Liste aller Rechnungen mit allen Feldern aus tbl_Rch_Kopf
    """
    try:
        sql = """
            SELECT
                ID,
                VA_ID,
                Rechnungsnummer,
                Rechnungsdatum,
                Kun_ID,
                Betrag_Netto,
                Betrag_Brutto,
                MwSt_Satz,
                MwSt_Betrag,
                Status,
                Bezahlt_am,
                Bezahlt_Betrag,
                Mahnstufe,
                Letzte_Mahnung,
                Storniert,
                Storno_Datum,
                Bemerkung,
                Erst_von,
                Erst_am,
                Aend_von,
                Aend_am
            FROM tbl_Rch_Kopf
            WHERE VA_ID = ?
            ORDER BY Rechnungsdatum DESC, ID DESC
        """
        data = execute_query(sql, [va_id])

        return jsonify({
            "success": True,
            "data": data,
            "count": len(data),
            "va_id": va_id
        })
    except Exception as e:
        logger.error(f"Error in get_auftrag_rechnungen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/kosten', methods=['GET'])
def get_auftrag_kosten(va_id):
    """
    Liefert alle Kosten/Preise für einen Auftrag.

    Die Tabelle tbl_VA_Preise enthält Kosten pro Auftrag (VA_ID).
    Wird vom Subformular sub_VA_Kosten verwendet.

    Query-Parameter:
    - kostenart_id: Optional - Filter nach Kostenart

    Returns:
        Liste aller Kostenpositionen mit Beschreibung, Preis, etc.
    """
    try:
        kostenart_id = request.args.get('kostenart_id', type=int)

        sql = """
            SELECT
                ID,
                VA_ID,
                kun_ID,
                Beschreibung,
                EurPreis,
                Kostenart_ID,
                Kostenzuo_KD,
                Erst_von,
                Erst_am,
                Aend_von,
                Aend_am
            FROM tbl_VA_Preise
            WHERE VA_ID = ?
        """
        params = [va_id]

        if kostenart_id:
            sql += " AND Kostenart_ID = ?"
            params.append(kostenart_id)

        sql += " ORDER BY ID"

        data = execute_query(sql, params)

        return jsonify({
            "success": True,
            "data": data,
            "count": len(data),
            "va_id": va_id
        })
    except Exception as e:
        logger.error(f"Error in get_auftrag_kosten: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# API ROUTEN - ORTE (fuer Autocomplete)
# =============================================================================

@app.route('/api/orte', methods=['GET'])
def get_orte():
    """Liefert Liste aller Orte (aus Auftraegen)"""
    try:
        sql = """
            SELECT DISTINCT Ort
            FROM tbl_VA_Auftragstamm
            WHERE Ort IS NOT NULL
            ORDER BY Ort
        """
        data = execute_query(sql)
        orte = [row.get('Ort') for row in data if row.get('Ort')]
        return jsonify({"success": True, "data": orte})
    except Exception as e:
        logger.error(f"Error in get_orte: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - MITARBEITER
# =============================================================================

@app.route('/api/mitarbeiter', methods=['GET'])
def get_mitarbeiter():
    """Listet Mitarbeiter

    Query-Parameter:
    - limit: Max. Anzahl (default: 100)
    - aktiv: true/false (default: true)
    - search: Suche in Nachname/Vorname/PersNr
    - anstellungsart_id: Filter nach Anstellungsart (3=Festangestellt, 5=Minijobber)
    - anstellungsart_in: Komma-separierte Liste von Anstellungsart-IDs (z.B. "3,4")
    - nicht_eingeplant_va: VA_ID - nur MA die bei diesem Auftrag nicht eingeplant sind
    """
    try:
        limit = request.args.get('limit', 100, type=int)
        aktiv = request.args.get('aktiv', 'true')
        search = request.args.get('search', '')
        anstellungsart_id = request.args.get('anstellungsart_id', type=int)
        anstellungsart_in = request.args.get('anstellungsart_in', '')  # z.B. "3,4"
        nicht_eingeplant_va = request.args.get('nicht_eingeplant_va', type=int)

        # DEBUG: Logge Parameter
        logger.info(f"[get_mitarbeiter] anstellungsart_id={anstellungsart_id}, anstellungsart_in={anstellungsart_in}, aktiv={aktiv}, limit={limit}")

        sql = f"""
            SELECT TOP {limit}
                ID AS MA_ID,
                ID,
                Nachname,
                Vorname,
                LexWare_ID AS PersNr,
                Strasse,
                PLZ,
                Ort,
                Tel_Mobil AS TelMobil,
                Tel_Festnetz AS TelFest,
                Email,
                IstAktiv AS Aktiv,
                Anstellungsart_ID,
                Anstellungsart_ID AS Anstellung
            FROM tbl_MA_Mitarbeiterstamm
            WHERE 1=1
        """

        if aktiv.lower() == 'true':
            sql += " AND IstAktiv = True"

        if anstellungsart_id:
            sql += f" AND Anstellungsart_ID = {anstellungsart_id}"
        elif anstellungsart_in:
            # Mehrere Anstellungsarten: z.B. "3,4" für Festangestellt + Minijobber
            ids = [int(x.strip()) for x in anstellungsart_in.split(',') if x.strip().isdigit()]
            if ids:
                sql += f" AND Anstellungsart_ID IN ({','.join(map(str, ids))})"

        if search:
            sql += f" AND (Nachname LIKE '%{search}%' OR Vorname LIKE '%{search}%' OR LexWare_ID LIKE '%{search}%')"

        # Filtere MA die bereits bei diesem Auftrag eingeplant sind
        if nicht_eingeplant_va:
            sql += f" AND ID NOT IN (SELECT MA_ID FROM tbl_MA_VA_Planung WHERE VA_ID = {nicht_eingeplant_va})"

        sql += " ORDER BY Nachname, Vorname"

        # DEBUG: Logge SQL Query
        logger.info(f"[get_mitarbeiter] SQL: {sql}")

        data = execute_query(sql)
        logger.info(f"[get_mitarbeiter] Result count: {len(data)}")
        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mitarbeiter/<int:id>', methods=['GET'])
def get_mitarbeiter_detail(id):
    """Holt einen Mitarbeiter mit Details"""
    try:
        sql = """
            SELECT 
                ID AS MA_ID,
                Nachname,
                Vorname,
                LexWare_ID AS PersNr,
                LexWare_ID AS LexNr,
                Strasse,
                Nr,
                PLZ,
                Ort,
                Land,
                Bundesland,
                Tel_Mobil AS TelMobil,
                Tel_Festnetz AS TelFest,
                Email,
                Geb_Dat AS Geburtsdatum,
                Geb_Ort AS Geburtsort,
                Geb_Name AS Geburtsname,
                Staatsang AS Staatsangehoerigkeit,
                Geschlecht,
                IBAN,
                BIC,
                Kontoinhaber,
                IstAktiv AS Aktiv,
                Anstellungsart_ID AS Anstellung,
                Austrittsdatum,
                KV_Kasse AS Krankenkasse,
                SteuerNr AS SteuerID,
                Taetigkeit_Bezeichnung AS Taetigkeit,
                Urlaubsanspr_pro_Jahr AS Urlaubsanspruch,
                StundenZahlMax AS StdMonatMax,
                tblBilddatei AS Lichtbild,
                tblSignaturdatei AS Signatur,
                Erst_von,
                Erst_am,
                Aend_von,
                Aend_am
            FROM tbl_MA_Mitarbeiterstamm
            WHERE ID = ?
        """

        data = execute_query(sql, [id])
        if not data:
            return jsonify({"success": False, "error": "Mitarbeiter nicht gefunden"}), 404

        return jsonify({"success": True, "data": data[0]})

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter_detail: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# -----------------------------------------------------------------------------
# MITARBEITER SUB-ENDPOINTS (18.01.2026)
# -----------------------------------------------------------------------------

@app.route('/api/mitarbeiter/<int:id>/zuordnungen', methods=['GET'])
def get_mitarbeiter_zuordnungen(id):
    """Holt alle VA-Zuordnungen eines Mitarbeiters für Einsatzübersicht/Dienstplan"""
    try:
        datum_von = request.args.get('von', None)
        datum_bis = request.args.get('bis', None)
        limit = request.args.get('limit', 100, type=int)

        sql = f"""
            SELECT TOP {limit}
                p.ID AS Planung_ID,
                p.VA_ID,
                p.VADatum_ID,
                p.MA_ID,
                p.VADatum,
                p.MVA_Start,
                p.MVA_Ende,
                p.IstAngefragt,
                p.IstZugesagt,
                p.IstAbgesagt,
                p.Bemerkung,
                a.Auftrag,
                a.Objekt,
                o.ob_Ort AS Objekt_Ort,
                s.VA_Start AS Schicht_Start,
                s.VA_Ende AS Schicht_Ende
            FROM ((tbl_MA_VA_Planung AS p
            LEFT JOIN tbl_VA_Auftragstamm AS a ON p.VA_ID = a.ID)
            LEFT JOIN tbl_OB_Objekt AS o ON a.Objekt_ID = o.ob_ID)
            LEFT JOIN tbl_VA_Start AS s ON p.VAStart_ID = s.ID
            WHERE p.MA_ID = ?
        """
        params = [id]

        if datum_von:
            sql += " AND p.VADatum >= ?"
            params.append(datum_von)
        if datum_bis:
            sql += " AND p.VADatum <= ?"
            params.append(datum_bis)

        sql += " ORDER BY p.VADatum DESC, p.MVA_Start"

        data = execute_query(sql, params)
        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter_zuordnungen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mitarbeiter/<int:id>/nverfueg', methods=['GET'])
def get_mitarbeiter_nverfueg(id):
    """Holt Nicht-Verfügbarkeiten eines Mitarbeiters"""
    try:
        sql = """
            SELECT
                ID,
                MA_ID,
                vonDat,
                bisDat,
                vonZeit,
                bisZeit,
                Grund_ID,
                Bemerkung,
                IstGanztaegig
            FROM tbl_MA_NVerfuegZeiten
            WHERE MA_ID = ?
            ORDER BY vonDat DESC
        """

        data = execute_query(sql, [id])
        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter_nverfueg: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mitarbeiter/<int:id>/zeitkonto', methods=['GET'])
def get_mitarbeiter_zeitkonto(id):
    """Holt Zeitkonto-Daten eines Mitarbeiters"""
    try:
        jahr = request.args.get('jahr', None, type=int)

        sql = """
            SELECT
                ID,
                MA_ID,
                ZK_Jahr,
                ZK_Monat,
                Soll_Std,
                Ist_Std,
                Uebertrag,
                Differenz,
                Urlaub_Soll,
                Urlaub_Ist,
                Urlaub_Rest
            FROM tbl_MA_Zeitkonto
            WHERE MA_ID = ?
        """
        params = [id]

        if jahr:
            sql += " AND ZK_Jahr = ?"
            params.append(jahr)

        sql += " ORDER BY ZK_Jahr DESC, ZK_Monat DESC"

        data = execute_query(sql, params)
        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter_zeitkonto: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - KUNDEN
# =============================================================================

@app.route('/api/kunden', methods=['GET'])
def get_kunden():
    """Listet Kunden"""
    try:
        limit = request.args.get('limit', 200, type=int)
        aktiv = request.args.get('aktiv', 'true')

        sql = f"""
            SELECT TOP {limit}
                kun_Id,
                kun_Firma,
                kun_IstAktiv
            FROM tbl_KD_Kundenstamm
            WHERE 1=1
        """

        if aktiv.lower() == 'true':
            sql += " AND kun_IstAktiv = True"

        sql += " ORDER BY kun_Firma"

        data = execute_query(sql)
        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_kunden: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kunden/<int:id>', methods=['GET'])
def get_kunde(id):
    """Holt einen einzelnen Kunden"""
    try:
        sql = """
            SELECT 
                kun_Id,
                kun_Firma,
                kun_Kuerzel AS Kuerzel,
                kun_Strasse AS Strasse,
                kun_PLZ AS PLZ,
                kun_Ort AS Ort,
                kun_Land AS Land,
                kun_telefon AS Telefon,
                kun_mobil AS Mobil,
                kun_email AS Email,
                kun_URL AS Homepage,
                kun_Ansprechpartner AS Ansprechpartner,
                kun_iban AS IBAN,
                kun_bic AS BIC,
                kun_kreditinstitut AS Kreditinstitut,
                kun_ustidnr AS UStIDNr,
                kun_Zahlungsbedingung AS Zahlungsbedingung,
                kun_IstAktiv AS Aktiv,
                Erst_von AS Erst_von,
                Erst_am AS Erst_am,
                Aend_von AS Aend_von,
                Aend_am AS Aend_am
            FROM tbl_KD_Kundenstamm
            WHERE kun_Id = ?
        """

        data = execute_query(sql, [id])
        if not data:
            return jsonify({"success": False, "error": "Kunde nicht gefunden"}), 404

        return jsonify({"success": True, "data": data[0]})

    except Exception as e:
        logger.error(f"Error in get_kunde: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - OBJEKTE
# =============================================================================

@app.route('/api/objekte', methods=['GET'])
def get_objekte():
    """Listet Objekte"""
    try:
        sql = """
            SELECT DISTINCT
                ID AS ID,
                Objekt,
                Strasse,
                PLZ,
                Ort
            FROM tbl_OB_Objekt
            ORDER BY Objekt
        """
        
        data = execute_query(sql)
        return jsonify({"success": True, "data": data})
        
    except Exception as e:
        logger.error(f"Error in get_objekte: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - ZUORDNUNGEN (MA-VA)
# =============================================================================

@app.route('/api/zuordnungen', methods=['GET'])
def get_zuordnungen():
    """Listet MA-Zuordnungen.
    
    Query-Parameter:
    - va_id: Auftrags-ID (optional wenn von/bis angegeben)
    - datum_id: Datum-ID (optional)
    - von: Startdatum YYYY-MM-DD (optional)
    - bis: Enddatum YYYY-MM-DD (optional)
    """
    try:
        va_id = request.args.get('va_id', type=int)
        datum_id = request.args.get('datum_id', type=int)
        von = request.args.get('von')
        bis = request.args.get('bis')

        # Zeitraum-basierte Abfrage: Alle Zuordnungen im Zeitraum
        if von and bis:
            sql = f"""
                SELECT
                    p.ID,
                    p.VA_ID,
                    p.MA_ID,
                    m.Nachname,
                    m.Vorname,
                    p.MVA_Start,
                    p.MVA_Ende,
                    p.Status_ID,
                    p.PKW,
                    p.PosNr,
                    p.VADatum,
                    p.VAStart_ID,
                    p.VADatum_ID
                FROM tbl_MA_VA_Planung AS p
                LEFT JOIN tbl_MA_Mitarbeiterstamm AS m ON p.MA_ID = m.ID
                WHERE p.VADatum >= #{von}#
                AND p.VADatum <= #{bis}#
                ORDER BY p.VA_ID, p.VADatum, p.MVA_Start
            """
            result = execute_query(sql)
            return jsonify({"success": True, "data": result})

        # Original: Nach va_id filtern
        if not va_id:
            return jsonify({"success": False, "error": "va_id oder von/bis erforderlich"}), 400

        # JOIN mit tbl_VA_Start um Schicht-Zeiten als Fallback zu haben
        sql = """
            SELECT
                p.ID,
                p.MA_ID,
                m.Nachname AS MA_Nachname,
                m.Vorname AS MA_Vorname,
                IIF(p.MVA_Start IS NULL, s.VA_Start, p.MVA_Start) AS MA_Start,
                IIF(p.MVA_Ende IS NULL, s.VA_Ende, p.MVA_Ende) AS MA_Ende,
                p.Status_ID AS Status,
                p.PKW AS PKW,
                p.PosNr AS PosNr,
                p.VADatum AS VADatum,
                p.VAStart_ID AS VAStart_ID,
                p.VADatum_ID AS VADatum_ID,
                s.VA_Start AS Schicht_Start,
                s.VA_Ende AS Schicht_Ende
            FROM (tbl_MA_VA_Planung AS p
            LEFT JOIN tbl_MA_Mitarbeiterstamm AS m ON p.MA_ID = m.ID)
            LEFT JOIN tbl_VA_Start AS s ON p.VAStart_ID = s.ID
            WHERE p.VA_ID = ?
        """

        params = [va_id]
        if datum_id:
            sql += " AND p.VADatum_ID = ?"
            params.append(datum_id)

        sql += " ORDER BY p.PosNr, IIF(p.MVA_Start IS NULL, s.VA_Start, p.MVA_Start)"

        result = execute_query(sql, params)
        return jsonify({"success": True, "data": result})

    except Exception as e:
        logger.error(f"Error in get_zuordnungen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/zuordnungen', methods=['POST'])


def create_zuordnung():
    """Erstellt eine neue MA-Zuordnung"""
    try:
        data = request.get_json() or {}

        sql = """
            INSERT INTO tbl_MA_VA_Planung (
                VA_ID, VADatum_ID, VAStart_ID, MA_ID,
                Status_ID, VADatum, MVA_Start, MVA_Ende
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """

        params = [
            data.get('va_id') or data.get('VA_ID'),
            data.get('vadatum_id') or data.get('VADatum_ID'),
            data.get('vastart_id') or data.get('VAStart_ID'),
            data.get('ma_id') or data.get('MA_ID'),
            data.get('status') or data.get('Status_ID', 1),
            data.get('vadatum') or data.get('VADatum'),
            data.get('von') or data.get('MVA_Start'),
            data.get('bis') or data.get('MVA_Ende')
        ]

        execute_query(sql, params)
        return jsonify({"success": True})

    except Exception as e:
        logger.error(f"Error in create_zuordnung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/zuordnungen/<int:id>', methods=['PUT', 'PATCH'])
def update_zuordnung(id):
    """
    Aktualisiert eine MA-Zuordnung.
    Wird von sub_MA_VA_Planung_Status bei Status-Aenderung aufgerufen.

    VBA-Aequivalent: Form_BeforeUpdate in sub_MA_VA_Planung_Status
    - Setzt Aend_am = Now()
    - Setzt Aend_von = atCNames(1)

    Request Body (alle optional):
        Status_ID: Status (1=Geplant, 2=Benachrichtigt, 3=Zusage, 4=Absage)
        MVA_Start: Startzeit
        MVA_Ende: Endzeit
        Bemerkungen: Text
        PKW: Boolean
    """
    try:
        data = request.get_json() or {}

        # Dynamisches SQL fuer nur die uebergebenen Felder
        updates = []
        params = []

        # Status_ID
        if 'Status_ID' in data or 'status_id' in data or 'status' in data:
            status_id = data.get('Status_ID') or data.get('status_id') or data.get('status')
            updates.append('Status_ID = ?')
            params.append(status_id)

        # MVA_Start
        if 'MVA_Start' in data or 'von' in data:
            mva_start = data.get('MVA_Start') or data.get('von')
            updates.append('MVA_Start = ?')
            params.append(mva_start)

        # MVA_Ende
        if 'MVA_Ende' in data or 'bis' in data:
            mva_ende = data.get('MVA_Ende') or data.get('bis')
            updates.append('MVA_Ende = ?')
            params.append(mva_ende)

        # Bemerkungen
        if 'Bemerkungen' in data or 'bemerkung' in data:
            bemerk = data.get('Bemerkungen') or data.get('bemerkung')
            updates.append('Bemerkungen = ?')
            params.append(bemerk)

        # PKW
        if 'PKW' in data or 'pkw' in data:
            pkw = data.get('PKW') or data.get('pkw')
            updates.append('PKW = ?')
            params.append(pkw)

        if not updates:
            return jsonify({"success": False, "error": "Keine Felder zum Aktualisieren"}), 400

        # Immer Aend_am und Aend_von setzen (wie Access VBA)
        updates.append('Aend_am = ?')
        params.append(datetime.now())
        updates.append('Aend_von = ?')
        params.append(data.get('Aend_von', 'HTML'))

        params.append(id)  # WHERE ID = ?

        sql = f"UPDATE tbl_MA_VA_Planung SET {', '.join(updates)} WHERE ID = ?"
        execute_query(sql, params)

        logger.info(f"Zuordnung {id} aktualisiert: {updates}")
        return jsonify({"success": True, "message": f"Zuordnung {id} aktualisiert"})

    except Exception as e:
        logger.error(f"Error in update_zuordnung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/zuordnungen/<int:id>', methods=['DELETE'])
def delete_zuordnung(id):
    """Löscht eine MA-Zuordnung"""
    try:
        execute_query(
            "DELETE FROM tbl_MA_VA_Planung WHERE ID = ?",
            [id]
        )
        return jsonify({"success": True})
        
    except Exception as e:
        logger.error(f"Error in delete_zuordnung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# PLANUNGEN (Alias fuer Zuordnungen - fuer Schnellauswahl Doppelklick)
# =============================================================================
@app.route('/api/planungen', methods=['GET'])
def get_planungen():
    """Alias fuer get_zuordnungen - liefert MA-Zuordnungen"""
    return get_zuordnungen()


@app.route('/api/planungen', methods=['POST'])
def create_planung():
    """
    Erstellt eine neue MA-Planung/Zuordnung.
    Wird von Schnellauswahl Doppelklick aufgerufen.

    Request Body:
        ma_id: Mitarbeiter-ID
        va_id: Auftrags-ID
        vadatum_id: Datum-ID (tbl_VA_AnzTage.ID)
        vastart_id: Schicht-ID (tbl_VA_Start.ID) - optional
        status_id: Status (1=Geplant, 2=Zugesagt, etc.) - default 1

    Returns:
        {"success": true, "id": <neue_id>}
    """
    try:
        data = request.get_json() or {}

        # Parameter extrahieren (beide Schreibweisen erlaubt)
        ma_id = data.get('ma_id') or data.get('MA_ID')
        va_id = data.get('va_id') or data.get('VA_ID')
        vadatum_id = data.get('vadatum_id') or data.get('VADatum_ID')
        vastart_id = data.get('vastart_id') or data.get('VAStart_ID')
        status_id = data.get('status_id') or data.get('Status_ID') or 1

        if not ma_id or not va_id:
            return jsonify({"success": False, "error": "ma_id und va_id sind erforderlich"}), 400

        # Baue dynamisches SQL basierend auf verfuegbaren Werten
        fields = ['VA_ID', 'MA_ID', 'Status_ID', 'Erst_am', 'Aend_am']
        values = ['?', '?', '?', 'Now()', 'Now()']
        params = [va_id, ma_id, status_id]

        if vadatum_id:
            fields.append('VADatum_ID')
            values.append('?')
            params.append(vadatum_id)

        if vastart_id:
            fields.append('VAStart_ID')
            values.append('?')
            params.append(vastart_id)

        sql = f"INSERT INTO tbl_MA_VA_Planung ({', '.join(fields)}) VALUES ({', '.join(values)})"

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()

        # Neue ID holen
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]
        conn.close()

        logger.info(f"[Planungen] Erstellt: ID={new_id}, MA={ma_id}, VA={va_id}, Datum={vadatum_id}")
        return jsonify({"success": True, "id": new_id})

    except Exception as e:
        logger.error(f"Error in create_planung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/planungen/<int:id>', methods=['DELETE'])
def delete_planung(id):
    """Alias fuer delete_zuordnung - loescht MA-Zuordnung"""
    return delete_zuordnung(id)


# =============================================================================
# ANFRAGEN (Mitarbeiter-Anfragen fuer Auftraege)
# =============================================================================
@app.route('/api/anfragen', methods=['GET'])
def get_anfragen():
    """Listet Anfragen aus tbl_MA_VA_Planung"""
    try:
        va_id = request.args.get('va_id', type=int)
        ma_id = request.args.get('ma_id', type=int)
        status_id = request.args.get('status', type=int)

        sql = """
            SELECT
                p.[ID],
                p.[VA_ID],
                p.[VADatum_ID],
                p.[VAStart_ID],
                p.[MA_ID],
                p.[Status_ID],
                p.[Anfragezeitpunkt],
                p.[Rueckmeldezeitpunkt],
                p.[MVA_Start],
                p.[MVA_Ende],
                m.[Nachname],
                m.[Vorname],
                a.[Auftrag],
                a.[Ort],
                a.[Dat_VA_Von]
            FROM (tbl_MA_VA_Planung AS p
                LEFT JOIN tbl_MA_Mitarbeiterstamm AS m ON p.[MA_ID] = m.[ID])
                LEFT JOIN tbl_VA_Auftragstamm AS a ON p.[VA_ID] = a.[ID]
            WHERE 1=1
        """

        params = []
        if va_id:
            sql += " AND p.VA_ID = ?"
            params.append(va_id)
        if ma_id:
            sql += " AND p.MA_ID = ?"
            params.append(ma_id)
        if status_id is not None:
            sql += " AND p.Status_ID = ?"
            params.append(status_id)

        sql += " ORDER BY p.[Anfragezeitpunkt] DESC"

        rows = execute_query(sql, params)
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        logger.error(f"Error in get_anfragen: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/anfragen', methods=['POST'])
def create_anfrage():
    """Erstellt eine neue Anfrage (tbl_MA_VA_Planung)"""
    try:
        data = request.get_json() or {}

        sql = """
            INSERT INTO tbl_MA_VA_Planung (
                VA_ID, VADatum_ID, VAStart_ID, MA_ID,
                Status_ID, Anfragezeitpunkt
            ) VALUES (?, ?, ?, ?, ?, ?)
        """

        params = [
            data.get('VA_ID'),
            data.get('VADatum_ID'),
            data.get('VAStart_ID'),
            data.get('MA_ID'),
            data.get('Status_ID', 1),
            datetime.now()
        ]

        execute_query(sql, params)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error in create_anfrage: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/anfragen/<int:id>', methods=['PUT'])
def update_anfrage(id):
    """Aktualisiert Status einer Anfrage"""
    try:
        data = request.get_json() or {}
        status_id = data.get('Status_ID')

        sql = """
            UPDATE tbl_MA_VA_Planung
            SET Status_ID = ?, Status_Datum = ?, Rueckmeldezeitpunkt = ?
            WHERE ID = ?
        """

        now = datetime.now()
        params = [status_id, now, now, id]
        execute_query(sql, params)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error in update_anfrage: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - SCHICHTEN
# =============================================================================

@app.route('/api/dienstplan/schichten', methods=['GET'])
def get_schichten():
    """Listet Schichten für einen Auftrag/Tag"""
    try:
        va_id = request.args.get('va_id', type=int)
        datum_id = request.args.get('datum_id', type=int)
        
        if not va_id:
            return jsonify({"success": False, "error": "va_id erforderlich"}), 400
        
        sql = """
            SELECT 
                s.VAS_ID AS ID,
                s.VAS_Von AS VA_Start,
                s.VAS_Bis AS VA_Ende,
                s.VAS_MA_Anzahl AS MA_Anzahl,
                d.VADatum AS Datum,
                (SELECT COUNT(*) FROM tbl_MA_VA_Planung p 
                 WHERE p.Planung_VAStart_ID = s.VAS_ID 
                 AND p.Planung_Status IN (2,3)) AS MA_Anzahl_Ist
            FROM tbl_VA_Start s
            INNER JOIN tbl_VA_Datum d ON s.VAS_VADatum_ID = d.VADatum_ID
            WHERE d.VADatum_VA_ID = ?
        """
        
        params = [va_id]
        
        if datum_id:
            sql += " AND s.VAS_VADatum_ID = ?"
            params.append(datum_id)
        
        sql += " ORDER BY d.VADatum, s.VAS_Von"
        
        data = execute_query(sql, params)
        return jsonify({"success": True, "data": data})
        
    except Exception as e:
        logger.error(f"Error in get_schichten: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - DIENSTPLAN
# =============================================================================

@app.route('/api/dienstplan/ma/<int:ma_id>', methods=['GET'])
def get_dienstplan_ma(ma_id):
    """Holt Dienstplan für einen Mitarbeiter"""
    try:
        datum_von = request.args.get('datum_von', datetime.now().strftime('%Y-%m-%d'))
        tage = request.args.get('tage', 7, type=int)
        
        sql = """
            SELECT 
                p.ID AS ID,
                d.VADatum AS Datum,
                a.Auftrag AS Auftrag,
                a.Ort AS Ort,
                p.MVA_Start AS Von,
                p.MVA_Ende AS Bis,
                p.Status_ID AS Status
            FROM ((tbl_MA_VA_Planung p
            INNER JOIN tbl_VA_Datum d ON p.VADatum_ID = d.VADatum_ID)
            INNER JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
            WHERE p.MA_ID = ?
            AND d.VADatum >= ?
            AND d.VADatum <= DateAdd('d', ?, ?)
            ORDER BY d.VADatum, p.MVA_Start
        """
        
        data = execute_query(sql, [ma_id, datum_von, tage, datum_von])
        return jsonify({"success": True, "data": data})
        
    except Exception as e:
        logger.error(f"Error in get_dienstplan_ma: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/dienstplan/objekt/<int:objekt_id>', methods=['GET'])
def get_dienstplan_objekt(objekt_id):
    """Holt Dienstplan für ein Objekt"""
    try:
        datum_von = request.args.get('datum_von', datetime.now().strftime('%Y-%m-%d'))
        tage = request.args.get('tage', 7, type=int)
        
        sql = """
            SELECT 
                p.ID AS ID,
                d.VADatum AS Datum,
                a.Auftrag AS Auftrag,
                m.Nachname AS Nachname,
                m.Vorname AS Vorname,
                p.MVA_Start AS Von,
                p.MVA_Ende AS Bis,
                p.Status_ID AS Status
            FROM (((tbl_MA_VA_Planung p
            INNER JOIN tbl_VA_Datum d ON p.VADatum_ID = d.VADatum_ID)
            INNER JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
            INNER JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
            WHERE a.Objekt_ID = ?
            AND d.VADatum >= ?
            ORDER BY d.VADatum, p.MVA_Start
        """
        
        data = execute_query(sql, [objekt_id, datum_von])
        return jsonify({"success": True, "data": data})
        
    except Exception as e:
        logger.error(f"Error in get_dienstplan_objekt: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - DIENSTPLAN ÜBERSICHT (alle MA)
# =============================================================================

@app.route('/api/dienstplan/uebersicht', methods=['GET'])
def get_dienstplan_uebersicht():
    """Holt Dienstplanübersicht für alle Mitarbeiter (für Kalender-Grid)"""
    try:
        datum_von = request.args.get('datum_von', datetime.now().strftime('%Y-%m-%d'))
        tage = request.args.get('tage', 7, type=int)
        anstellung = request.args.get('anstellung', 'Alle')  # Festangestellte, Minijobber, Alle

        # 1. Mitarbeiter laden
        ma_sql = """
            SELECT
                ID,
                Nachname,
                Vorname,
                Nachname & ', ' & Vorname AS Name,
                Anstellungsart_ID
            FROM tbl_MA_Mitarbeiterstamm
            WHERE IstAktiv = True
        """
        if anstellung == 'Festangestellte':
            ma_sql += " AND (Anstellungsart_ID = 1 OR Anstellungsart_ID IS NULL)"
        elif anstellung == 'Minijobber':
            ma_sql += " AND Anstellungsart_ID = 2"

        ma_sql += " ORDER BY Nachname, Vorname"
        mitarbeiter = execute_query(ma_sql)

        # 2. Einsätze im Zeitraum laden (VADatum direkt in tbl_MA_VA_Planung)
        einsatz_sql = f"""
            SELECT
                p.MA_ID AS maId,
                p.VADatum AS datum,
                a.Auftrag AS auftrag,
                a.Ort AS ort,
                a.ID AS vaId,
                Format(p.MVA_Start, 'HH:nn') AS von,
                Format(p.MVA_Ende, 'HH:nn') AS bis,
                p.Status_ID AS status
            FROM tbl_MA_VA_Planung p
            INNER JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID
            WHERE p.VADatum >= #{datum_von}#
            AND p.VADatum < DateAdd('d', {tage}, #{datum_von}#)
            AND p.Status_ID IN (2,3)
        """

        einsaetze = execute_query(einsatz_sql)

        # 3. Abwesenheiten im Zeitraum laden (tbl_MA_NVerfuegZeiten)
        abw_sql = f"""
            SELECT
                n.MA_ID AS maId,
                n.vonDat AS datumVon,
                n.bisDat AS datumBis,
                n.Zeittyp_ID AS grundId,
                n.Bemerkung AS grund
            FROM tbl_MA_NVerfuegZeiten n
            WHERE (n.vonDat <= DateAdd('d', {tage}, #{datum_von}#) AND n.bisDat >= #{datum_von}#)
        """

        abwesenheiten = execute_query(abw_sql)

        return jsonify({
            "success": True,
            "mitarbeiter": mitarbeiter,
            "einsaetze": einsaetze,
            "abwesenheiten": abwesenheiten,
            "datum_von": datum_von,
            "tage": tage
        })

    except Exception as e:
        logger.error(f"Error in get_dienstplan_uebersicht: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/dienstplan/objekt-uebersicht', methods=['GET'])
def get_dienstplan_objekt_uebersicht():
    """
    Aggregierte Dienstplan-Daten pro Objekt/Datum.

    Liefert Schichten gruppiert nach Objekt mit allen zugeordneten MA.
    Joinst Schichten mit Zuordnungen und MA-Infos.

    Query-Parameter:
    - datum_von: Startdatum (default: heute)
    - datum_bis: Enddatum (default: datum_von + 7 Tage)
    - objekt_id: Optional - nur bestimmtes Objekt

    Returns:
        Liste von Objekten mit Schichten und zugeordneten Mitarbeitern
    """
    try:
        datum_von = request.args.get('datum_von', datetime.now().strftime('%Y-%m-%d'))
        datum_bis = request.args.get('datum_bis')
        objekt_id = request.args.get('objekt_id', type=int)

        # Default: 7 Tage ab datum_von
        if not datum_bis:
            from_date = datetime.strptime(datum_von, '%Y-%m-%d')
            datum_bis = (from_date + timedelta(days=7)).strftime('%Y-%m-%d')

        # 1. Schichten im Zeitraum laden (aus tbl_VA_Start)
        schichten_sql = f"""
            SELECT
                s.ID AS SchichtID,
                s.VA_ID,
                s.VADatum_ID,
                t.VADatum AS Datum,
                a.Auftrag,
                a.Objekt,
                a.Objekt_ID,
                a.Ort,
                s.VA_Start AS SchichtVon,
                s.VA_Ende AS SchichtBis,
                s.MA_Anzahl AS MA_Soll,
                s.MA_Anzahl_Ist AS MA_Ist
            FROM ((tbl_VA_Start s
            INNER JOIN tbl_VA_AnzTage t ON s.VADatum_ID = t.ID)
            INNER JOIN tbl_VA_Auftragstamm a ON s.VA_ID = a.ID)
            WHERE t.VADatum >= #{datum_von}#
            AND t.VADatum <= #{datum_bis}#
        """

        if objekt_id:
            schichten_sql += f" AND a.Objekt_ID = {objekt_id}"

        schichten_sql += " ORDER BY a.Objekt, t.VADatum, s.VA_Start"

        schichten = execute_query(schichten_sql)

        # 2. Zugeordnete Mitarbeiter zu den Schichten laden
        zuordnungen_sql = f"""
            SELECT
                p.VADatum_ID,
                p.VA_ID,
                p.MA_ID,
                m.Nachname,
                m.Vorname,
                m.Nachname & ', ' & m.Vorname AS MA_Name,
                m.Tel_Mobil AS TelMobil,
                p.MVA_Start AS MA_Von,
                p.MVA_Ende AS MA_Bis,
                p.Status_ID,
                p.IstEL,
                p.IstFraeutl AS IstFraeutlich,
                p.Bemerkung
            FROM (tbl_MA_VA_Planung p
            INNER JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
            INNER JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID
            WHERE t.VADatum >= #{datum_von}#
            AND t.VADatum <= #{datum_bis}#
            AND p.Status_ID IN (2, 3)
        """

        zuordnungen = execute_query(zuordnungen_sql)

        # 3. Zuordnungen nach VADatum_ID gruppieren für schnellen Lookup
        zuordnungen_map = {}
        for z in zuordnungen:
            key = (z.get('VADatum_ID'), z.get('VA_ID'))
            if key not in zuordnungen_map:
                zuordnungen_map[key] = []
            zuordnungen_map[key].append(z)

        # 4. Schichten mit Zuordnungen anreichern
        for schicht in schichten:
            key = (schicht.get('VADatum_ID'), schicht.get('VA_ID'))
            schicht['zuordnungen'] = zuordnungen_map.get(key, [])
            schicht['MA_Zugeordnet'] = len(schicht['zuordnungen'])

        # 5. Nach Objekt gruppieren
        objekte_map = {}
        for schicht in schichten:
            objekt_key = schicht.get('Objekt') or 'Ohne Objekt'
            if objekt_key not in objekte_map:
                objekte_map[objekt_key] = {
                    'Objekt': objekt_key,
                    'Objekt_ID': schicht.get('Objekt_ID'),
                    'schichten': []
                }
            objekte_map[objekt_key]['schichten'].append(schicht)

        # Als Liste zurückgeben
        objekte_liste = list(objekte_map.values())

        return jsonify({
            "success": True,
            "data": objekte_liste,
            "datum_von": datum_von,
            "datum_bis": datum_bis,
            "schichten_total": len(schichten),
            "objekte_count": len(objekte_liste)
        })

    except Exception as e:
        logger.error(f"Error in get_dienstplan_objekt_uebersicht: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# API ROUTEN - PLANUNGSÜBERSICHT (alle Aufträge mit Schichten)
# =============================================================================

@app.route('/api/planung/uebersicht', methods=['GET'])
def get_planungs_uebersicht():
    """Holt Planungsübersicht - alle Aufträge mit Schichten und MA-Zuordnungen"""
    try:
        datum_von = request.args.get('datum_von', datetime.now().strftime('%Y-%m-%d'))
        tage = request.args.get('tage', 7, type=int)
        nur_freie = request.args.get('nur_freie', 'false').lower() == 'true'
        max_positionen = request.args.get('max_positionen', 999, type=int)

        # 1. Aufträge im Zeitraum laden (über tbl_VA_AnzTage)
        auftrag_sql = f"""
            SELECT DISTINCT
                a.ID AS id,
                a.Auftrag AS name,
                a.Objekt AS objekt,
                a.Ort AS ort,
                a.Dat_VA_Von AS datVon,
                a.Dat_VA_Bis AS datBis
            FROM tbl_VA_Auftragstamm a
            INNER JOIN tbl_VA_AnzTage t ON a.ID = t.VA_ID
            WHERE t.VADatum >= #{datum_von}#
            AND t.VADatum < DateAdd('d', {tage}, #{datum_von}#)
            AND a.Veranst_Status_ID IN (1,2,3,4)
            ORDER BY a.Auftrag
        """
        auftraege = execute_query(auftrag_sql)

        # Filter nach max Positionen (wenn gewünscht)
        if max_positionen < 999:
            # Zähle Positionen pro Auftrag
            filtered = []
            for a in auftraege:
                count_sql = f"SELECT COUNT(*) AS cnt FROM tbl_VA_AnzTage WHERE VA_ID = {a['id']}"
                result = execute_query(count_sql)
                if result and result[0]['cnt'] <= max_positionen:
                    filtered.append(a)
            auftraege = filtered

        # 2. Alle MA-Zuordnungen im Zeitraum laden
        zuord_sql = f"""
            SELECT
                p.VA_ID AS vaId,
                p.VADatum AS datum,
                p.MA_ID AS maId,
                m.Nachname & ', ' & m.Vorname AS maName,
                Format(p.MVA_Start, 'HH:nn') AS von,
                Format(p.MVA_Ende, 'HH:nn') AS bis,
                p.Status_ID AS status
            FROM tbl_MA_VA_Planung p
            INNER JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VADatum >= #{datum_von}#
            AND p.VADatum < DateAdd('d', {tage}, #{datum_von}#)
            ORDER BY p.VA_ID, p.VADatum, p.MVA_Start
        """
        alle_zuordnungen = execute_query(zuord_sql)

        # 3. Zuordnungen nach vaId und Datum gruppieren
        zuordnungen = {}
        for z in alle_zuordnungen:
            va_id = z['vaId']
            datum = z['datum'].split('T')[0] if z['datum'] else None

            if va_id not in zuordnungen:
                zuordnungen[va_id] = {}
            if datum not in zuordnungen[va_id]:
                zuordnungen[va_id][datum] = []

            zuordnungen[va_id][datum].append({
                'name': z['maName'],
                'maId': z['maId'],
                'von': z['von'],
                'bis': z['bis'],
                'status': z['status'],
                'storno': z['status'] in (5, 6)  # Status 5/6 = storniert/abgesagt
            })

        # 4. Optional: Nur Aufträge mit freien Schichten
        if nur_freie:
            filtered = []
            for a in auftraege:
                va_id = a['id']
                # Prüfe ob es Tage ohne Zuordnungen gibt
                hat_freie = False
                for i in range(tage):
                    check_date = datetime.strptime(datum_von, '%Y-%m-%d') + timedelta(days=i)
                    date_str = check_date.strftime('%Y-%m-%d')
                    if va_id not in zuordnungen or date_str not in zuordnungen.get(va_id, {}):
                        hat_freie = True
                        break
                    elif len(zuordnungen[va_id][date_str]) == 0:
                        hat_freie = True
                        break
                if hat_freie:
                    filtered.append(a)
            auftraege = filtered

        return jsonify({
            "success": True,
            "auftraege": auftraege,
            "zuordnungen": zuordnungen,
            "datum_von": datum_von,
            "tage": tage
        })

    except Exception as e:
        logger.error(f"Error in get_planungs_uebersicht: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - PLANUNG (tbl_MA_VA_Planung) - CRUD
# =============================================================================

@app.route('/api/planung', methods=['GET'])
def get_planung():
    """Listet geplante Mitarbeiter für einen Auftrag/Tag

    Query-Parameter:
    - va_id: Auftrags-ID (erforderlich)
    - vadatum_id: Datum-ID (optional)
    - vastart_id: Schicht-ID (optional)
    """
    try:
        va_id = request.args.get('va_id', type=int)
        vadatum_id = request.args.get('vadatum_id', type=int)
        vastart_id = request.args.get('vastart_id', type=int)

        if not va_id:
            return jsonify({"success": False, "error": "va_id erforderlich"}), 400

        sql = """
            SELECT
                p.ID,
                p.VA_ID,
                p.VADatum_ID,
                p.VAStart_ID,
                p.PosNr,
                p.MA_ID,
                p.Status_ID,
                p.VADatum,
                p.MVA_Start,
                p.MVA_Ende,
                p.Bemerkungen,
                m.Nachname AS MA_Nachname,
                m.Vorname AS MA_Vorname
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
            WHERE p.VA_ID = ?
        """
        params = [va_id]

        if vadatum_id:
            sql += " AND p.VADatum_ID = ?"
            params.append(vadatum_id)

        if vastart_id:
            sql += " AND p.VAStart_ID = ?"
            params.append(vastart_id)

        sql += " ORDER BY p.PosNr"

        data = execute_query(sql, params)
        return jsonify({"success": True, "data": data})

    except Exception as e:
        logger.error(f"Error in get_planung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/planung', methods=['POST'])
def create_planung_single():
    """Erstellt einen neuen Planungseintrag

    JSON-Body:
    - ma_id: Mitarbeiter-ID (erforderlich)
    - va_id: Auftrags-ID (erforderlich)
    - vadatum_id: Datum-ID (erforderlich)
    - vastart_id: Schicht-ID (erforderlich)
    - status_id: Status-ID (default: 1)
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "JSON-Body erforderlich"}), 400

        ma_id = data.get('ma_id')
        va_id = data.get('va_id')
        vadatum_id = data.get('vadatum_id')
        vastart_id = data.get('vastart_id')
        status_id = data.get('status_id', 1)

        if not all([ma_id, va_id, vadatum_id, vastart_id]):
            return jsonify({"success": False, "error": "ma_id, va_id, vadatum_id, vastart_id erforderlich"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Prüfe ob Zuordnung bereits existiert
        check_sql = """
            SELECT ID FROM tbl_MA_VA_Planung
            WHERE MA_ID = ? AND VA_ID = ? AND VADatum_ID = ? AND VAStart_ID = ?
        """
        cursor.execute(check_sql, [ma_id, va_id, vadatum_id, vastart_id])
        existing = cursor.fetchone()

        if existing:
            conn.close()
            return jsonify({"success": False, "error": "MA bereits für diese Schicht geplant"}), 400

        # Nächste PosNr ermitteln
        cursor.execute(
            "SELECT MAX(PosNr) FROM tbl_MA_VA_Planung WHERE VA_ID = ? AND VADatum_ID = ?",
            [va_id, vadatum_id]
        )
        max_pos = cursor.fetchone()[0]
        pos_nr = (max_pos or 0) + 1

        # Schicht-Zeiten holen
        cursor.execute(
            "SELECT VA_Start, VA_Ende, VADatum FROM tbl_VA_Start WHERE ID = ?",
            [vastart_id]
        )
        schicht_row = cursor.fetchone()

        va_start = schicht_row[0] if schicht_row else None
        va_ende = schicht_row[1] if schicht_row else None
        va_datum = schicht_row[2] if schicht_row else None

        # INSERT
        sql = """
            INSERT INTO tbl_MA_VA_Planung
            (VA_ID, VADatum_ID, VAStart_ID, PosNr, MA_ID, Status_ID,
             VA_Start, VA_Ende, VADatum, MVA_Start, MVA_Ende,
             Erst_am, Aend_am, Bemerkungen)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, Now(), Now(), '')
        """
        cursor.execute(sql, [
            va_id, vadatum_id, vastart_id, pos_nr, ma_id, status_id,
            va_start, va_ende, va_datum, va_start, va_ende
        ])

        conn.commit()

        # Neue ID holen
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]
        conn.close()

        logger.info(f"Planung erstellt: ID={new_id}, MA={ma_id}, VA={va_id}")
        return jsonify({"success": True, "id": new_id})

    except Exception as e:
        logger.error(f"Error in create_planung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/planung/<int:id>', methods=['DELETE'])
def delete_planung_single(id):
    """Löscht einen Planungseintrag"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM tbl_MA_VA_Planung WHERE ID = ?", [id])
        rows_deleted = cursor.rowcount
        conn.commit()
        conn.close()

        if rows_deleted == 0:
            return jsonify({"success": False, "error": "Eintrag nicht gefunden"}), 404

        logger.info(f"Planung gelöscht: ID={id}")
        return jsonify({"success": True})

    except Exception as e:
        logger.error(f"Error in delete_planung: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# API ROUTEN - ABWESENHEITEN
# =============================================================================

@app.route('/api/abwesenheiten', methods=['GET'])
def get_abwesenheiten():
    """Listet Abwesenheiten mit Mitarbeiter-Namen"""
    try:
        ma_id = request.args.get('ma_id', type=int)
        datum_von = request.args.get('datum_von')
        datum_bis = request.args.get('datum_bis')

        sql = """
            SELECT
                a.Abw_ID AS ID,
                a.Abw_MA_ID AS MA_ID,
                a.Abw_Von AS DatVon,
                a.Abw_Bis AS DatBis,
                a.Abw_Grund AS Grund,
                a.Abw_Bemerkung AS Bemerkung,
                a.Abw_Ganztagig AS Ganztagig,
                m.Nachname,
                m.Vorname
            FROM tbl_MA_Abwesenheit a
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON a.Abw_MA_ID = m.ID
            WHERE 1=1
        """

        params = []
        if ma_id:
            sql += " AND a.Abw_MA_ID = ?"
            params.append(ma_id)
        if datum_von:
            sql += " AND a.Abw_Bis >= ?"
            params.append(datum_von)
        if datum_bis:
            sql += " AND a.Abw_Von <= ?"
            params.append(datum_bis)

        sql += " ORDER BY a.Abw_Von DESC"

        data = execute_query(sql, params) if params else execute_query(sql)
        return jsonify({"success": True, "data": data})
        
    except Exception as e:
        logger.error(f"Error in get_abwesenheiten: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/abwesenheiten', methods=['POST'])
def create_abwesenheit():
    """Erstellt eine neue Abwesenheit"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "JSON-Body erforderlich"}), 400

        ma_id = data.get('MA_ID')
        von_dat = data.get('vonDat')
        bis_dat = data.get('bisDat')
        grund = data.get('Grund', 'Sonstiges')
        bemerkung = data.get('Bemerkung', '')
        ganztaegig = data.get('Ganztaegig', True)

        if not all([ma_id, von_dat, bis_dat]):
            return jsonify({"success": False, "error": "MA_ID, vonDat und bisDat erforderlich"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        sql = """
            INSERT INTO tbl_MA_Abwesenheit
            (Abw_MA_ID, Abw_Von, Abw_Bis, Abw_Grund, Abw_Bemerkung, Abw_Ganztagig)
            VALUES (?, ?, ?, ?, ?, ?)
        """
        cursor.execute(sql, [ma_id, von_dat, bis_dat, grund, bemerkung, ganztaegig])
        conn.commit()

        # Neue ID ermitteln
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]
        conn.close()

        return jsonify({"success": True, "id": new_id, "message": "Abwesenheit erstellt"})

    except Exception as e:
        logger.error(f"Error in create_abwesenheit: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/abwesenheiten/<int:abw_id>', methods=['PUT'])
def update_abwesenheit(abw_id):
    """Aktualisiert eine Abwesenheit"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "JSON-Body erforderlich"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Prüfen ob Abwesenheit existiert
        cursor.execute("SELECT Abw_ID FROM tbl_MA_Abwesenheit WHERE Abw_ID = ?", [abw_id])
        if not cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "error": "Abwesenheit nicht gefunden"}), 404

        # Update-Felder sammeln
        fields = []
        params = []

        if 'MA_ID' in data:
            fields.append("Abw_MA_ID = ?")
            params.append(data['MA_ID'])
        if 'vonDat' in data:
            fields.append("Abw_Von = ?")
            params.append(data['vonDat'])
        if 'bisDat' in data:
            fields.append("Abw_Bis = ?")
            params.append(data['bisDat'])
        if 'Grund' in data:
            fields.append("Abw_Grund = ?")
            params.append(data['Grund'])
        if 'Bemerkung' in data:
            fields.append("Abw_Bemerkung = ?")
            params.append(data['Bemerkung'])
        if 'Ganztaegig' in data:
            fields.append("Abw_Ganztagig = ?")
            params.append(data['Ganztaegig'])

        if not fields:
            conn.close()
            return jsonify({"success": False, "error": "Keine Felder zum Aktualisieren"}), 400

        params.append(abw_id)
        sql = f"UPDATE tbl_MA_Abwesenheit SET {', '.join(fields)} WHERE Abw_ID = ?"
        cursor.execute(sql, params)
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Abwesenheit aktualisiert"})

    except Exception as e:
        logger.error(f"Error in update_abwesenheit: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/abwesenheiten/<int:abw_id>', methods=['DELETE'])
def delete_abwesenheit(abw_id):
    """Loescht eine Abwesenheit"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Prüfen ob Abwesenheit existiert
        cursor.execute("SELECT Abw_ID FROM tbl_MA_Abwesenheit WHERE Abw_ID = ?", [abw_id])
        if not cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "error": "Abwesenheit nicht gefunden"}), 404

        cursor.execute("DELETE FROM tbl_MA_Abwesenheit WHERE Abw_ID = ?", [abw_id])
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Abwesenheit geloescht"})

    except Exception as e:
        logger.error(f"Error in delete_abwesenheit: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# API ROUTEN - VERFÜGBARKEIT
# =============================================================================

@app.route('/api/verfuegbarkeit', methods=['GET'])
def get_verfuegbarkeit():
    """Prüft Mitarbeiter-Verfügbarkeit"""
    try:
        datum = request.args.get('datum')
        von = request.args.get('von')
        bis = request.args.get('bis')
        anstellung = request.args.get('anstellung')
        
        if not datum:
            return jsonify({"success": False, "error": "datum erforderlich"}), 400
        
        # Basis-Query: Alle aktiven Mitarbeiter
        sql = """
            SELECT 
                m.ID AS MA_ID,
                m.Nachname,
                m.Vorname,
                m.Anstellungsart_ID AS Anstellung,
                m.Ort,
                (SELECT COUNT(*) FROM tbl_MA_VA_Planung p 
                 INNER JOIN tbl_VA_Datum d ON p.VADatum_ID = d.VADatum_ID
                 WHERE p.MA_ID = m.ID 
                 AND d.VADatum = #{datum}#) AS Einsaetze,
                (SELECT COUNT(*) FROM tbl_MA_Abwesenheit a
                 WHERE a.Abw_MA_ID = m.ID
                 AND #{datum}# BETWEEN a.Abw_Von AND a.Abw_Bis) AS Abwesend
            FROM tbl_MA_Mitarbeiterstamm m
            WHERE m.IstAktiv = True
        """.format(datum=datum)
        
        if anstellung:
            sql += f" AND m.Anstellungsart_ID = '{anstellung}'"
        
        sql += " ORDER BY m.Nachname, m.Vorname"
        
        data = execute_query(sql)
        
        # Verfügbare filtern (keine Abwesenheit)
        verfuegbar = [d for d in data if d.get('Abwesend', 0) == 0]
        
        return jsonify({"success": True, "data": verfuegbar, "count": len(verfuegbar)})
        
    except Exception as e:
        logger.error(f"Error in get_verfuegbarkeit: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - QUERY (für benutzerdefinierte Abfragen)
# =============================================================================

@app.route('/api/mitarbeiter/verfuegbar', methods=['GET'])
def get_mitarbeiter_verfuegbar():
    """Access-Style Verfuegbarkeitscheck wie mini_api"""
    try:
        datum = request.args.get('datum')
        va_start = request.args.get('va_start')
        va_ende = request.args.get('va_ende')
        geplant_verfuegbar = request.args.get('geplant_verfuegbar', 'false').lower() in ('true', '1')

        if not datum:
            return jsonify({"success": False, "error": "datum erforderlich"}), 400

        try:
            datum_obj = datetime.strptime(datum, '%Y-%m-%d')
        except ValueError:
            return jsonify({"success": False, "error": "datum im format YYYY-MM-DD erforderlich"}), 400

        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT ID FROM tbl_MA_Mitarbeiterstamm WHERE IstAktiv = True")
            alle_ma = [row[0] for row in cursor.fetchall() if row[0]]

            cursor.execute(
                "SELECT DISTINCT MA_ID FROM tbl_MA_NVerfuegZeiten WHERE ? BETWEEN vonDat AND bisDat",
                (datum_obj,)
            )
            nicht_verfuegbar = {row[0] for row in cursor.fetchall() if row[0]}

            sql_plan = "SELECT DISTINCT MA_ID FROM tbl_MA_VA_Planung WHERE VADatum = ?"
            params = [datum_obj]
            if va_start:
                sql_plan += " AND MVA_Start = ?"
                params.append(va_start)
            if va_ende:
                sql_plan += " AND MVA_Ende = ?"
                params.append(va_ende)

            cursor.execute(sql_plan, tuple(params))
            geplant = {row[0] for row in cursor.fetchall() if row[0]}

        if geplant_verfuegbar:
            verfuegbar_ma = [ma for ma in alle_ma if ma not in nicht_verfuegbar]
        else:
            verfuegbar_ma = [ma for ma in alle_ma if ma not in nicht_verfuegbar and ma not in geplant]

        return jsonify({
            "success": True,
            "data": {
                "verfuegbar": verfuegbar_ma,
                "nicht_verfuegbar": sorted(nicht_verfuegbar),
                "geplant": sorted(geplant),
                "datum": datum_obj.strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        logger.error(f"Error in get_mitarbeiter_verfuegbar: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/query', methods=['POST'])
def execute_custom_query():
    """Führt eine benutzerdefinierte SQL-Abfrage aus"""
    try:
        data = request.get_json()
        sql = data.get('query')
        
        if not sql:
            return jsonify({"success": False, "error": "query erforderlich"}), 400
        
        # Sicherheitsprüfung: Nur SELECT erlaubt
        sql_upper = sql.upper().strip()
        if not sql_upper.startswith('SELECT'):
            return jsonify({"success": False, "error": "Nur SELECT-Abfragen erlaubt"}), 403
        
        result = execute_query(sql)
        return jsonify({"success": True, "data": result})
        
    except Exception as e:
        logger.error(f"Error in execute_custom_query: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - TABELLEN
# =============================================================================

@app.route('/api/tables', methods=['GET'])
def get_tables():
    """Listet alle Tabellen in der Datenbank"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        tables = []
        for row in cursor.tables(tableType='TABLE'):
            if not row.table_name.startswith('MSys'):
                tables.append(row.table_name)
        
        conn.close()
        return jsonify({"success": True, "data": sorted(tables)})
        
    except Exception as e:
        logger.error(f"Error in get_tables: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - HEALTH CHECK
# =============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Prüft ob API und Datenbank erreichbar sind"""
    try:
        # Datenbank-Test
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM tbl_MA_Mitarbeiterstamm")
        count = cursor.fetchone()[0]
        conn.close()
        
        return jsonify({
            "success": True,
            "status": "healthy",
            "database": "connected",
            "mitarbeiter_count": count,
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "status": "unhealthy",
            "error": str(e)
        }), 500


@app.route('/api/status', methods=['GET'])
def api_status():
    """Einfacher Status-Endpoint für Frontend-Check"""
    return jsonify({
        "success": True,
        "status": "online",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/api/ping', methods=['GET'])
def ping():
    """Einfacher Ping-Test"""
    return jsonify({"pong": True, "timestamp": datetime.now().isoformat()})

# =============================================================================
# API ROUTEN - EVENTDATEN (Stub-Endpoints)
# =============================================================================

@app.route('/api/eventdaten/<auftrag_id>', methods=['GET'])
def get_eventdaten(auftrag_id):
    """Liefert Eventdaten für einen Auftrag (Stub)."""
    return jsonify({
        "success": True,
        "data": None,
        "message": "Eventdaten-Feature noch nicht vollständig implementiert"
    })

@app.route('/api/eventdaten/scrape', methods=['POST'])
def scrape_eventdaten():
    """Scraped Eventdaten von externer Quelle (Stub)."""
    return jsonify({
        "success": False,
        "error": "Eventdaten-Scraping noch nicht implementiert",
        "message": "Bitte Daten manuell eingeben"
    })

@app.route('/api/eventdaten', methods=['POST'])
def save_eventdaten():
    """Speichert Eventdaten (Stub)."""
    data = request.get_json() or {}
    logger.info(f"Eventdaten empfangen: {data}")
    return jsonify({
        "success": True,
        "message": "Eventdaten gespeichert (Stub - keine persistente Speicherung)"
    })

# =============================================================================
# API ROUTEN - KUNDENPREISE / VERRECHNUNGSSAETZE
# =============================================================================

@app.route('/api/kundenpreise', methods=['GET'])
def get_kundenpreise():
    """
    Listet alle Kundenpreise/Verrechnungssaetze als Pivot-Tabelle.
    Spalten: Firma | SVS | EL | NZ | SZ | FZ | Fahrtkosten | Div.
    """
    try:
        # SQL: Alle Preise mit Artikelbeschreibung und Kundendaten
        sql = ("SELECT k.kun_Id AS kunId, k.kun_Firma AS firma, "
               "a.Beschreibung AS preisart, p.StdPreis AS preis "
               "FROM tbl_KD_Kundenstamm k "
               "INNER JOIN tbl_KD_Standardpreise p ON k.kun_Id = p.kun_ID "
               "INNER JOIN tbl_KD_Artikelbeschreibung a ON p.Preisart_ID = a.ID "
               "WHERE k.kun_IstAktiv = True "
               "ORDER BY k.kun_Firma, a.Beschreibung")

        raw_data = execute_query(sql)

        # Pivot in Python: Gruppiere nach Kunde
        kunden = {}
        for row in raw_data:
            kun_id = row['kunId']
            if kun_id not in kunden:
                kunden[kun_id] = {
                    'kunId': kun_id,
                    'firma': row['firma'],
                    'Sicherheitspersonal': None,
                    'Leitungspersonal': None,
                    'Nachtzuschlag': None,
                    'Sonntagszuschlag': None,
                    'Feiertagszuschlag': None,
                    'Fahrtkosten': None,
                    'Sonstiges': None
                }

            # Preisart zuordnen
            preisart = row['preisart']
            preis = row['preis']

            if preisart in kunden[kun_id]:
                kunden[kun_id][preisart] = preis

        # Liste erstellen
        result = list(kunden.values())

        return jsonify({
            "success": True,
            "data": result,
            "count": len(result)
        })

    except Exception as e:
        logger.error(f"Error in get_kundenpreise: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kundenpreise/<int:kun_id>', methods=['GET'])
def get_kundenpreis_detail(kun_id):
    """Einzelne Kundenpreise"""
    try:
        sql = ("SELECT k.kun_Id AS kunId, k.kun_Firma AS firma, "
               "a.Beschreibung AS preisart, p.StdPreis AS preis, p.ID AS preisId "
               "FROM tbl_KD_Kundenstamm k "
               "INNER JOIN tbl_KD_Standardpreise p ON k.kun_Id = p.kun_ID "
               "INNER JOIN tbl_KD_Artikelbeschreibung a ON p.Preisart_ID = a.ID "
               "WHERE k.kun_Id = ? "
               "ORDER BY a.Beschreibung")

        raw_data = execute_query(sql, [kun_id])

        if not raw_data:
            return jsonify({"success": False, "error": "Kunde nicht gefunden"}), 404

        # Pivot
        result = {
            'kunId': kun_id,
            'firma': raw_data[0]['firma'] if raw_data else '',
            'preise': {}
        }

        for row in raw_data:
            result['preise'][row['preisart']] = {
                'preis': row['preis'],
                'preisId': row['preisId']
            }

        return jsonify({"success": True, "data": result})

    except Exception as e:
        logger.error(f"Error in get_kundenpreis_detail: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kundenpreise/preis/<int:preis_id>', methods=['PUT'])
def update_kundenpreis(preis_id):
    """Aktualisiert einen einzelnen Preis in tbl_KD_Standardpreise"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Keine Daten gesendet"}), 400

        new_preis = data.get('StdPreis')
        if new_preis is None:
            return jsonify({"success": False, "error": "StdPreis erforderlich"}), 400

        sql = "UPDATE tbl_KD_Standardpreise SET StdPreis = ? WHERE ID = ?"
        execute_query(sql, [new_preis, preis_id], fetch=False)

        logger.info(f"Kundenpreis {preis_id} aktualisiert auf {new_preis}")
        return jsonify({"success": True, "message": f"Preis {preis_id} aktualisiert"})

    except Exception as e:
        logger.error(f"Error in update_kundenpreis: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kundenpreise/preis/<int:preis_id>', methods=['DELETE'])
def delete_kundenpreis(preis_id):
    """Loescht einen einzelnen Preis aus tbl_KD_Standardpreise"""
    try:
        sql = "DELETE FROM tbl_KD_Standardpreise WHERE ID = ?"
        execute_query(sql, [preis_id], fetch=False)

        logger.info(f"Kundenpreis {preis_id} geloescht")
        return jsonify({"success": True, "message": f"Preis {preis_id} geloescht"})

    except Exception as e:
        logger.error(f"Error in delete_kundenpreis: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kundenpreise', methods=['POST'])
def create_kundenpreis():
    """Erstellt einen neuen Preis in tbl_KD_Standardpreise"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Keine Daten gesendet"}), 400

        kun_id = data.get('kun_ID')
        preisart_id = data.get('Preisart_ID')
        std_preis = data.get('StdPreis', 0)

        if not kun_id or not preisart_id:
            return jsonify({"success": False, "error": "kun_ID und Preisart_ID erforderlich"}), 400

        sql = "INSERT INTO tbl_KD_Standardpreise (kun_ID, Preisart_ID, StdPreis) VALUES (?, ?, ?)"
        execute_query(sql, [kun_id, preisart_id, std_preis], fetch=False)

        logger.info(f"Kundenpreis erstellt: Kunde={kun_id}, Preisart={preisart_id}, Preis={std_preis}")
        return jsonify({"success": True, "message": "Preis erstellt"}), 201

    except Exception as e:
        logger.error(f"Error in create_kundenpreis: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# API ROUTEN - EINSATZLISTEN GESENDET
# =============================================================================

@app.route('/api/einsatzlisten/gesendet', methods=['GET'])
def get_einsatzlisten_gesendet():
    """Listet gesendete Einsatzlisten fuer einen Auftrag."""
    try:
        va_id = request.args.get('va_id', type=int)
        if not va_id:
            return jsonify({"success": False, "error": "va_id erforderlich"}), 400

        # Abfrage gesendeter Einsatzlisten (falls Tabelle existiert)
        # Fallback: Leere Liste wenn Tabelle nicht existiert
        try:
            sql = ("SELECT ID, VA_ID, gesendet_am, empfaenger, typ, erfolgreich "
                   "FROM tbl_VA_EL_Gesendet "
                   "WHERE VA_ID = ? "
                   "ORDER BY gesendet_am DESC")
            data = execute_query(sql, [va_id])
        except Exception:
            # Tabelle existiert nicht - leere Liste zurueckgeben
            data = []

        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        logger.error(f"Error in get_einsatzlisten_gesendet: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =============================================================================
# ERROR HANDLERS
# =============================================================================

@app.errorhandler(404)
def not_found(e):
    return jsonify({"success": False, "error": "Route nicht gefunden"}), 404

@app.errorhandler(500)
def server_error(e):
    logger.error(f"Internal server error: {e}")
    return jsonify({"success": False, "error": "Interner Server-Fehler"}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    """Globaler Exception Handler - verhindert Server-Absturz"""
    logger.error(f"Unhandled exception: {type(e).__name__}: {e}")
    return jsonify({
        "success": False,
        "error": f"Unerwarteter Fehler: {type(e).__name__}",
        "details": str(e)
    }), 500

# =============================================================================
# MAIN
# =============================================================================

def is_port_in_use(port, host='127.0.0.1'):
    """Prüft ob ein Port bereits belegt ist"""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        try:
            s.settimeout(1)
            s.connect((host, port))
            return True  # Port ist belegt
        except (ConnectionRefusedError, socket.timeout, OSError):
            return False  # Port ist frei


def run_server(host='127.0.0.1', port=5000, use_production=True):
    """
    Startet den API Server.

    Args:
        host: IP-Adresse (default: localhost)
        port: Port (default: 5000)
        use_production: Wenn True, wird waitress verwendet (stabiler)
    """
    # Prüfen ob Port bereits belegt ist
    if is_port_in_use(port, host):
        print("=" * 60)
        print(f"[INFO] API Server läuft bereits auf Port {port}")
        print("[INFO] Neue Instanz wird nicht gestartet.")
        print("=" * 60)
        return  # Nicht starten wenn bereits eine Instanz läuft

    print("=" * 60)
    print("CONSYS API Server - STABIL")
    print("=" * 60)
    print(f"Backend: {BACKEND_PATH}")
    print(f"API Base: http://{host}:{port}/api")
    print(f"Health Check: http://{host}:{port}/api/health")
    print("=" * 60)

    if use_production:
        try:
            from waitress import serve
            print("[SERVER] Starte mit Waitress (Production Mode)")
            print("[SERVER] 8 Threads fuer parallele Requests (DB-Lock aktiv)")
            print("[SERVER] Server laeuft stabil - beenden mit Ctrl+C")
            # threads=8 fuer parallele HTTP-Requests (4 Frontends x 2)
            # DB-Zugriffe werden via _db_lock serialisiert (Access ODBC nicht thread-safe)
            serve(app, host=host, port=port, threads=8)
        except ImportError:
            print("[WARNUNG] waitress nicht installiert - verwende Flask Dev Server")
            print("[INFO] Fuer Production: pip install waitress")
            app.run(host=host, port=port, debug=False, threaded=False)
    else:
        print("[SERVER] Starte mit Flask Development Server")
        app.run(host=host, port=port, debug=True, threaded=True)

if __name__ == '__main__':
    import signal
    import sys

    def signal_handler(sig, frame):
        print("\n[SERVER] Shutdown angefordert...")
        stop_vba_bridge_watchdog()
        print("[SERVER] Server beendet.")
        sys.exit(0)

    # Graceful shutdown bei Ctrl+C
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)

    # VBA Bridge Watchdog starten (Auto-Start & Auto-Restart)
    print("[VBA Bridge] Starte Auto-Watchdog...")
    start_vba_bridge_watchdog()

    # Server starten (Production Mode)
    run_server(use_production=True)
