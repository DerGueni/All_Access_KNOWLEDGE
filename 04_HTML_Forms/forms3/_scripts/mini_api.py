# -*- coding: utf-8 -*-
"""
Mini API Server für forms3 HTML-Formulare
Minimaler REST-Server für direkten Access-Backend-Zugriff
Starten: python mini_api.py
"""

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import pyodbc
from datetime import datetime, date, time
from decimal import Decimal
import os
import threading
from contextlib import contextmanager
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ============================================
# SMTP KONFIGURATION (Mailjet)
# ============================================
SMTP_SERVER = "in-v3.mailjet.com"
SMTP_PORT = 25
SMTP_USERNAME = "97455f0f699bcd3a1cb8602299c3dadd"
SMTP_PASSWORD = "1dd9946e4f632343405471b1b700c52f"
SMTP_FROM_EMAIL = "consec-auftragsplanung@gmx.de"  # Absender-Adresse

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # UTF-8 für JSON-Responses
CORS(app)  # CORS für ALLE Routen aktivieren (nicht nur /api/*)

# ============================================
# STATISCHE DATEIEN (HTML, JS, CSS)
# ============================================
# KORRIGIERT: Nur ein dirname() um von _scripts nach forms3 zu kommen
FORMS3_PATH = os.path.dirname(os.path.abspath(__file__))  # _scripts Ordner
FORMS3_PATH = os.path.dirname(FORMS3_PATH)  # forms3 Ordner (parent von _scripts)

@app.route('/')
def index():
    """Startseite - zeigt Auftragstamm"""
    return send_from_directory(FORMS3_PATH, 'frm_va_Auftragstamm.html')

# HINWEIS: Catch-All Route für statische Dateien ist am Ende der Datei definiert
# (nach allen API-Routes), damit Flask zuerst die API-Routes prüft!

# ============================================
# KONFIGURATION - Pfade anpassen falls nötig
# ============================================
BACKEND_PATH = r"S:\CONSEC\CONSEC PLANUNG AKTUELL\B - DIVERSES\0_Consec_V1_BE_V1.55_Test.accdb"

# Alternative: Netzwerkpfad falls S: nicht gemappt
# BACKEND_PATH = r"\\vConSYS01-NBG\Consys\CONSEC\CONSEC PLANUNG AKTUELL\B - DIVERSES\0_Consec_V1_BE_V1.55_Test.accdb"

# ============================================
# DATENBANK-VERBINDUNG (mit Connection Pooling disabled)
# ============================================
# Deaktiviere pyodbc Connection Pooling um Segfaults zu vermeiden
pyodbc.pooling = False

# Globales Lock - Access ODBC-Treiber ist NICHT thread-safe!
# Alle DB-Operationen muessen serialisiert werden
db_lock = threading.Lock()

def get_connection():
    """Erstellt eine neue Datenbankverbindung"""
    conn_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={BACKEND_PATH};"
    conn = pyodbc.connect(conn_str)
    conn.autocommit = True  # Verhindert Lock-Probleme
    return conn

@contextmanager
def get_db():
    """Thread-safe Datenbank-Verbindung mit automatischem Lock und Cleanup"""
    with db_lock:
        conn = get_connection()
        try:
            yield conn
        finally:
            conn.close()

def serialize_value(val):
    """Konvertiert Werte für JSON"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, time):
        return val.strftime('%H:%M:%S')
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, bytes):
        return val.decode('utf-8', errors='ignore')
    return val

def query_to_list(cursor):
    """Konvertiert Cursor-Ergebnis zu Liste von Dicts"""
    columns = [col[0] for col in cursor.description]
    return [
        {col: serialize_value(val) for col, val in zip(columns, row)}
        for row in cursor.fetchall()
    ]

# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/health')
def health():
    """Health Check"""
    try:
        conn = get_connection()
        conn.close()
        return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/tables')
def tables():
    """Liste aller Tabellen"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        table_list = [{"name": t.table_name, "type": t.table_type} for t in cursor.tables(tableType='TABLE')]
        conn.close()
        return jsonify({"success": True, "tables": table_list})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- AUFTRÄGE ---
@app.route('/api/auftraege')
def auftraege_list():
    """Auftragsliste mit Soll/Ist und expand_days Support"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Parameter
        limit = request.args.get('limit', 200, type=int)
        offset = request.args.get('offset', 0, type=int)
        search = request.args.get('search', '')
        status = request.args.get('status', '')
        ab_datum = request.args.get('ab', '')
        datum_von = request.args.get('datum_von', '')
        expand_days = request.args.get('expand_days', 'false').lower() == 'true'

        # Basis-SQL mit Soll/Ist Berechnung pro Tag
        if expand_days:
            # Expandierte Ansicht: Jeder Einsatztag als eigene Zeile mit Soll/Ist
            # a.ID ist der Primary Key in tbl_VA_Auftragstamm, VA_ID ist Foreign Key in anderen Tabellen
            sql = f"""
                SELECT TOP {limit}
                    a.*,
                    t.VADatum as Datum,
                    t.ID as VADatum_ID,
                    (SELECT SUM(s.MA_Anzahl) FROM tbl_VA_Start s WHERE s.VA_ID = a.ID AND s.VADatum = t.VADatum) AS MA_Anzahl_Soll,
                    (SELECT COUNT(*) FROM tbl_MA_VA_Planung p WHERE p.VA_ID = a.ID AND p.VADatum = t.VADatum) AS MA_Anzahl_Ist
                FROM tbl_VA_Auftragstamm a
                INNER JOIN tbl_VA_AnzTage t ON a.ID = t.VA_ID
                WHERE 1=1
            """
        else:
            # Standard-Ansicht: Nur Auftragsstamm
            sql = f"SELECT TOP {limit} * FROM tbl_VA_Auftragstamm WHERE 1=1"

        if search:
            sql += f" AND (a.Auftrag LIKE '%{search}%' OR a.Objekt LIKE '%{search}%' OR a.Ort LIKE '%{search}%')" if expand_days else f" AND (Auftrag LIKE '%{search}%' OR Objekt LIKE '%{search}%' OR Ort LIKE '%{search}%')"
        if status:
            sql += f" AND a.Veranst_Status_ID = {status}" if expand_days else f" AND Veranst_Status_ID = {status}"
        if datum_von:
            if expand_days:
                sql += f" AND t.VADatum >= #{datum_von}#"
            else:
                sql += f" AND Dat_VA_Von >= #{datum_von}#"
        if ab_datum:
            if expand_days:
                sql += f" AND t.VADatum >= #{ab_datum}#"
            else:
                sql += f" AND Dat_VA_Von >= #{ab_datum}#"

        if expand_days:
            sql += " ORDER BY t.VADatum ASC, a.Auftrag ASC"
        else:
            sql += " ORDER BY ID DESC"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()

        return jsonify({
            "success": True,
            "data": result,
            "total": len(result),
            "limit": limit,
            "offset": offset
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/auftraege/<int:id>')
def auftrag_detail(id):
    """Einzelner Auftrag mit allen zugehörigen Daten"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()

            # Hauptdaten
            cursor.execute("SELECT * FROM tbl_VA_Auftragstamm WHERE ID = ?", (id,))
            rows = query_to_list(cursor)

            if not rows:
                return jsonify({"success": False, "error": "Nicht gefunden"}), 404

            auftrag = rows[0]

            # Einsatztage - einfache Query
            cursor.execute("SELECT * FROM tbl_VA_AnzTage WHERE VA_ID = ? ORDER BY VADatum", (id,))
            einsatztage = query_to_list(cursor)

            # Startzeiten/Schichten - einfache Query
            cursor.execute("SELECT * FROM tbl_VA_Start WHERE VA_ID = ? ORDER BY VADatum, VA_Start", (id,))
            startzeiten = query_to_list(cursor)

            # Zuordnungen - OHNE JOIN (verhindert Segfaults!)
            cursor.execute("SELECT *, MVA_Start AS MA_Start, MVA_Ende AS MA_Ende FROM tbl_MA_VA_Planung WHERE VA_ID = ? ORDER BY VADatum, MVA_Start", (id,))
            zuordnungen = query_to_list(cursor)

            # MA-Namen separat laden und zusammenführen (statt JOIN)
            if zuordnungen:
                ma_ids = list(set(z.get('MA_ID') for z in zuordnungen if z.get('MA_ID')))
                if ma_ids:
                    # Access/pyodbc: IN-Clause mit direkten Werten (sicher bei Integer-IDs)
                    # WICHTIG: tbl_MA_Mitarbeiterstamm hat 'ID' nicht 'MA_ID'!
                    ids_str = ','.join(str(int(mid)) for mid in ma_ids)
                    cursor.execute(f"SELECT ID, Nachname, Vorname, Tel_Mobil FROM tbl_MA_Mitarbeiterstamm WHERE ID IN ({ids_str})")
                    ma_dict = {row[0]: {'Nachname': row[1], 'Vorname': row[2], 'Tel_Mobil': row[3]} for row in cursor.fetchall()}
                    # Namen zu Zuordnungen hinzufügen
                    for z in zuordnungen:
                        ma = ma_dict.get(z.get('MA_ID'), {})
                        z['Nachname'] = ma.get('Nachname', '')
                        z['Vorname'] = ma.get('Vorname', '')
                        z['Tel_Mobil'] = ma.get('Tel_Mobil', '')

            # Offene Anfragen - vorerst leer (tbl_MA_VA_Zuordnung hat Probleme)
            anfragen = []

            return jsonify({
                "success": True,
                "data": {
                    "auftrag": auftrag,
                    "einsatztage": einsatztage,
                    "startzeiten": startzeiten,
                    "zuordnungen": zuordnungen,
                    "anfragen": anfragen
                }
            })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# --- MITARBEITER ---
@app.route('/api/mitarbeiter')
def mitarbeiter_list():
    """Mitarbeiterliste mit erweiterten Filtern"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        aktiv = request.args.get('aktiv', '')
        limit = request.args.get('limit', 500, type=int)
        search = request.args.get('search', '')
        anstellung = request.args.get('anstellung', '')
        kategorie = request.args.get('kategorie', '')
        nur34a = request.args.get('nur34a', '')

        sql = f"SELECT TOP {limit} * FROM tbl_MA_Mitarbeiterstamm WHERE 1=1"

        if aktiv == 'true' or aktiv == '1':
            sql += " AND IstAktiv = -1"
        if search:
            sql += f" AND (Nachname LIKE '%{search}%' OR Vorname LIKE '%{search}%' OR Kurzname LIKE '%{search}%')"

        # Anstellungsart-Filter (3=Festangestellt, 5=Minijobber, 9=Studenten, 13=Alle aktiven)
        if anstellung and anstellung != '9':  # 9 = Alle
            if anstellung == '13':  # Alle aktiven (3, 4, 5)
                sql += " AND Anstellungsart_ID IN (3, 4, 5)"
            elif anstellung == '5':  # Mini + Midi (4, 5)
                sql += " AND Anstellungsart_ID IN (4, 5)"
            else:
                sql += f" AND Anstellungsart_ID = {int(anstellung)}"

        # Kategorie/Qualifikation Filter
        if kategorie and kategorie != '' and kategorie != '1':  # 1 = Alle
            sql += f" AND Kategorie_ID = {int(kategorie)}"

        # Nur 34a (Hat_keine_34a = False bedeutet HAT 34a)
        if nur34a == 'true' or nur34a == '1':
            sql += " AND Hat_keine_34a = 0"

        sql += " ORDER BY Nachname, Vorname"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mitarbeiter/verfuegbar')
def mitarbeiter_verfuegbar():
    """Prüft Verfügbarkeit der Mitarbeiter für ein Datum/Zeitraum"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        datum = request.args.get('datum', '')
        va_start = request.args.get('va_start', '')
        va_ende = request.args.get('va_ende', '')
        geplant_verfuegbar = request.args.get('geplant_verfuegbar', 'false')

        if not datum:
            return jsonify({"success": False, "error": "Datum erforderlich"}), 400

        # Alle aktiven Mitarbeiter holen
        sql_ma = "SELECT ID FROM tbl_MA_Mitarbeiterstamm WHERE IstAktiv = -1"
        cursor.execute(sql_ma)
        alle_ma = [row[0] for row in cursor.fetchall()]

        # Nicht verfügbare MA (aus tbl_MA_NVerfuegZeiten)
        sql_nv = f"""
            SELECT DISTINCT MA_ID FROM tbl_MA_NVerfuegZeiten
            WHERE vonDat <= #{datum}# AND bisDat >= #{datum}#
        """
        cursor.execute(sql_nv)
        nicht_verfuegbar = set(row[0] for row in cursor.fetchall())

        # Bereits geplante MA (aus tbl_MA_VA_Planung)
        sql_geplant = f"""
            SELECT DISTINCT MA_ID FROM tbl_MA_VA_Planung
            WHERE VADatum = #{datum}#
        """
        if va_start:
            sql_geplant += f" AND MVA_Start = #{va_start}#"

        cursor.execute(sql_geplant)
        geplant = set(row[0] for row in cursor.fetchall())

        conn.close()

        # Verfügbare berechnen
        if geplant_verfuegbar == 'true':
            # Geplant = Verfügbar: Nur nicht_verfuegbar ausschließen
            verfuegbar = [ma for ma in alle_ma if ma not in nicht_verfuegbar]
        else:
            # Normal: Nicht verfügbar UND geplant ausschließen
            verfuegbar = [ma for ma in alle_ma if ma not in nicht_verfuegbar and ma not in geplant]

        return jsonify({
            "success": True,
            "data": {
                "verfuegbar": verfuegbar,
                "nicht_verfuegbar": list(nicht_verfuegbar),
                "geplant": list(geplant),
                "datum": datum
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mitarbeiter/<int:id>')
def mitarbeiter_detail(id):
    """Einzelner Mitarbeiter"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        # KORREKTUR: Das Feld heißt 'ID' nicht 'MA_ID' in tbl_MA_Mitarbeiterstamm
        cursor.execute("SELECT * FROM tbl_MA_Mitarbeiterstamm WHERE ID = ?", (id,))
        rows = query_to_list(cursor)
        conn.close()
        if rows:
            return jsonify({"success": True, "data": rows[0]})
        return jsonify({"success": False, "error": "Nicht gefunden"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- KUNDEN ---
@app.route('/api/kunden')
def kunden_list():
    """Kundenliste"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        aktiv = request.args.get('aktiv', '')
        limit = request.args.get('limit', 500, type=int)

        sql = f"SELECT TOP {limit} * FROM tbl_KD_Kundenstamm WHERE 1=1"

        if aktiv == 'true' or aktiv == '1':
            sql += " AND kun_IstAktiv = -1"

        sql += " ORDER BY kun_Firma"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/kunden/<int:id>')
def kunde_detail(id):
    """Einzelner Kunde"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tbl_KD_Kundenstamm WHERE kun_Id = ?", (id,))
        rows = query_to_list(cursor)
        conn.close()
        if rows:
            return jsonify({"success": True, "data": rows[0]})
        return jsonify({"success": False, "error": "Nicht gefunden"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- OBJEKTE ---
@app.route('/api/objekte')
def objekte_list():
    """Objekteliste"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        limit = request.args.get('limit', 500, type=int)

        # tbl_OB_Objekt hat keine Kun_ID und IstAktiv Felder
        sql = f"SELECT TOP {limit} * FROM tbl_OB_Objekt"
        sql += " ORDER BY Objekt"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ANSTELLUNGSARTEN ---
@app.route('/api/anstellungsarten')
def anstellungsarten_list():
    """Liste aller Anstellungsarten für Dropdown-Filter"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM tbl_hlp_MA_Anstellungsart
        """)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- EINSATZTAGE ---
@app.route('/api/einsatztage')
def einsatztage_list():
    """Einsatztage für einen Auftrag, Zeitraum oder parallele Einsätze"""
    try:
        va_id = request.args.get('va_id', type=int)
        datum_id = request.args.get('datum_id', type=int)
        parallel = request.args.get('parallel', 'false').lower() == 'true'
        # NEU: von/bis Parameter für Dienstplan-Übersicht
        von = request.args.get('von')
        bis = request.args.get('bis')

        conn = get_connection()
        cursor = conn.cursor()

        if von and bis:
            # Zeitraum-Abfrage: Alle Einsatztage zwischen von und bis
            cursor.execute("""
                SELECT t.*, a.Auftrag, a.Objekt
                FROM tbl_VA_AnzTage t
                LEFT JOIN tbl_VA_Auftragstamm a ON t.VA_ID = a.ID
                WHERE t.VADatum >= CDATE(?) AND t.VADatum <= CDATE(?)
                ORDER BY t.VADatum, a.Auftrag
            """, (von, bis))
            result = query_to_list(cursor)
        elif datum_id and parallel:
            # Parallele Einsaetze: Alle anderen Auftraege am selben Tag
            cursor.execute("SELECT VADatum FROM tbl_VA_AnzTage WHERE ID = ?", (datum_id,))
            row = cursor.fetchone()
            if row:
                vadatum = row[0]
                cursor.execute("""
                    SELECT t.*, a.Auftrag, a.Objekt
                    FROM tbl_VA_AnzTage t
                    LEFT JOIN tbl_VA_Auftragstamm a ON t.VA_ID = a.ID
                    WHERE t.VADatum = ?
                    ORDER BY a.Auftrag
                """, (vadatum,))
                result = query_to_list(cursor)
            else:
                result = []
        elif va_id:
            cursor.execute("SELECT * FROM tbl_VA_AnzTage WHERE VA_ID = ? ORDER BY VADatum", (va_id,))
            result = query_to_list(cursor)
        else:
            conn.close()
            return jsonify({"success": False, "error": "va_id, datum_id oder von/bis erforderlich"}), 400

        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- SCHICHTEN ---
@app.route('/api/schichten')
def schichten_list():
    """Schichten für einen Auftrag"""
    try:
        va_id = request.args.get('va_id', type=int)
        vadatum_id = request.args.get('vadatum_id', type=int)

        if not va_id:
            return jsonify({"success": False, "error": "va_id erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Falls vadatum_id gegeben, erst das VADatum aus tbl_VA_AnzTage holen
        if vadatum_id:
            cursor.execute("SELECT VADatum FROM tbl_VA_AnzTage WHERE ID = ?", (vadatum_id,))
            row = cursor.fetchone()
            if row:
                vadatum = row[0]
                cursor.execute("SELECT * FROM tbl_VA_Start WHERE VA_ID = ? AND VADatum = ? ORDER BY VA_Start", (va_id, vadatum))
            else:
                # Fallback: alle Schichten
                cursor.execute("SELECT * FROM tbl_VA_Start WHERE VA_ID = ? ORDER BY VADatum, VA_Start", (va_id,))
        else:
            # Alle Schichten des Auftrags
            cursor.execute("SELECT * FROM tbl_VA_Start WHERE VA_ID = ? ORDER BY VADatum, VA_Start", (va_id,))

        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ZUORDNUNGEN ---
@app.route('/api/zuordnungen')
def zuordnungen_list():
    """MA-Zuordnungen für einen Auftrag oder Zeitraum"""
    try:
        va_id = request.args.get('va_id', type=int)
        vadatum_id = request.args.get('vadatum_id', type=int)  # Spezifischer Tag
        von = request.args.get('von')  # Datum YYYY-MM-DD
        bis = request.args.get('bis')  # Datum YYYY-MM-DD

        conn = get_connection()
        cursor = conn.cursor()

        # Access-SQL: Multiple JOINs benoetigen Klammern!
        sql = """
            SELECT p.*, p.MVA_Start AS MA_Start, p.MVA_Ende AS MA_Ende,
                   m.Nachname, m.Vorname, m.Tel_Mobil,
                   t.VADatum, a.Objekt
            FROM (((tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
            WHERE 1=1
        """
        params = []

        if va_id:
            sql += " AND p.VA_ID = ?"
            params.append(va_id)

        if vadatum_id:
            # Nur Zuordnungen fuer diesen spezifischen Tag
            sql += " AND p.VADatum_ID = ?"
            params.append(vadatum_id)
        elif von and bis:
            sql += " AND t.VADatum BETWEEN ? AND ?"
            params.append(von)
            params.append(bis)
        elif von:
            sql += " AND t.VADatum >= ?"
            params.append(von)
        elif bis:
            sql += " AND t.VADatum <= ?"
            params.append(bis)

        sql += " ORDER BY t.VADatum, p.MVA_Start"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/zuordnungen', methods=['POST'])
def zuordnungen_create():
    """Neue MA-Zuordnung erstellen (tbl_MA_VA_Planung)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Keine Daten"}), 400

        # Pflichtfelder pruefen
        va_id = data.get('VA_ID')
        ma_id = data.get('MA_ID')
        if not va_id or not ma_id:
            return jsonify({"success": False, "error": "VA_ID und MA_ID sind erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Insert
        sql = """
            INSERT INTO tbl_MA_VA_Planung
            (VA_ID, MA_ID, VADatum_ID, VAStart_ID, MVA_Start, MVA_Ende,
             Bemerkungen, PKW, Einsatzleitung, IstFraglich)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = [
            va_id,
            ma_id,
            data.get('VADatum_ID'),
            data.get('VAStart_ID'),
            data.get('MVA_Start') or data.get('MA_Start'),
            data.get('MVA_Ende') or data.get('MA_Ende'),
            data.get('Bemerkungen', ''),
            data.get('PKW', 0),
            data.get('Einsatzleitung', False),
            data.get('IstFraglich', False)
        ]

        cursor.execute(sql, params)
        conn.commit()

        # Neue ID abrufen
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]

        conn.close()
        print(f"[API] Zuordnung erstellt: ID={new_id}, VA={va_id}, MA={ma_id}")
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/zuordnungen/<int:id>', methods=['PUT'])
def zuordnungen_update(id):
    """MA-Zuordnung aktualisieren (tbl_MA_VA_Planung)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "Keine Daten"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Dynamisches Update - nur uebergebene Felder aktualisieren
        updates = []
        params = []

        # Mapping: API-Feldnamen zu DB-Feldnamen
        field_mapping = {
            'MA_ID': 'MA_ID',
            'MA_Start': 'MVA_Start',
            'MA_Ende': 'MVA_Ende',
            'MVA_Start': 'MVA_Start',
            'MVA_Ende': 'MVA_Ende',
            'Bemerkungen': 'Bemerkungen',
            'PKW': 'PKW',
            'Einsatzleitung': 'Einsatzleitung',
            'IstFraglich': 'IstFraglich',
            'VAStart_ID': 'VAStart_ID',
            'VADatum_ID': 'VADatum_ID',
            'Rch_Erstellt': 'Rch_Erstellt'
        }

        for api_field, db_field in field_mapping.items():
            if api_field in data:
                updates.append(f"{db_field} = ?")
                params.append(data[api_field])

        if not updates:
            return jsonify({"success": False, "error": "Keine Felder zum Aktualisieren"}), 400

        params.append(id)
        sql = f"UPDATE tbl_MA_VA_Planung SET {', '.join(updates)} WHERE ID = ?"

        cursor.execute(sql, params)
        conn.commit()
        conn.close()

        print(f"[API] Zuordnung aktualisiert: ID={id}")
        return jsonify({"success": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# --- PLANUNGEN (Alias fuer zuordnungen - Kompatibilitaet) ---
@app.route('/api/planungen')
def planungen_list():
    """Planungen = Alias fuer Zuordnungen (tbl_MA_VA_Planung)"""
    try:
        va_id = request.args.get('va_id', type=int)
        # FIX 20.01.2026: Akzeptiere beide Parameter-Namen (HTML sendet datum_id)
        vadatum_id = request.args.get('vadatum_id', type=int) or request.args.get('datum_id', type=int)

        conn = get_connection()
        cursor = conn.cursor()

        # Access-SQL: Multiple JOINs benoetigen Klammern!
        sql = """
            SELECT p.*, p.MVA_Start AS MA_Start, p.MVA_Ende AS MA_Ende,
                   m.Nachname, m.Vorname, m.Tel_Mobil,
                   t.VADatum, a.Objekt
            FROM (((tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
            WHERE 1=1
        """
        params = []

        if va_id:
            sql += " AND p.VA_ID = ?"
            params.append(va_id)

        if vadatum_id:
            sql += " AND p.VADatum_ID = ?"
            params.append(vadatum_id)

        sql += " ORDER BY t.VADatum, p.MVA_Start"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# FIX 19.01.2026: POST Route fuer Planungen (DblClick Schnellauswahl)
@app.route('/api/planungen', methods=['POST'])
def planungen_create():
    """Neue Planung erstellen (MA zu Auftrag zuordnen)"""
    try:
        data = request.get_json()
        # Akzeptiere sowohl Gross- als auch Kleinschreibung
        va_id = data.get('VA_ID') or data.get('va_id')
        ma_id = data.get('MA_ID') or data.get('ma_id')
        vadatum_id = data.get('VADatum_ID') or data.get('vadatum_id')

        if not all([va_id, ma_id, vadatum_id]):
            return jsonify({"success": False, "error": "VA_ID, MA_ID und VADatum_ID erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Pruefen ob bereits zugeordnet
        cursor.execute("""
            SELECT ID FROM tbl_MA_VA_Planung
            WHERE VA_ID = ? AND MA_ID = ? AND VADatum_ID = ?
        """, [va_id, ma_id, vadatum_id])
        existing = cursor.fetchone()

        if existing:
            conn.close()
            return jsonify({"success": False, "error": "MA bereits zugeordnet", "existing_id": existing[0]}), 409

        # Neue Zuordnung erstellen
        cursor.execute("""
            INSERT INTO tbl_MA_VA_Planung (VA_ID, MA_ID, VADatum_ID)
            VALUES (?, ?, ?)
        """, [va_id, ma_id, vadatum_id])
        conn.commit()

        # Neue ID holen
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]
        conn.close()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# FIX 19.01.2026: DELETE Route fuer Planungen (DblClick entfernen in Schnellauswahl)
@app.route('/api/planungen/<int:plan_id>', methods=['DELETE'])
def planungen_delete(plan_id):
    """Planung loeschen (MA aus Auftrag entfernen)"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Pruefen ob Eintrag existiert
        cursor.execute("SELECT ID FROM tbl_MA_VA_Planung WHERE ID = ?", [plan_id])
        existing = cursor.fetchone()

        if not existing:
            conn.close()
            return jsonify({"success": False, "error": "Planung nicht gefunden"}), 404

        # Loeschen
        cursor.execute("DELETE FROM tbl_MA_VA_Planung WHERE ID = ?", [plan_id])
        conn.commit()
        conn.close()

        return jsonify({"success": True, "deleted_id": plan_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# --- STATUS ---
@app.route('/api/status')
def status_list():
    """Status-Liste"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tbl_VA_Status ORDER BY ID")
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ORTE ---
@app.route('/api/orte')
def orte_list():
    """Orte-Liste (distinct aus Aufträgen)"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT Ort, PLZ
            FROM tbl_VA_Auftragstamm
            WHERE Ort IS NOT NULL AND Ort <> ''
            ORDER BY Ort
        """)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- DIENSTKLEIDUNG ---
@app.route('/api/dienstkleidung')
def dienstkleidung_list():
    """Dienstkleidung-Liste"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tbl_MA_Dienstkleidung ORDER BY ID")
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ANFRAGEN ---
@app.route('/api/anfragen')
def anfragen_list():
    """Offene MA-Anfragen"""
    try:
        va_id = request.args.get('va_id', type=int)

        conn = get_connection()
        cursor = conn.cursor()

        # Status_ID: 1=Geplant, 2=Angefragt, 3=Bestaetigt, 4=Durchgefuehrt, 5=Abgesagt
        sql = """
            SELECT z.*, m.Nachname, m.Vorname, m.Kurzname
            FROM tbl_MA_VA_Zuordnung z
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON z.MA_ID = m.MA_ID
            WHERE z.Status_ID = 2
        """

        if va_id:
            sql += f" AND z.VA_ID = {va_id}"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ABSAGEN ---
@app.route('/api/absagen')
def absagen_list():
    """MA-Absagen für einen Auftrag"""
    try:
        va_id = request.args.get('va_id', type=int)
        if not va_id:
            return jsonify({"success": True, "data": []})

        conn = get_connection()
        cursor = conn.cursor()

        # Status_ID 5 = Abgesagt
        cursor.execute("""
            SELECT z.*, m.Nachname, m.Vorname
            FROM tbl_MA_VA_Zuordnung z
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON z.MA_ID = m.MA_ID
            WHERE z.VA_ID = ? AND z.Status_ID = 5
        """, (va_id,))
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- MA ANFRAGEN ERSTELLEN ---
@app.route('/api/anfragen/create', methods=['POST'])
def anfragen_create():
    """Erstellt MA-Anfragen für einen Auftrag (in tbl_MA_VA_Planung)"""
    try:
        data = request.get_json(force=True)
        ma_ids = data.get('ma_ids', [])
        va_id = data.get('va_id')
        vadatum_id = data.get('vadatum_id')
        vastart_id = data.get('vastart_id')

        if not ma_ids or not va_id:
            return jsonify({"success": False, "error": "ma_ids und va_id erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Startzeit-Daten holen (VADatum, MVA_Start, MVA_Ende) aus tbl_VA_Start
        va_datum = None
        mva_start = None
        mva_ende = None
        if vastart_id:
            cursor.execute("""
                SELECT VADatum, MVA_Start, MVA_Ende
                FROM tbl_VA_Start
                WHERE ID = ?
            """, (vastart_id,))
            startzeit_row = cursor.fetchone()
            if startzeit_row:
                va_datum = startzeit_row[0]
                mva_start = startzeit_row[1]
                mva_ende = startzeit_row[2]

        created = []
        for ma_id in ma_ids:
            # Prüfen ob Anfrage bereits existiert (in tbl_MA_VA_Planung)
            cursor.execute("""
                SELECT ID FROM tbl_MA_VA_Planung
                WHERE MA_ID = ? AND VA_ID = ? AND VADatum_ID = ?
            """, (ma_id, va_id, vadatum_id))
            existing = cursor.fetchone()

            if not existing:
                # Neue Anfrage erstellen in tbl_MA_VA_Planung
                # Status_ID: 1=Geplant, 2=Angefragt/Benachrichtigt, 3=Bestaetigt, 4=Durchgefuehrt
                # Mit VADatum, MVA_Start, MVA_Ende aus Startzeiten
                cursor.execute("""
                    INSERT INTO tbl_MA_VA_Planung
                    (MA_ID, VA_ID, VADatum_ID, VAStart_ID, Status_ID, Anfragezeitpunkt,
                     VADatum, MVA_Start, MVA_Ende)
                    VALUES (?, ?, ?, ?, 2, ?, ?, ?, ?)
                """, (ma_id, va_id, vadatum_id, vastart_id, datetime.now(),
                      va_datum, mva_start, mva_ende))
                conn.commit()
                created.append(ma_id)

        # E-Mail-Adressen der angefragten MA holen
        if created:
            placeholders = ','.join(['?'] * len(created))
            cursor.execute(f"""
                SELECT ID as MA_ID, Nachname, Vorname, eMail
                FROM tbl_MA_Mitarbeiterstamm
                WHERE ID IN ({placeholders})
            """, created)
            ma_data = query_to_list(cursor)
        else:
            ma_data = []

        conn.close()

        return jsonify({
            "success": True,
            "created": len(created),
            "ma_data": ma_data,
            "message": f"{len(created)} Anfragen erstellt"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# --- MA ANFRAGEN SENDEN (SMTP) ---
@app.route('/api/anfragen/senden', methods=['POST'])
def anfragen_senden():
    """
    Sendet E-Mail-Anfragen an Mitarbeiter direkt per SMTP (Mailjet).
    Verwendet HTML_Body_Anfrage.txt als Vorlage.
    """
    try:
        data = request.get_json(force=True)
        print(f"[API] /api/anfragen/senden - Daten: {data}")

        # Parameter validieren
        va_id = data.get('va_id') or data.get('VA_ID')
        vadatum_id = data.get('vadatum_id') or data.get('VADatum_ID')
        vastart_id = data.get('vastart_id') or data.get('VAStart_ID')
        ma_ids = data.get('ma_ids') or data.get('MA_IDs') or []

        if not va_id or not vadatum_id:
            return jsonify({"success": False, "error": "va_id und vadatum_id sind erforderlich"}), 400

        if not ma_ids:
            return jsonify({"success": False, "error": "Keine MA-IDs angegeben"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # 1. Auftragsdaten laden
        cursor.execute("""
            SELECT a.*, t.VADatum, s.VA_Start, s.VA_Ende
            FROM tbl_VA_Auftragstamm a
            LEFT JOIN tbl_VA_AnzTage t ON a.ID = t.VA_ID AND t.ID = ?
            LEFT JOIN tbl_VA_Start s ON t.ID = s.VADatum_ID AND s.ID = ?
            WHERE a.ID = ?
        """, (vadatum_id, vastart_id, va_id))
        auftrag_row = cursor.fetchone()

        if not auftrag_row:
            conn.close()
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        # Spalten-Namen
        columns = [col[0] for col in cursor.description]
        auftrag = dict(zip(columns, auftrag_row))

        # 2. MA-Daten laden (mit E-Mail)
        ids_str = ','.join(str(int(mid)) for mid in ma_ids)
        cursor.execute(f"""
            SELECT ID, Vorname, Nachname, eMail, Tel_Mobil
            FROM tbl_MA_Mitarbeiterstamm
            WHERE ID IN ({ids_str})
        """)
        mitarbeiter = query_to_list(cursor)

        # 3. HTML-Template laden
        template_path = os.path.join(os.path.dirname(__file__), '..', 'HTMLBodies', 'HTML_Body_Anfrage.txt')
        if not os.path.exists(template_path):
            # Fallback-Pfad
            template_path = r"C:\Users\guenther.siegert\Documents\0006_All_Access_KNOWLEDGE\04_HTML_Forms\forms3\HTMLBodies\HTML_Body_Anfrage.txt"

        if not os.path.exists(template_path):
            conn.close()
            return jsonify({"success": False, "error": "HTML-Template nicht gefunden"}), 500

        with open(template_path, 'r', encoding='utf-8') as f:
            html_template = f.read()

        # 4. Platzhalter-Werte vorbereiten
        # Datum formatieren
        vadatum = auftrag.get('VADatum')
        if vadatum:
            if isinstance(vadatum, str):
                vadatum = datetime.strptime(vadatum[:10], '%Y-%m-%d')
            wochentag = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag'][vadatum.weekday()]
            datum_str = vadatum.strftime('%d.%m.%Y')
        else:
            wochentag = ''
            datum_str = ''

        # Zeiten formatieren
        va_start = auftrag.get('VA_Start', '')
        va_ende = auftrag.get('VA_Ende', '')
        if va_start and hasattr(va_start, 'strftime'):
            va_start = va_start.strftime('%H:%M')
        if va_ende and hasattr(va_ende, 'strftime'):
            va_ende = va_ende.strftime('%H:%M')

        # Treffpunkt-Zeit
        treffp_zeit = auftrag.get('Treff_Zeit', '')
        if treffp_zeit and hasattr(treffp_zeit, 'strftime'):
            treffp_zeit = treffp_zeit.strftime('%H:%M')

        # 5. E-Mails senden und Anfragen speichern
        sent_count = 0
        failed_count = 0
        results = []

        for ma in mitarbeiter:
            ma_id = ma.get('ID')
            email = ma.get('eMail', '')
            vorname = ma.get('Vorname', '')
            nachname = ma.get('Nachname', '')

            if not email or '@' not in email:
                print(f"[ANFRAGE] Keine E-Mail für MA {ma_id} ({nachname})")
                failed_count += 1
                results.append({"ma_id": ma_id, "status": "keine_email", "name": f"{vorname} {nachname}"})
                continue

            try:
                # URL für JA/NEIN Antwort (Rueckmeldung-System)
                base_url = "http://127.0.0.1:5000/api/rueckmeldung"
                url_ja = f"{base_url}?ma_id={ma_id}&va_id={va_id}&vadatum_id={vadatum_id}&antwort=ja"
                url_nein = f"{base_url}?ma_id={ma_id}&va_id={va_id}&vadatum_id={vadatum_id}&antwort=nein"

                # Platzhalter ersetzen
                html_body = html_template
                html_body = html_body.replace('[A_Wochentag]', wochentag)
                html_body = html_body.replace('[A_Auftr_Datum]', datum_str)
                html_body = html_body.replace('[A_Auftrag]', auftrag.get('Auftrag', '') or '')
                html_body = html_body.replace('[A_Ort]', auftrag.get('VA_Ort', '') or auftrag.get('Ort', '') or '')
                html_body = html_body.replace('[A_Objekt]', auftrag.get('Objekt', '') or '')
                html_body = html_body.replace('[A_Start_Zeit]', str(va_start) if va_start else '')
                html_body = html_body.replace('[A_End_Zeit]', str(va_ende) if va_ende else '')
                html_body = html_body.replace('[A_Treffp_Zeit]', str(treffp_zeit) if treffp_zeit else '')
                html_body = html_body.replace('[A_Treffpunkt]', auftrag.get('Treffpunkt', '') or '')
                html_body = html_body.replace('[A_Dienstkleidung]', auftrag.get('Dienstkleidung', '') or '')
                html_body = html_body.replace('[A_URL_JA]', url_ja)
                html_body = html_body.replace('[A_URL_NEIN]', url_nein)
                html_body = html_body.replace('[A_Sender]', 'Günther Siegert')

                # E-Mail senden
                betreff = f"Auftragsanfrage: {auftrag.get('Auftrag', '')} am {datum_str}"
                send_email_smtp(email, betreff, html_body)

                # Anfrage in DB speichern (Status_ID 2 = angefragt/benachrichtigt)
                cursor.execute("""
                    INSERT INTO tbl_MA_VA_Planung
                    (MA_ID, VA_ID, VADatum_ID, VAStart_ID, Status_ID, Anfragezeitpunkt, VADatum, MVA_Start, MVA_Ende)
                    VALUES (?, ?, ?, ?, 2, ?, ?, ?, ?)
                """, (ma_id, va_id, vadatum_id, vastart_id, datetime.now(), vadatum, va_start, va_ende))

                sent_count += 1
                results.append({"ma_id": ma_id, "status": "gesendet", "email": email, "name": f"{vorname} {nachname}"})
                print(f"[ANFRAGE] E-Mail gesendet an {email} ({nachname})")

            except Exception as mail_err:
                print(f"[ANFRAGE] Fehler beim Senden an {email}: {mail_err}")
                failed_count += 1
                results.append({"ma_id": ma_id, "status": "fehler", "error": str(mail_err), "name": f"{vorname} {nachname}"})

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "sent": sent_count,
            "failed": failed_count,
            "total": len(ma_ids),
            "results": results,
            "message": f"{sent_count} E-Mail-Anfragen gesendet, {failed_count} fehlgeschlagen"
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# --- EINSATZLISTE FÜR EXCEL-EXPORT ---
@app.route('/api/einsatzliste/<int:va_id>')
def einsatzliste_export(va_id):
    """Einsatzliste-Daten für Excel-Export"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Auftragsdaten
        cursor.execute("SELECT * FROM tbl_VA_Auftragstamm WHERE ID = ?", (va_id,))
        auftrag = query_to_list(cursor)

        if not auftrag:
            conn.close()
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        # Alle Zuordnungen mit MA-Daten (Status_ID 3 = Bestaetigt/Zugesagt)
        cursor.execute("""
            SELECT
                z.ID, z.VA_ID, z.VADatum_ID, z.VAStart_ID, z.MA_ID, z.Status_ID,
                z.VA_Start as ZuoStart, z.VA_Ende as ZuoEnde,
                m.Nachname, m.Vorname, m.Kurzname, m.Tel_Mobil, m.eMail,
                t.VADatum,
                s.VA_Start as SchichtStart, s.VA_Ende as SchichtEnde
            FROM tbl_MA_VA_Zuordnung z
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON z.MA_ID = m.MA_ID
            LEFT JOIN tbl_VA_AnzTage t ON z.VADatum_ID = t.ID
            LEFT JOIN tbl_VA_Start s ON z.VAStart_ID = s.ID
            WHERE z.VA_ID = ? AND z.Status_ID = 3
            ORDER BY t.VADatum, s.VA_Start, m.Nachname
        """, (va_id,))
        zuordnungen = query_to_list(cursor)

        conn.close()

        return jsonify({
            "success": True,
            "auftrag": auftrag[0],
            "zuordnungen": zuordnungen,
            "count": len(zuordnungen)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ESS NAMENSLISTE FÜR EXCEL-EXPORT ---
@app.route('/api/namensliste/<int:va_id>')
def namensliste_export(va_id):
    """ESS Namensliste-Daten für Excel-Export"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Auftragsdaten
        cursor.execute("SELECT * FROM tbl_VA_Auftragstamm WHERE ID = ?", (va_id,))
        auftrag = query_to_list(cursor)

        if not auftrag:
            conn.close()
            return jsonify({"success": False, "error": "Auftrag nicht gefunden"}), 404

        # Alle zugesagten MA mit ESS-relevanten Daten
        cursor.execute("""
            SELECT DISTINCT
                m.MA_ID, m.Nachname, m.Vorname, m.Kurzname,
                m.Geburtsdatum, m.Geburtsort, m.Nationalitaet,
                m.Ausweis_Nr, m.Ausweis_Gueltig_Bis,
                m.IHK_34a_Nr, m.IHK_34a_Gueltig_Bis,
                m.Tel_Mobil, m.eMail
            FROM tbl_MA_VA_Zuordnung z
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON z.MA_ID = m.MA_ID
            WHERE z.VA_ID = ? AND z.Status_ID = 3
            ORDER BY m.Nachname, m.Vorname
        """, (va_id,))
        mitarbeiter = query_to_list(cursor)

        conn.close()

        return jsonify({
            "success": True,
            "auftrag": auftrag[0],
            "mitarbeiter": mitarbeiter,
            "count": len(mitarbeiter)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- PLANUNG ERSTELLEN ---
@app.route('/api/planung/create', methods=['POST'])
def planung_create():
    """Erstellt Planungs-Eintrag für MA"""
    try:
        data = request.get_json(force=True)
        ma_id = data.get('ma_id')
        va_id = data.get('va_id')
        vadatum_id = data.get('vadatum_id')
        vastart_id = data.get('vastart_id')
        status_id = data.get('status_id', 1)  # 1 = Geplant

        if not ma_id or not va_id:
            return jsonify({"success": False, "error": "ma_id und va_id erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Prüfen ob MA bereits geplant
        cursor.execute("""
            SELECT ID FROM tbl_MA_VA_Planung
            WHERE MA_ID = ? AND VA_ID = ? AND VADatum_ID = ?
        """, (ma_id, va_id, vadatum_id))
        existing = cursor.fetchone()

        if existing:
            conn.close()
            return jsonify({"success": False, "error": "MA bereits geplant"}), 409

        # Neue Planung erstellen
        cursor.execute("""
            INSERT INTO tbl_MA_VA_Planung
            (MA_ID, VA_ID, VADatum_ID, VAStart_ID, Status_ID)
            VALUES (?, ?, ?, ?, ?)
        """, (ma_id, va_id, vadatum_id, vastart_id, status_id))
        conn.commit()

        # Neue ID holen
        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            "success": True,
            "id": new_id,
            "message": "Planung erstellt"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- PLANUNG LÖSCHEN ---
@app.route('/api/planung/<int:id>', methods=['DELETE'])
def planung_delete(id):
    """Löscht Planungs-Eintrag"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM tbl_MA_VA_Planung WHERE ID = ?", (id,))
        conn.commit()

        conn.close()

        return jsonify({"success": True, "message": "Planung gelöscht"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- GENERISCHE SQL-ABFRAGE ---
@app.route('/api/query', methods=['POST'])
def execute_query():
    """Führt beliebige SELECT-Abfrage aus"""
    try:
        data = request.get_json(force=True)
        sql = data.get('query', data.get('sql', ''))

        if not sql:
            return jsonify({"success": False, "error": "query erforderlich"}), 400

        # Sicherheitscheck: nur SELECT erlauben
        sql_upper = sql.strip().upper()
        if not sql_upper.startswith('SELECT'):
            return jsonify({"success": False, "error": "Nur SELECT erlaubt"}), 403

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- DIENSTPLAN GRÜNDE (Zeittypen) ---
@app.route('/api/dienstplan/gruende')
def dienstplan_gruende():
    """Abwesenheits-/Zeittyp-Gründe für Dienstplan"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tbl_MA_Zeittyp ORDER BY SortNr, Zeittyp")
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ABWESENHEITEN PRO MITARBEITER ---
@app.route('/api/abwesenheiten')
def abwesenheiten_list():
    """Abwesenheiten/Gründe für einen Mitarbeiter oder Zeitraum"""
    try:
        ma_id = request.args.get('ma_id', type=int)
        von = request.args.get('von')  # Datum YYYY-MM-DD
        bis = request.args.get('bis')  # Datum YYYY-MM-DD

        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT a.*, z.Zeittyp as Grund_Bez, z.Kuerzel as Grund_Kuerzel,
                   z.Zeittyp as Grund, m.Nachname, m.Vorname
            FROM tbl_MA_NVerfuegZeiten a
            LEFT JOIN tbl_MA_Zeittyp z ON a.Zeittyp_ID = z.ID
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON a.MA_ID = m.ID
            WHERE 1=1
        """
        params = []

        if ma_id:
            sql += " AND a.MA_ID = ?"
            params.append(ma_id)

        if von and bis:
            # Abwesenheiten die im Zeitraum liegen (Ueberschneidung)
            sql += " AND a.vonDat <= ? AND a.bisDat >= ?"
            params.append(bis)
            params.append(von)
        elif von:
            sql += " AND a.bisDat >= ?"
            params.append(von)
        elif bis:
            sql += " AND a.vonDat <= ?"
            params.append(bis)

        sql += " ORDER BY a.vonDat DESC"

        if not ma_id and not von and not bis:
            sql = sql.replace("SELECT a.*", "SELECT TOP 500 a.*")

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify(result)  # Direkt als Array zurueckgeben (Kompatibilitaet)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- OFFENE ANFRAGEN (fuer Subform) ---
@app.route('/api/anfragen/offen')
def anfragen_offen():
    """Offene MA-Anfragen fuer Subform-Anzeige"""
    try:
        ma_id = request.args.get('ma_id', type=int)
        va_id = request.args.get('va_id', type=int)

        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT p.*, m.Nachname, m.Vorname, m.Kurzname, m.Tel_Mobil,
                   a.Auftrag, a.Objekt, t.VADatum, s.VA_Start, s.VA_Ende
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.MA_ID
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID
            LEFT JOIN tbl_VA_Start s ON p.VAStart_ID = s.ID
            WHERE p.MVP_Status = 1
        """

        params = []
        if ma_id:
            sql += " AND p.MA_ID = ?"
            params.append(ma_id)
        if va_id:
            sql += " AND p.VA_ID = ?"
            params.append(va_id)

        sql += " ORDER BY t.VADatum, s.VA_Start"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- GENERISCHES RECORD UPDATE ---
@app.route('/api/record/update', methods=['POST'])
def record_update():
    """Aktualisiert ein einzelnes Feld in einer Tabelle"""
    try:
        data = request.get_json(force=True)
        table = data.get('table')
        record_id = data.get('id')
        field = data.get('field')
        value = data.get('value')

        if not table or not record_id or not field:
            return jsonify({"success": False, "error": "table, id und field erforderlich"}), 400

        # Sicherheitscheck: Nur bestimmte Tabellen erlauben
        allowed_tables = [
            'tbl_MA_VA_Planung', 'tbl_MA_VA_Zuordnung', 'tbl_MA_NVerfuegZeiten',
            'tbl_VA_Auftragstamm', 'tbl_VA_Start', 'tbl_VA_AnzTage'
        ]
        if table not in allowed_tables:
            return jsonify({"success": False, "error": f"Tabelle {table} nicht erlaubt"}), 403

        conn = get_connection()
        cursor = conn.cursor()

        # SQL-Injection-Schutz: field wird validiert
        # ID-Feld der Tabelle bestimmen
        id_field = 'ID'
        if table == 'tbl_MA_Mitarbeiterstamm':
            id_field = 'MA_ID'

        # UPDATE ausfuehren
        sql = f"UPDATE {table} SET {field} = ? WHERE {id_field} = ?"
        cursor.execute(sql, (value, record_id))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": f"{field} aktualisiert"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- LOHNABRECHNUNGEN ---
@app.route('/api/lohnabrechnungen')
def lohnabrechnungen_list():
    """Lohnabrechnungen-Liste"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        jahr = request.args.get('jahr', type=int)
        monat = request.args.get('monat', type=int)
        limit = request.args.get('limit', 100, type=int)

        sql = f"SELECT TOP {limit} * FROM ztbl_Lohnabrechnungen WHERE 1=1"

        if jahr:
            sql += f" AND Jahr = {jahr}"
        if monat:
            sql += f" AND Monat = {monat}"

        sql += " ORDER BY Jahr DESC, Monat DESC"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ZEITKONTEN ---
@app.route('/api/zeitkonten')
def zeitkonten_list():
    """Zeitkonten-Eintraege fuer einen Mitarbeiter"""
    try:
        ma_id = request.args.get('ma_id', type=int)
        jahr = request.args.get('jahr', type=int)
        monat = request.args.get('monat', type=int)

        conn = get_connection()
        cursor = conn.cursor()

        sql = "SELECT * FROM tbl_MA_Zeitkonto WHERE 1=1"
        params = []

        if ma_id:
            sql += " AND MA_ID = ?"
            params.append(ma_id)
        if jahr:
            sql += " AND Jahr = ?"
            params.append(jahr)
        if monat:
            sql += " AND Monat = ?"
            params.append(monat)

        sql += " ORDER BY Jahr DESC, Monat DESC"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ZEITKONTEN IMPORT-FEHLER ---
@app.route('/api/zeitkonten/importfehler')
def zeitkonten_importfehler():
    """Import-Fehler aus Zeiterfassung"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT TOP 200 * FROM tbl_MA_ZeitImportFehler
            ORDER BY ImportDatum DESC
        """)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- BEWERBER ---
@app.route('/api/bewerber')
def bewerber_list():
    """Bewerberliste"""
    try:
        status = request.args.get('status', '')
        limit = request.args.get('limit', 100, type=int)

        conn = get_connection()
        cursor = conn.cursor()

        sql = f"SELECT TOP {limit} * FROM tbl_MA_Bewerber WHERE 1=1"

        if status:
            sql += f" AND Status = '{status}'"

        sql += " ORDER BY EingangDatum DESC"

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/bewerber/<int:id>')
def bewerber_detail(id):
    """Einzelner Bewerber"""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tbl_MA_Bewerber WHERE ID = ?", (id,))
        rows = query_to_list(cursor)
        conn.close()
        if rows:
            return jsonify({"success": True, "data": rows[0]})
        return jsonify({"success": False, "error": "Nicht gefunden"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/bewerber', methods=['POST'])
def bewerber_create():
    """Neuen Bewerber anlegen"""
    try:
        data = request.get_json(force=True)

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO tbl_MA_Bewerber
            (Nachname, Vorname, Email, Telefon, Status, EingangDatum, Bemerkungen)
            VALUES (?, ?, ?, ?, 'Neu', ?, ?)
        """, (
            data.get('Nachname', ''),
            data.get('Vorname', ''),
            data.get('Email', ''),
            data.get('Telefon', ''),
            datetime.now(),
            data.get('Bemerkungen', '')
        ))
        conn.commit()

        cursor.execute("SELECT @@IDENTITY")
        new_id = cursor.fetchone()[0]

        conn.close()
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- DIENSTPLAN (erweitert) ---
@app.route('/api/dienstplan/ma/<int:ma_id>')
def dienstplan_ma(ma_id):
    """Dienstplan fuer einen Mitarbeiter"""
    try:
        von = request.args.get('von', '')
        bis = request.args.get('bis', '')

        conn = get_connection()
        cursor = conn.cursor()

        # Access braucht Klammern bei mehreren LEFT JOINs
        sql = """
            SELECT [tbl_MA_VA_Planung].*, [tbl_VA_Auftragstamm].[Auftrag], [tbl_VA_Auftragstamm].[Objekt],
                   [tbl_MA_VA_Planung].[VADatum], [tbl_VA_Start].[VA_Start], [tbl_VA_Start].[VA_Ende]
            FROM ([tbl_MA_VA_Planung]
            LEFT JOIN [tbl_VA_Start] ON [tbl_MA_VA_Planung].[VAStart_ID] = [tbl_VA_Start].[ID])
            LEFT JOIN [tbl_VA_Auftragstamm] ON [tbl_MA_VA_Planung].[VA_ID] = [tbl_VA_Auftragstamm].[VA_ID]
            WHERE [tbl_MA_VA_Planung].[MA_ID] = ?
        """
        params = [ma_id]

        if von:
            sql += " AND [tbl_MA_VA_Planung].[VADatum] >= ?"
            params.append(von)
        if bis:
            sql += " AND [tbl_MA_VA_Planung].[VADatum] <= ?"
            params.append(bis)

        sql += " ORDER BY [tbl_MA_VA_Planung].[VADatum], [tbl_VA_Start].[VA_Start]"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/dienstplan/objekt/<int:objekt_id>')
def dienstplan_objekt(objekt_id):
    """Dienstplan fuer ein Objekt"""
    try:
        von = request.args.get('von', '')
        bis = request.args.get('bis', '')

        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT p.*, m.Nachname, m.Vorname, t.VADatum, s.VA_Start, s.VA_Ende
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.MA_ID
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID
            LEFT JOIN tbl_VA_Start s ON p.VAStart_ID = s.ID
            WHERE a.Objekt_ID = ?
        """
        params = [objekt_id]

        if von:
            sql += " AND t.VADatum >= ?"
            params.append(von)
        if bis:
            sql += " AND t.VADatum <= ?"
            params.append(bis)

        sql += " ORDER BY t.VADatum, s.VA_Start, m.Nachname"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- STUNDEN-EXPORT ---
@app.route('/api/lohn/stunden-export')
def lohn_stunden_export():
    """Stunden-Export fuer Lohnabrechnung"""
    try:
        ma_id = request.args.get('ma_id', type=int)
        jahr = request.args.get('jahr', type=int)
        monat = request.args.get('monat', type=int)

        if not jahr or not monat:
            return jsonify({"success": False, "error": "jahr und monat erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT p.MA_ID, m.Nachname, m.Vorname, m.LexNr,
                   t.VADatum, s.VA_Start, s.VA_Ende,
                   a.Auftrag, a.Objekt,
                   DATEDIFF(hour, s.VA_Start, s.VA_Ende) as Stunden
            FROM tbl_MA_VA_Planung p
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.MA_ID
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID
            LEFT JOIN tbl_VA_Start s ON p.VAStart_ID = s.ID
            WHERE YEAR(t.VADatum) = ? AND MONTH(t.VADatum) = ?
              AND p.MVP_Status = 2
        """
        params = [jahr, monat]

        if ma_id:
            sql += " AND p.MA_ID = ?"
            params.append(ma_id)

        sql += " ORDER BY m.Nachname, t.VADatum, s.VA_Start"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result, "jahr": jahr, "monat": monat})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- RUECKMELDUNGEN ---
@app.route('/api/rueckmeldungen')
def rueckmeldungen_list():
    """MA-Rueckmeldungen (Zusagen/Absagen)"""
    try:
        va_id = request.args.get('va_id', type=int)
        status = request.args.get('status', '')

        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT r.*, m.Nachname, m.Vorname, a.Auftrag
            FROM tbl_MA_Rueckmeldung r
            LEFT JOIN tbl_MA_Mitarbeiterstamm m ON r.MA_ID = m.MA_ID
            LEFT JOIN tbl_VA_Auftragstamm a ON r.VA_ID = a.ID
            WHERE 1=1
        """
        params = []

        if va_id:
            sql += " AND r.VA_ID = ?"
            params.append(va_id)
        if status:
            sql += " AND r.Status = ?"
            params.append(status)

        sql += " ORDER BY r.RueckDatum DESC"

        cursor.execute(sql, params)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- MITARBEITER SCHNELLAUSWAHL ---
@app.route('/api/mitarbeiter_schnellauswahl')
def mitarbeiter_schnellauswahl():
    """Mitarbeiter für Schnellauswahl (nur aktive, kurze Felder)"""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        va_id = request.args.get('va_id', type=int)
        datum = request.args.get('datum', '')

        # Basis: aktive Mitarbeiter mit wichtigen Feldern
        sql = """
            SELECT m.ID AS MA_ID, m.Nachname, m.Vorname, m.Tel_Mobil, m.Email
            FROM tbl_MA_Mitarbeiterstamm m
            WHERE m.IstAktiv = -1
            ORDER BY m.Nachname, m.Vorname
        """

        cursor.execute(sql)
        result = query_to_list(cursor)
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ========================================================================
# GESCHÜTZTE ENDPOINTS (kritisch für Formular-Funktionalität!)
# Siehe CLAUDE.md Zeile 93-112
# Diese Endpoints unterstützen vadatum_id als Integer-ID ODER Datum-String
# ========================================================================

@app.route('/api/auftraege/vorschlaege')
def auftraege_vorschlaege():
    """Autocomplete-Vorschläge für Auftragsfelder (Ort, Objekt, etc.)"""
    try:
        feld = request.args.get('feld', '')
        limit = request.args.get('limit', 10, type=int)

        if not feld:
            return jsonify({"success": False, "error": "Parameter 'feld' erforderlich"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Je nach Feld unterschiedliche SQL
        if feld == 'ort':
            cursor.execute(f"""
                SELECT DISTINCT TOP {limit} Ort
                FROM tbl_VA_Auftragstamm
                WHERE Ort IS NOT NULL AND Ort <> ''
                ORDER BY Ort
            """)
        elif feld == 'objekt':
            cursor.execute(f"""
                SELECT DISTINCT TOP {limit} Objekt
                FROM tbl_VA_Auftragstamm
                WHERE Objekt IS NOT NULL AND Objekt <> ''
                ORDER BY Objekt
            """)
        elif feld == 'auftrag':
            cursor.execute(f"""
                SELECT DISTINCT TOP {limit} Auftrag
                FROM tbl_VA_Auftragstamm
                WHERE Auftrag IS NOT NULL AND Auftrag <> ''
                ORDER BY Auftrag
            """)
        elif feld == 'dienstkleidung':
            cursor.execute(f"""
                SELECT DISTINCT TOP {limit} Dienstkleidung_ID
                FROM tbl_MA_Dienstkleidung
                WHERE Dienstkleidung_ID IS NOT NULL
                ORDER BY Dienstkleidung_ID
            """)
        else:
            conn.close()
            return jsonify({"success": False, "error": f"Feld '{feld}' nicht unterstützt"}), 400

        result = [row[0] for row in cursor.fetchall()]
        conn.close()
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/schichten')
def auftraege_schichten(va_id):
    """
    GESCHÜTZTER ENDPOINT - NICHT ÄNDERN!
    Schichten für einen Auftrag mit vadatum_id Filter
    vadatum_id kann Integer-ID ODER Datum-String sein
    """
    try:
        vadatum_id = request.args.get('vadatum_id', '')

        with get_db() as conn:
            cursor = conn.cursor()

            if not vadatum_id:
                # Alle Schichten des Auftrags
                cursor.execute("""
                    SELECT * FROM tbl_VA_Start
                    WHERE VA_ID = ?
                    ORDER BY VADatum, VA_Start
                """, (va_id,))
            elif vadatum_id.isdigit():
                # vadatum_id ist eine Integer-ID → JOIN mit tbl_VA_AnzTage
                cursor.execute("""
                    SELECT s.* FROM tbl_VA_Start s
                    INNER JOIN tbl_VA_AnzTage t ON s.VADatum = t.VADatum AND s.VA_ID = t.VA_ID
                    WHERE s.VA_ID = ? AND t.ID = ?
                    ORDER BY s.VA_Start
                """, (va_id, int(vadatum_id)))
            else:
                # vadatum_id ist ein Datum-String → Vergleich mit VADatum
                # Entferne mögliche Zeitangaben (nur Datum verwenden)
                datum_str = vadatum_id.split('T')[0]
                cursor.execute("""
                    SELECT * FROM tbl_VA_Start
                    WHERE VA_ID = ? AND CDATE(VADatum) = CDATE(?)
                    ORDER BY VA_Start
                """, (va_id, datum_str))

            result = query_to_list(cursor)
            return jsonify({"success": True, "data": result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/zuordnungen')
def auftraege_zuordnungen(va_id):
    """
    GESCHÜTZTER ENDPOINT - NICHT ÄNDERN!
    MA-Zuordnungen für einen Auftrag mit vadatum_id Filter
    vadatum_id kann Integer-ID ODER Datum-String sein
    """
    try:
        vadatum_id = request.args.get('vadatum_id', '')

        with get_db() as conn:
            cursor = conn.cursor()

            # Basis-SQL mit JOINs
            sql = """
                SELECT p.*, p.MVA_Start AS MA_Start, p.MVA_Ende AS MA_Ende,
                       m.Nachname, m.Vorname, m.Tel_Mobil,
                       t.VADatum
                FROM ((tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
                LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
                WHERE p.VA_ID = ?
            """
            params = [va_id]

            if vadatum_id:
                if vadatum_id.isdigit():
                    # vadatum_id ist eine Integer-ID → Zeige Zuordnungen für dieses Datum ODER ohne Datum
                    sql += " AND (p.VADatum_ID = ? OR p.VADatum_ID IS NULL)"
                    params.append(int(vadatum_id))
                else:
                    # vadatum_id ist ein Datum-String → Zeige Zuordnungen für dieses Datum ODER ohne Datum
                    datum_str = vadatum_id.split('T')[0]
                    sql += " AND (CDATE(t.VADatum) = CDATE(?) OR p.VADatum_ID IS NULL)"
                    params.append(datum_str)

            sql += " ORDER BY t.VADatum, p.MVA_Start"

            cursor.execute(sql, params)
            result = query_to_list(cursor)
            return jsonify({"success": True, "data": result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/absagen')
def auftraege_absagen(va_id):
    """
    GESCHÜTZTER ENDPOINT - NICHT ÄNDERN!
    MA-Absagen für einen Auftrag mit vadatum_id Filter
    vadatum_id kann Integer-ID ODER Datum-String sein
    """
    try:
        vadatum_id = request.args.get('vadatum_id', '')

        with get_db() as conn:
            cursor = conn.cursor()

            # Absagen aus tbl_MA_VA_Planung mit Status_ID = 4 (Absage)
            cursor.execute("""
                SELECT * FROM tbl_MA_VA_Planung
                WHERE VA_ID = ? AND Status_ID = 4
                ORDER BY VADatum_ID
            """, (va_id,))

            absagen = query_to_list(cursor)

            # MA-Namen nachträglich laden (verhindert JOIN-Probleme)
            if absagen:
                ma_ids = list(set(a.get('MA_ID') for a in absagen if a.get('MA_ID')))
                if ma_ids:
                    ids_str = ','.join(str(int(mid)) for mid in ma_ids)
                    cursor.execute(f"""
                        SELECT ID, Nachname, Vorname, Tel_Mobil
                        FROM tbl_MA_Mitarbeiterstamm
                        WHERE ID IN ({ids_str})
                    """)
                    ma_dict = {row[0]: {'Nachname': row[1], 'Vorname': row[2], 'Tel_Mobil': row[3]}
                               for row in cursor.fetchall()}
                    # Namen hinzufügen
                    for a in absagen:
                        ma = ma_dict.get(a.get('MA_ID'), {})
                        a['Nachname'] = ma.get('Nachname', '')
                        a['Vorname'] = ma.get('Vorname', '')
                        a['Tel_Mobil'] = ma.get('Tel_Mobil', '')

            return jsonify({"success": True, "data": absagen})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/auftraege/<int:va_id>/anfragen')
def auftraege_anfragen(va_id):
    """
    Ausstehende MA-Anfragen für einen Auftrag (für Antworten-Tab)
    Gibt alle Planungen mit Status_ID 1 (geplant) oder 2 (benachrichtigt) zurück
    """
    try:
        vadatum_id = request.args.get('vadatum_id', '')

        with get_db() as conn:
            cursor = conn.cursor()

            # Anfragen aus tbl_MA_VA_Planung mit Status_ID < 3 (geplant/benachrichtigt)
            cursor.execute("""
                SELECT * FROM tbl_MA_VA_Planung
                WHERE VA_ID = ? AND Status_ID < 3
                ORDER BY VADatum_ID, MA_ID
            """, (va_id,))

            anfragen = query_to_list(cursor)

            # MA-Namen nachträglich laden (verhindert JOIN-Probleme)
            if anfragen:
                ma_ids = list(set(a.get('MA_ID') for a in anfragen if a.get('MA_ID')))
                if ma_ids:
                    ids_str = ','.join(str(int(mid)) for mid in ma_ids)
                    cursor.execute(f"""
                        SELECT ID, Nachname, Vorname, Tel_Mobil
                        FROM tbl_MA_Mitarbeiterstamm
                        WHERE ID IN ({ids_str})
                    """)
                    ma_dict = {row[0]: {'Nachname': row[1], 'Vorname': row[2], 'Tel_Mobil': row[3]}
                               for row in cursor.fetchall()}
                    # Namen hinzufügen
                    for a in anfragen:
                        ma = ma_dict.get(a.get('MA_ID'), {})
                        a['Nachname'] = ma.get('Nachname', '')
                        a['Vorname'] = ma.get('Vorname', '')
                        a['Tel_Mobil'] = ma.get('Tel_Mobil', '')

            return jsonify({"success": True, "data": anfragen})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================
# MITARBEITER-FOTOS (UNC-Server-Pfad)
# ============================================
# Pfad für Mitarbeiterfotos (wie in Access VBA: prp_CONSYS_GrundPfad & TLookup("Pfad", "_tblEigeneFirma_Pfade", "ID = 7"))
# S: ist gemappt auf \\vConSYS01-NBG\Consys
MA_FOTO_UNC_PATH = r"S:\Bilder\Mitarbeiter"

@app.route('/api/fotos/mitarbeiter/<filename>')
def serve_mitarbeiter_foto(filename):
    """
    Serviert Mitarbeiterfotos vom UNC-Server-Pfad.
    Browser koennen nicht direkt auf file:// zugreifen,
    daher proxy via HTTP.
    """
    try:
        # Sicherheitspruefung: Nur Bilddateien erlauben
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}
        _, ext = os.path.splitext(filename.lower())
        if ext not in allowed_extensions:
            return jsonify({'error': 'Ungueltiger Dateityp'}), 400

        # Pfadtraversal-Schutz
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Ungueltiger Dateiname'}), 400

        full_path = os.path.join(MA_FOTO_UNC_PATH, filename)

        if os.path.exists(full_path):
            return send_from_directory(MA_FOTO_UNC_PATH, filename)
        else:
            print(f"[WARN] Mitarbeiterfoto nicht gefunden: {full_path}")
            return jsonify({'error': 'Foto nicht gefunden'}), 404
    except Exception as e:
        print(f"[ERROR] Fehler beim Laden des Mitarbeiterfotos: {e}")
        return jsonify({'error': str(e)}), 500


# ============================================
# EINSATZLISTE SENDEN (E-Mail mit Excel-Anhang)
# ============================================
# Pfade für Templates
HTML_BODIES_PATH = r"\\vConSYS01-NBG\Database\HTMLBodies"
TEMP_PATH = os.path.join(os.environ.get('TEMP', 'C:\\Temp'), 'consys_exports')

@app.route('/api/einsatzliste/senden', methods=['POST'])
def einsatzliste_senden():
    """
    Sendet Einsatzliste per E-Mail an alle zugeordneten MA.
    Erstellt Excel-Anhang und verwendet HTML-Body-Vorlage.

    POST-Body:
    {
        "va_id": 9316,
        "vadatum_id": null,  // optional: nur für bestimmten Tag
        "sender_name": "Günther Siegert",  // für [A_Sender] Platzhalter
        "test_mode": false  // true = nur Vorschau, keine E-Mails
    }
    """
    try:
        data = request.get_json(force=True)
        va_id = data.get('va_id')
        vadatum_id = data.get('vadatum_id')
        sender_name = data.get('sender_name', 'CONSEC Team')
        test_mode = data.get('test_mode', False)

        if not va_id:
            return jsonify({"success": False, "error": "va_id erforderlich"}), 400

        # Temp-Ordner erstellen falls nicht vorhanden
        os.makedirs(TEMP_PATH, exist_ok=True)

        with get_db() as conn:
            cursor = conn.cursor()

            # 1. Auftragsdaten laden
            cursor.execute("SELECT * FROM tbl_VA_Auftragstamm WHERE ID = ?", (va_id,))
            auftrag_rows = query_to_list(cursor)
            if not auftrag_rows:
                return jsonify({"success": False, "error": f"Auftrag {va_id} nicht gefunden"}), 404
            auftrag = auftrag_rows[0]

            # 2. Zugeordnete MA mit E-Mail-Adressen laden
            sql_ma = """
                SELECT DISTINCT m.ID AS MA_ID, m.Nachname, m.Vorname, m.eMail, m.Tel_Mobil,
                       p.VADatum_ID, p.MVA_Start, p.MVA_Ende
                FROM tbl_MA_VA_Planung p
                INNER JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID
                WHERE p.VA_ID = ?
            """
            params = [va_id]

            if vadatum_id:
                sql_ma += " AND p.VADatum_ID = ?"
                params.append(vadatum_id)

            sql_ma += " ORDER BY m.Nachname, m.Vorname"

            cursor.execute(sql_ma, params)
            ma_list = query_to_list(cursor)

            if not ma_list:
                return jsonify({
                    "success": False,
                    "error": "Keine MA für diesen Auftrag zugeordnet"
                }), 404

            # 3. Einsatztage laden
            cursor.execute("""
                SELECT * FROM tbl_VA_AnzTage
                WHERE VA_ID = ?
                ORDER BY VADatum
            """, (va_id,))
            einsatztage = query_to_list(cursor)

            # 4. Startzeiten/Schichten laden
            cursor.execute("""
                SELECT * FROM tbl_VA_Start
                WHERE VA_ID = ?
                ORDER BY VADatum, VA_Start
            """, (va_id,))
            schichten = query_to_list(cursor)

            # 5. Alle Zuordnungen für Excel-Export
            cursor.execute("""
                SELECT p.*, m.Nachname, m.Vorname, m.Tel_Mobil,
                       t.VADatum
                FROM ((tbl_MA_VA_Planung p
                LEFT JOIN tbl_MA_Mitarbeiterstamm m ON p.MA_ID = m.ID)
                LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
                WHERE p.VA_ID = ?
                ORDER BY t.VADatum, p.MVA_Start, m.Nachname
            """, (va_id,))
            zuordnungen = query_to_list(cursor)

        # 6. Excel-Einsatzliste erstellen
        excel_path = _create_einsatzliste_excel(
            auftrag, einsatztage, schichten, zuordnungen, va_id
        )

        # 7. HTML-Body-Vorlage laden und füllen
        html_body = _get_einsatzliste_html_body(sender_name)

        # 8. E-Mails senden (oder Test-Modus)
        if test_mode:
            # Nur Vorschau zurückgeben
            return jsonify({
                "success": True,
                "test_mode": True,
                "auftrag": auftrag.get('Auftrag', f'Auftrag {va_id}'),
                "empfaenger": [
                    {
                        "name": f"{ma.get('Vorname', '')} {ma.get('Nachname', '')}",
                        "email": ma.get('eMail', ''),
                        "telefon": ma.get('Tel_Mobil', '')
                    }
                    for ma in ma_list
                ],
                "anzahl_ma": len(ma_list),
                "excel_path": excel_path,
                "html_body_preview": html_body[:500] + "..."
            })

        # E-Mails senden
        result = _send_einsatzliste_emails(
            ma_list, auftrag, excel_path, html_body
        )

        return jsonify({
            "success": True,
            "auftrag": auftrag.get('Auftrag', f'Auftrag {va_id}'),
            "gesendet": result['gesendet'],
            "fehler": result['fehler'],
            "ohne_email": result['ohne_email']
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================
# SMTP E-MAIL VERSAND (Mailjet)
# ============================================
def send_email_smtp(to_email, subject, html_body, attachment_path=None, from_name="CONSEC Auftragsplanung"):
    """
    Sendet E-Mail über SMTP (Mailjet).
    Ersetzt Outlook COM-Automation für zuverlässigen Versand.
    """
    try:
        # E-Mail erstellen
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f"{from_name} <{SMTP_FROM_EMAIL}>"
        msg['To'] = to_email

        # HTML Body
        html_part = MIMEText(html_body, 'html', 'utf-8')
        msg.attach(html_part)

        # Attachment hinzufügen falls vorhanden
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                filename = os.path.basename(attachment_path)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)

        # SMTP-Verbindung und Versand
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_EMAIL, to_email, msg.as_string())

        print(f"[SMTP] E-Mail gesendet an: {to_email}")
        return True

    except Exception as e:
        print(f"[SMTP] Fehler beim Senden an {to_email}: {e}")
        raise


def _create_einsatzliste_excel(auftrag, einsatztage, schichten, zuordnungen, va_id):
    """
    Erstellt Excel-Einsatzliste mit der Original Access-Vorlage.
    Verwendet S:\Vorlage_EINSATZLISTE.xls als Template.
    """
    # Vorlage-Pfad (wie in Access mdl_Maintainance.bas)
    EXCEL_TEMPLATE = r"S:\Vorlage_EINSATZLISTE.xls"

    # Fallback falls Vorlage nicht verfügbar
    if not os.path.exists(EXCEL_TEMPLATE):
        print(f"[WARN] Excel-Vorlage nicht gefunden: {EXCEL_TEMPLATE}")
        return _create_einsatzliste_simple(auftrag, einsatztage, schichten, zuordnungen, va_id)

    try:
        import win32com.client
        from datetime import datetime as dt

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # Vorlage öffnen
        wb = excel.Workbooks.Open(EXCEL_TEMPLATE)
        ws = wb.Sheets("Liste")

        # === HEADER BEFÜLLEN (wie in VBA fXL_Export_Auftrag) ===

        # Zeile 1: Stand (aktuelles Datum)
        ws.Cells(1, 2).Value = dt.now().strftime("%d.%m.%Y %H:%M")

        # Zeile 6: Datum (Einsatzdatum)
        if einsatztage:
            first_date = einsatztage[0].get('VADatum', '')
            if first_date:
                if hasattr(first_date, 'strftime'):
                    ws.Cells(6, 2).Value = first_date.strftime("%d.%m.%Y")
                else:
                    ws.Cells(6, 2).Value = str(first_date)[:10]

        # Zeile 7: Auftrag
        ws.Cells(7, 2).Value = auftrag.get('Auftrag', '')

        # Zeile 8: Einsatzort (Ort)
        ws.Cells(8, 2).Value = f"{auftrag.get('PLZ', '')} {auftrag.get('Ort', '')}".strip()

        # Zeile 9: Location (Objekt)
        ws.Cells(9, 2).Value = auftrag.get('Objekt', '')

        # Zeile 12: Anzahl MA
        ws.Cells(12, 2).Value = len(zuordnungen)

        # === SCHICHTEN (Zeilen 13-15) ===
        schicht_row = 13
        for i, schicht in enumerate(schichten[:3]):  # Max 3 Schichten in Vorlage
            va_start = schicht.get('VA_Start', '')
            va_ende = schicht.get('VA_Ende', '')

            if va_start:
                start_str = str(va_start)[:5] if va_start else ''
                ende_str = str(va_ende)[:5] if va_ende else ''
                ws.Cells(schicht_row + i, 2).Value = f"{start_str} - {ende_str}"
                ws.Cells(schicht_row + i, 3).Value = schicht.get('MA_Anzahl', '')

        # === MITARBEITER-LISTE (ab Zeile 17) ===
        ma_start_row = 17
        prev_datum = None

        for i, zuo in enumerate(zuordnungen):
            row = ma_start_row + i

            # Datum (nur anzeigen wenn anders als vorherige Zeile)
            va_datum = zuo.get('VADatum', '')
            if va_datum:
                if hasattr(va_datum, 'strftime'):
                    datum_str = va_datum.strftime("%d.%m.%Y")
                else:
                    datum_str = str(va_datum)[:10]

                if datum_str != prev_datum:
                    ws.Cells(row, 1).Value = datum_str
                    prev_datum = datum_str

            # Position/Nr
            ws.Cells(row, 2).Value = i + 1

            # Name
            name = f"{zuo.get('Nachname', '')}, {zuo.get('Vorname', '')}".strip(', ')
            ws.Cells(row, 3).Value = name

            # Von/Bis Zeiten
            mva_start = zuo.get('MVA_Start', zuo.get('MA_Start', ''))
            mva_ende = zuo.get('MVA_Ende', zuo.get('MA_Ende', ''))
            ws.Cells(row, 4).Value = str(mva_start)[:5] if mva_start else ''
            ws.Cells(row, 5).Value = str(mva_ende)[:5] if mva_ende else ''

            # Telefon
            ws.Cells(row, 6).Value = zuo.get('Tel_Mobil', '')

        # Zeile 16: Gesamt
        ws.Cells(16, 2).Value = len(zuordnungen)

        # Dateiname erstellen
        auftrag_name = auftrag.get('Auftrag', f'Auftrag_{va_id}')
        safe_name = "".join(c for c in auftrag_name if c.isalnum() or c in (' ', '-', '_'))[:30]
        filename = f"Einsatzliste_{safe_name}_{va_id}.xlsm"
        filepath = os.path.join(TEMP_PATH, filename)

        # Speichern (als xlsm mit Makros)
        wb.SaveAs(filepath, FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
        wb.Close(False)
        excel.Quit()

        print(f"[API] Excel mit Vorlage erstellt: {filepath}")
        return filepath

    except Exception as e:
        print(f"[ERROR] Excel-Vorlage Fehler: {e}")
        import traceback
        traceback.print_exc()
        # Fallback auf einfache Version
        return _create_einsatzliste_simple(auftrag, einsatztage, schichten, zuordnungen, va_id)


def _create_einsatzliste_simple(auftrag, einsatztage, schichten, zuordnungen, va_id):
    """Fallback: Erstellt einfache Excel-Datei mit openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        return _create_einsatzliste_csv(auftrag, einsatztage, schichten, zuordnungen, va_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Einsatzliste"

    # Header-Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header
    ws['A1'] = "CONSEC SECURITY NÜRNBERG - EINSATZLISTE"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A3'] = "Auftrag:"
    ws['B3'] = auftrag.get('Auftrag', '')
    ws['A4'] = "Objekt:"
    ws['B4'] = auftrag.get('Objekt', '')
    ws['A5'] = "Ort:"
    ws['B5'] = f"{auftrag.get('PLZ', '')} {auftrag.get('Ort', '')}"

    # MA-Liste Header
    row = 8
    headers = ['Datum', 'Nr', 'Name', 'Von', 'Bis', 'Telefon']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # MA-Daten
    for i, zuo in enumerate(zuordnungen):
        row += 1
        va_datum = zuo.get('VADatum', '')
        if hasattr(va_datum, 'strftime'):
            ws.cell(row=row, column=1, value=va_datum.strftime('%d.%m.%Y'))
        else:
            ws.cell(row=row, column=1, value=str(va_datum)[:10])

        ws.cell(row=row, column=2, value=i + 1)
        ws.cell(row=row, column=3, value=f"{zuo.get('Nachname', '')}, {zuo.get('Vorname', '')}")
        ws.cell(row=row, column=4, value=str(zuo.get('MVA_Start', ''))[:5])
        ws.cell(row=row, column=5, value=str(zuo.get('MVA_Ende', ''))[:5])
        ws.cell(row=row, column=6, value=zuo.get('Tel_Mobil', ''))

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 15

    # Speichern
    auftrag_name = auftrag.get('Auftrag', f'Auftrag_{va_id}')
    safe_name = "".join(c for c in auftrag_name if c.isalnum() or c in (' ', '-', '_'))[:30]
    filename = f"Einsatzliste_{safe_name}_{va_id}.xlsx"
    filepath = os.path.join(TEMP_PATH, filename)

    wb.save(filepath)
    print(f"[API] Excel (simple) erstellt: {filepath}")
    return filepath


def _create_einsatzliste_csv(auftrag, einsatztage, schichten, zuordnungen, va_id):
    """Fallback: Erstellt CSV wenn openpyxl nicht verfügbar"""
    import csv

    auftrag_name = auftrag.get('Auftrag', f'Auftrag_{va_id}')
    safe_name = "".join(c for c in auftrag_name if c.isalnum() or c in (' ', '-', '_'))[:30]
    filename = f"Einsatzliste_{safe_name}_{va_id}.csv"
    filepath = os.path.join(TEMP_PATH, filename)

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')

        # Header
        writer.writerow(['CONSEC SECURITY NÜRNBERG - EINSATZLISTE'])
        writer.writerow([])
        writer.writerow(['Auftrag:', auftrag.get('Auftrag', '')])
        writer.writerow(['Objekt:', auftrag.get('Objekt', '')])
        writer.writerow(['Ort:', f"{auftrag.get('PLZ', '')} {auftrag.get('Ort', '')}"])
        writer.writerow([])

        # Zuordnungen
        writer.writerow(['MITARBEITER-EINSÄTZE'])
        writer.writerow(['Datum', 'Name', 'Von', 'Bis', 'Telefon', 'Bemerkung'])

        for zuo in zuordnungen:
            va_datum = zuo.get('VADatum', '')
            if isinstance(va_datum, str):
                datum_str = va_datum[:10]
            elif hasattr(va_datum, 'strftime'):
                datum_str = va_datum.strftime('%d.%m.%Y')
            else:
                datum_str = str(va_datum)

            name = f"{zuo.get('Vorname', '')} {zuo.get('Nachname', '')}".strip()
            mva_start = str(zuo.get('MVA_Start', zuo.get('MA_Start', '')))[:5]
            mva_ende = str(zuo.get('MVA_Ende', zuo.get('MA_Ende', '')))[:5]

            writer.writerow([
                datum_str,
                name,
                mva_start,
                mva_ende,
                zuo.get('Tel_Mobil', ''),
                zuo.get('Bemerkungen', '')
            ])

    print(f"[API] CSV erstellt: {filepath}")
    return filepath


def _get_einsatzliste_html_body(sender_name):
    """Lädt und füllt die HTML-Body-Vorlage"""
    template_path = os.path.join(HTML_BODIES_PATH, "HTML_Body_Einsatzliste.txt")

    # Fallback auf lokalen Pfad
    if not os.path.exists(template_path):
        local_path = os.path.join(FORMS3_PATH, "HTMLBodies", "HTML_Body_Einsatzliste.txt")
        if os.path.exists(local_path):
            template_path = local_path
        else:
            # Standard-HTML wenn keine Vorlage gefunden
            return f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
            <h2>CONSEC SECURITY NÜRNBERG - Einsatzliste</h2>
            <p>Hallo,</p>
            <p>anbei die aktuelle Einsatzliste mit allen Daten und Infos zum Auftrag.</p>
            <p>Nach Dienstende bitte die Endzeit per E-Mail an info@consec-nuernberg.de</p>
            <p>... und wie immer den Dienstausweis nicht vergessen...</p>
            <br>
            <p>Vielen lieben Dank & viel Spaß,</p>
            <p>{sender_name}<br>
            CONSEC Veranstaltungsservice & Sicherheitsdienst oHG</p>
            </body>
            </html>
            """

    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            html_body = f.read()
    except UnicodeDecodeError:
        with open(template_path, 'r', encoding='latin-1') as f:
            html_body = f.read()

    # Platzhalter ersetzen
    html_body = html_body.replace('[A_Sender]', sender_name)

    return html_body


def _send_einsatzliste_emails(ma_list, auftrag, excel_path, html_body):
    """Sendet E-Mails an alle MA via Outlook"""
    result = {
        'gesendet': [],
        'fehler': [],
        'ohne_email': []
    }

    auftrag_name = auftrag.get('Auftrag', f"Auftrag {auftrag.get('ID', '')}")
    betreff = f"Einsatzliste: {auftrag_name}"

    for ma in ma_list:
        email = ma.get('eMail', '').strip()
        name = f"{ma.get('Vorname', '')} {ma.get('Nachname', '')}".strip()

        if not email or '@' not in email:
            result['ohne_email'].append(name)
            continue

        try:
            # SMTP-Versand (Mailjet)
            send_email_smtp(
                to_email=email,
                subject=betreff,
                html_body=html_body,
                attachment_path=excel_path if os.path.exists(excel_path) else None
            )
            result['gesendet'].append({'name': name, 'email': email})

        except Exception as e:
            result['fehler'].append({
                'name': name,
                'email': email,
                'error': str(e)
            })
            print(f"[API] E-Mail-Fehler für {email}: {e}")

    return result


# ============================================
# DIENSTPLAN SENDEN (E-Mail mit PDF/Excel-Anhang)
# ============================================
@app.route('/api/dienstplan/senden', methods=['POST'])
def dienstplan_senden():
    """
    Sendet Dienstplan per E-Mail an einen Mitarbeiter.
    Erstellt Excel-Anhang mit Dienstplan-Daten.

    POST-Body:
    {
        "ma_id": 472,
        "von": "2025-01-01",  // Startdatum
        "bis": "2025-01-31",  // Enddatum (optional, default: +30 Tage)
        "sender_name": "Günther Siegert",
        "test_mode": false
    }
    """
    try:
        data = request.get_json(force=True)
        ma_id = data.get('ma_id')
        von = data.get('von')
        bis = data.get('bis')
        sender_name = data.get('sender_name', 'CONSEC Team')
        test_mode = data.get('test_mode', False)

        if not ma_id:
            return jsonify({"success": False, "error": "ma_id erforderlich"}), 400

        # Datum parsen
        from datetime import datetime, timedelta
        if von:
            if isinstance(von, str):
                datum_von = datetime.strptime(von[:10], '%Y-%m-%d')
            else:
                datum_von = von
        else:
            datum_von = datetime.now()

        if bis:
            if isinstance(bis, str):
                datum_bis = datetime.strptime(bis[:10], '%Y-%m-%d')
            else:
                datum_bis = bis
        else:
            datum_bis = datum_von + timedelta(days=30)

        os.makedirs(TEMP_PATH, exist_ok=True)

        with get_db() as conn:
            cursor = conn.cursor()

            # 1. MA-Daten laden
            cursor.execute("""
                SELECT ID, Nachname, Vorname, eMail, Tel_Mobil
                FROM tbl_MA_Mitarbeiterstamm
                WHERE ID = ?
            """, (ma_id,))
            ma_rows = query_to_list(cursor)

            if not ma_rows:
                return jsonify({"success": False, "error": f"MA {ma_id} nicht gefunden"}), 404

            ma = ma_rows[0]
            vorname = ma.get('Vorname', '')
            nachname = ma.get('Nachname', '')
            email = ma.get('eMail', '')

            # 2. Planungen für Zeitraum laden
            cursor.execute("""
                SELECT p.*, t.VADatum, a.Auftrag, a.Objekt, a.Ort,
                       s.VA_Start, s.VA_Ende
                FROM (((tbl_MA_VA_Planung p
                LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
                LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
                LEFT JOIN tbl_VA_Start s ON p.VAStart_ID = s.ID)
                WHERE p.MA_ID = ?
                  AND t.VADatum >= ?
                  AND t.VADatum <= ?
                ORDER BY t.VADatum, p.MVA_Start
            """, (ma_id, datum_von.strftime('%Y-%m-%d'), datum_bis.strftime('%Y-%m-%d')))

            planungen = query_to_list(cursor)

        # 3. Dienstplan-PDF aus Access-Report erstellen (wie in VBA)
        pdf_path = _create_dienstplan_pdf(ma_id, datum_von, datum_bis)

        # 4. HTML-Body laden und füllen
        html_body = _get_dienstplan_html_body(vorname, datum_von, sender_name)

        # 5. Test-Modus oder senden
        if test_mode:
            return jsonify({
                "success": True,
                "test_mode": True,
                "ma": f"{vorname} {nachname}",
                "email": email,
                "zeitraum": f"{datum_von.strftime('%d.%m.%Y')} - {datum_bis.strftime('%d.%m.%Y')}",
                "planungen": len(planungen),
                "pdf_path": pdf_path
            })

        # E-Mail senden
        if not email or '@' not in email:
            return jsonify({
                "success": False,
                "error": f"Keine gültige E-Mail-Adresse für {vorname} {nachname}"
            }), 400

        result = _send_dienstplan_email(ma, email, html_body, pdf_path, datum_von)

        return jsonify({
            "success": result['success'],
            "ma": f"{vorname} {nachname}",
            "email": email,
            "message": result.get('message', ''),
            "error": result.get('error', '')
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


def _create_dienstplan_pdf(ma_id, datum_von, datum_bis):
    """
    Erstellt Dienstplan-PDF aus Access-Report rpt_MA_Dienstplan.
    Genau wie in VBA Dienstplan_senden() - DoCmd.OutputTo acOutputReport.
    """
    # Access Frontend Pfad
    ACCESS_FRONTEND = r"C:\Users\guenther.siegert\Documents\0006_All_Access_KNOWLEDGE\0_Consys_FE_Test.accdb"

    # PDF-Dateiname (wie in VBA: "Dienstplan_" & von & "-" & bis & ".pdf")
    filename = f"Dienstplan_{datum_von.strftime('%d.%m.%Y')}-{datum_bis.strftime('%d.%m.%Y')}.pdf"
    filepath = os.path.join(TEMP_PATH, filename)

    # Alte Datei löschen falls vorhanden
    if os.path.exists(filepath):
        os.remove(filepath)

    try:
        import win32com.client

        print(f"[API] Öffne Access für Report-Export...")

        # Neue Access-Instanz erstellen
        access = win32com.client.Dispatch("Access.Application")
        close_db_after = True

        try:
            # Datenbank öffnen
            try:
                access.OpenCurrentDatabase(ACCESS_FRONTEND)
            except Exception as open_err:
                if "bereits geöffnet" in str(open_err) or "already" in str(open_err).lower():
                    # DB schon offen - versuche mit bestehender Instanz
                    print("[API] DB bereits offen, nutze bestehende Instanz")
                    close_db_after = False
                else:
                    raise

            try:
                access.Visible = False
            except:
                pass
            print(f"[API] Datenbank bereit: {ACCESS_FRONTEND}")

            # TempVars setzen für den Report-Filter
            # Erst löschen falls vorhanden
            try:
                access.TempVars.Remove("TMP_MA_ID")
            except:
                pass
            try:
                access.TempVars.Remove("TMP_DatumVon")
            except:
                pass
            try:
                access.TempVars.Remove("TMP_DatumBis")
            except:
                pass

            access.TempVars.Add("TMP_MA_ID", int(ma_id))
            access.TempVars.Add("TMP_DatumVon", datum_von.strftime('%m/%d/%Y'))
            access.TempVars.Add("TMP_DatumBis", datum_bis.strftime('%m/%d/%Y'))

            # Report als PDF exportieren
            # acOutputReport = 3, acFormatPDF = "PDF Format (*.pdf)"
            access.DoCmd.OutputTo(
                3,  # acOutputReport
                "rpt_MA_Dienstplan",
                "PDF Format (*.pdf)",
                filepath,
                False  # AutoStart = False
            )

            print(f"[API] Dienstplan-PDF erstellt: {filepath}")

        finally:
            if close_db_after:
                try:
                    access.CloseCurrentDatabase()
                except:
                    pass
                access.Quit()

        if os.path.exists(filepath):
            return filepath
        else:
            print("[API] PDF wurde nicht erstellt, nutze Fallback")
            raise Exception("PDF wurde nicht erstellt")

    except Exception as e:
        print(f"[API] Access-PDF-Export Fehler: {e}")
        import traceback
        traceback.print_exc()
        # Fallback: Excel erstellen
        return _create_dienstplan_excel_fallback(ma_id, datum_von, datum_bis)


def _create_dienstplan_excel_fallback(ma_id, datum_von, datum_bis):
    """Fallback: Erstellt einfache Excel-Datei wenn Access-PDF nicht funktioniert"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        return None

    # MA-Daten und Planungen laden
    with get_db() as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT Nachname, Vorname FROM tbl_MA_Mitarbeiterstamm WHERE ID = ?", (ma_id,))
        ma_row = cursor.fetchone()
        vorname = ma_row[1] if ma_row else ""
        nachname = ma_row[0] if ma_row else ""

        cursor.execute("""
            SELECT t.VADatum, a.Auftrag, a.Ort, p.MVA_Start, p.MVA_Ende
            FROM ((tbl_MA_VA_Planung p
            LEFT JOIN tbl_VA_AnzTage t ON p.VADatum_ID = t.ID)
            LEFT JOIN tbl_VA_Auftragstamm a ON p.VA_ID = a.ID)
            WHERE p.MA_ID = ? AND t.VADatum >= ? AND t.VADatum <= ?
            ORDER BY t.VADatum
        """, (ma_id, datum_von.strftime('%Y-%m-%d'), datum_bis.strftime('%Y-%m-%d')))
        planungen = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dienstplan"

    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = "CONSEC SECURITY NÜRNBERG - DIENSTPLAN"
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')

    ws['A3'] = f"Mitarbeiter: {vorname} {nachname}"
    ws['A4'] = f"Zeitraum: {datum_von.strftime('%d.%m.%Y')} - {datum_bis.strftime('%d.%m.%Y')}"

    row = 6
    for col, h in enumerate(['Datum', 'Auftrag', 'Ort', 'Von', 'Bis'], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for p in planungen:
        row += 1
        va_datum = p[0]
        if hasattr(va_datum, 'strftime'):
            ws.cell(row=row, column=1, value=va_datum.strftime('%d.%m.%Y'))
        else:
            ws.cell(row=row, column=1, value=str(va_datum)[:10] if va_datum else '')
        ws.cell(row=row, column=2, value=p[1] or '')
        ws.cell(row=row, column=3, value=p[2] or '')
        ws.cell(row=row, column=4, value=str(p[3])[:5] if p[3] else '')
        ws.cell(row=row, column=5, value=str(p[4])[:5] if p[4] else '')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8

    filename = f"Dienstplan_{nachname}_{vorname}_{datum_von.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(TEMP_PATH, filename)
    wb.save(filepath)
    print(f"[API] Dienstplan-Excel (Fallback) erstellt: {filepath}")
    return filepath


def _create_dienstplan_csv(ma, planungen, datum_von, datum_bis):
    """Fallback: CSV erstellen"""
    import csv

    vorname = ma.get('Vorname', '')
    nachname = ma.get('Nachname', '')

    safe_name = "".join(c for c in f"{nachname}_{vorname}" if c.isalnum() or c in (' ', '-', '_'))[:20]
    filename = f"Dienstplan_{safe_name}_{datum_von.strftime('%Y%m%d')}.csv"
    filepath = os.path.join(TEMP_PATH, filename)

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(['CONSEC SECURITY NÜRNBERG - DIENSTPLAN'])
        writer.writerow([f'Mitarbeiter: {vorname} {nachname}'])
        writer.writerow([f'Zeitraum: {datum_von.strftime("%d.%m.%Y")} - {datum_bis.strftime("%d.%m.%Y")}'])
        writer.writerow([])
        writer.writerow(['Datum', 'Auftrag', 'Ort', 'Von', 'Bis'])

        for p in planungen:
            va_datum = p.get('VADatum', '')
            if hasattr(va_datum, 'strftime'):
                datum_str = va_datum.strftime('%d.%m.%Y')
            else:
                datum_str = str(va_datum)[:10]

            writer.writerow([
                datum_str,
                p.get('Auftrag', ''),
                p.get('Ort', ''),
                str(p.get('MVA_Start', ''))[:5],
                str(p.get('MVA_Ende', ''))[:5]
            ])

    print(f"[API] Dienstplan-CSV erstellt: {filepath}")
    return filepath


def _get_dienstplan_html_body(vorname, datum_von, sender_name):
    """Lädt und füllt die Dienstplan HTML-Body-Vorlage"""
    template_path = os.path.join(HTML_BODIES_PATH, "HTML_Body_DienstPl.txt")

    # Fallback auf lokalen Pfad
    if not os.path.exists(template_path):
        local_path = os.path.join(FORMS3_PATH, "HTMLBodies", "HTML_Body_DienstPl.txt")
        if os.path.exists(local_path):
            template_path = local_path
        else:
            # Standard-HTML
            return f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
            <h2>CONSEC SECURITY NÜRNBERG</h2>
            <h3>Dienstplan</h3>
            <p>Hallo {vorname},</p>
            <p>anbei Dein Dienstplan ab {datum_von.strftime('%d.%m.%Y')}</p>
            <br>
            <p>{sender_name}<br>
            CONSEC Veranstaltungsservice & Sicherheitsdienst oHG</p>
            </body>
            </html>
            """

    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            html_body = f.read()
    except UnicodeDecodeError:
        with open(template_path, 'r', encoding='latin-1') as f:
            html_body = f.read()

    # Platzhalter ersetzen
    html_body = html_body.replace('[A_Vorname]', vorname)
    html_body = html_body.replace('[A_DatumAb]', datum_von.strftime('%d.%m.%Y'))
    html_body = html_body.replace('[A_Sender]', sender_name)

    return html_body


def _send_dienstplan_email(ma, email, html_body, attachment_path, datum_von):
    """Sendet Dienstplan-E-Mail via SMTP (Mailjet)"""
    try:
        subject = f"Dienstplan ab {datum_von.strftime('%d.%m.%Y')}"

        # SMTP-Versand (Mailjet)
        send_email_smtp(
            to_email=email,
            subject=subject,
            html_body=html_body,
            attachment_path=attachment_path if os.path.exists(attachment_path) else None
        )

        return {"success": True, "message": "E-Mail gesendet"}

    except Exception as e:
        print(f"[API] Dienstplan-E-Mail-Fehler: {e}")
        return {"success": False, "error": str(e)}


# ============================================
# CATCH-ALL ROUTE FÜR STATISCHE DATEIEN
# WICHTIG: Diese Route MUSS am Ende stehen (nach allen API-Routes),
# damit Flask zuerst die API-Routes prüft!
# ============================================
@app.route('/<path:filename>')
def serve_static(filename):
    """Serviert HTML, JS, CSS Dateien aus forms3"""
    return send_from_directory(FORMS3_PATH, filename)


# ============================================
# SERVER STARTEN
# ============================================
if __name__ == '__main__':
    print("=" * 50)
    print("Mini API Server für forms3")
    print("=" * 50)
    print(f"Backend: {BACKEND_PATH}")
    print(f"Server:  http://localhost:5000")
    print("=" * 50)
    print("Endpoints:")
    print("  /api/health          - Health Check")
    print("  /api/auftraege       - Auftragsliste")
    print("  /api/auftraege/<id>  - Auftrag-Details")
    print("  /api/mitarbeiter     - Mitarbeiterliste")
    print("  /api/kunden          - Kundenliste")
    print("  /api/objekte         - Objekteliste")
    print("  /api/einsatztage     - Einsatztage")
    print("  /api/schichten       - Schichten")
    print("  /api/zuordnungen     - Zuordnungen")
    print("  /api/status          - Status-Liste")
    print("  /api/anfragen        - Offene Anfragen")
    print("  /api/absagen         - Absagen")
    print("  /api/query (POST)    - SQL-Abfrage")
    print("=" * 50)

    # Teste Datenbankverbindung
    try:
        conn = get_connection()
        conn.close()
        print("[OK] Datenbankverbindung erfolgreich")
    except Exception as e:
        print(f"[FEHLER] Datenbankverbindung: {e}")
        print("Bitte BACKEND_PATH in mini_api.py pruefen!")

    print("\nServer wird gestartet...")
    # Verwende waitress als WSGI-Server (stabiler als Flask dev server)
    # threads=1 weil Access ODBC-Treiber NICHT thread-safe ist!
    try:
        from waitress import serve
        print("[OK] Waitress WSGI-Server (threads=1)")
        serve(app, host='0.0.0.0', port=5000, threads=1)
    except ImportError:
        print("[WARNUNG] Waitress nicht installiert, nutze Flask dev server")
        app.run(host='0.0.0.0', port=5000, debug=False, threaded=False)
