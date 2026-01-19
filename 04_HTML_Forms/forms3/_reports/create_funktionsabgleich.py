#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agent C: Funktionsabgleich HTML ↔ Access
Vergleicht HTML-Formulare mit Access-Formularen und erstellt vollständigen Abgleich
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Fix für Windows Console Encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Pfade
BASE_DIR = r"C:\Users\guenther.siegert\Documents\0006_All_Access_KNOWLEDGE\04_HTML_Forms\forms3\_reports"
HTML_JSON = os.path.join(BASE_DIR, "HTML_FORMULARE_ANALYSE_2026-01-15.json")
ACCESS_JSON = os.path.join(BASE_DIR, "ACCESS_FORMULARE_ANALYSE_2026-01-15.json")

# Output-Dateien
DATUM = datetime.now().strftime("%Y-%m-%d")
EXCEL_OUT = os.path.join(BASE_DIR, f"FUNKTIONS_ABGLEICH_{DATUM}.xlsx")
MD_OUT = os.path.join(BASE_DIR, f"FUNKTIONS_ABGLEICH_{DATUM}.md")
FEHLEND_OUT = os.path.join(BASE_DIR, f"FEHLENDE_FUNKTIONEN_{DATUM}.md")


def load_json(filepath):
    """Lädt JSON-Datei"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def normalize_name(name):
    """Normalisiert Formular-Namen für Matching"""
    # Entferne .html Extension
    name = name.replace('.html', '')
    # Kleinbuchstaben für Vergleich
    return name.lower()


def match_formulare(html_forms, access_forms):
    """
    Matched HTML-Formulare mit Access-Formularen
    Returns: dict {html_name: access_name}
    """
    matches = {}
    access_lookup = {normalize_name(f['name']): f['name'] for f in access_forms}

    for html_form in html_forms:
        html_name = html_form['name']
        normalized = normalize_name(html_name)

        if normalized in access_lookup:
            matches[html_name] = access_lookup[normalized]

    return matches


def map_control_type(access_type):
    """
    Mapped Access Control Type zu erwarteten HTML Typen
    """
    mapping = {
        109: ['input', 'textarea'],  # TextBox → input/textarea
        111: ['select', 'datalist'],  # ComboBox → select/datalist
        110: ['select'],  # ListBox → select[multiple]
        104: ['button'],  # CommandButton → button
        100: ['label', 'span', 'div'],  # Label → label/span/div
        112: ['iframe'],  # Subform → iframe
        106: ['input'],  # OptionButton → input[radio]
        122: ['input'],  # CheckBox → input[checkbox]
        108: ['img'],  # Image → img
        114: ['div'],  # Tab Control → div.tab-container
        # Weitere nach Bedarf
    }
    return mapping.get(access_type, ['div'])


def compare_controls(html_form, access_form):
    """
    Vergleicht Controls zwischen HTML und Access
    Returns: dict mit Statistiken und Abweichungen
    """
    # HTML Controls sind nach Typ gruppiert (input, select, button, etc.)
    html_controls_by_type = html_form.get('controls', {})
    html_controls_flat = []
    for ctrl_type, ctrl_list in html_controls_by_type.items():
        for ctrl in ctrl_list:
            # Nutze 'id' oder 'name' als Identifier
            identifier = ctrl.get('id') or ctrl.get('name')
            if identifier:
                html_controls_flat.append({
                    'name': identifier,
                    'type': ctrl_type,
                    **ctrl
                })

    # Access Controls sind eine Liste
    access_controls_list = access_form.get('controls', [])

    result = {
        'html_controls': len(html_controls_flat),
        'access_controls': len(access_controls_list),
        'matched': 0,
        'missing_in_html': [],
        'additional_in_html': [],
        'type_mismatches': [],
        'match_percentage': 0.0
    }

    html_controls = {c['name'].lower(): c for c in html_controls_flat if c.get('name')}
    access_controls = {c['Name'].lower(): c for c in access_controls_list if c.get('Name')}

    # Prüfe Access-Controls in HTML
    for access_name, access_ctrl in access_controls.items():
        if access_name in html_controls:
            html_ctrl = html_controls[access_name]
            result['matched'] += 1

            # Prüfe Typ-Mapping
            expected_types = map_control_type(access_ctrl.get('ControlType', 0))
            html_type = html_ctrl.get('type', 'unknown')

            if html_type not in expected_types:
                result['type_mismatches'].append({
                    'name': access_ctrl['Name'],
                    'access_type': access_ctrl.get('ControlType', 0),
                    'html_type': html_type,
                    'expected': expected_types
                })
        else:
            result['missing_in_html'].append({
                'name': access_ctrl['Name'],
                'type': access_ctrl.get('ControlType', 0),
                'caption': access_ctrl.get('Caption', '')
            })

    # Prüfe zusätzliche HTML-Controls
    for html_name, html_ctrl in html_controls.items():
        if html_name not in access_controls:
            result['additional_in_html'].append({
                'name': html_ctrl['name'],
                'type': html_ctrl.get('type', 'unknown')
            })

    # Berechne Match-Percentage
    total = len(access_controls)
    if total > 0:
        result['match_percentage'] = (result['matched'] / total) * 100.0

    return result


def compare_events(html_form, access_form):
    """
    Vergleicht Event-Handler zwischen HTML und Access
    """
    result = {
        'access_events': 0,
        'html_events': 0,
        'matched': 0,
        'missing_in_html': [],
        'match_percentage': 0.0
    }

    # Sammle Access Events
    access_events = defaultdict(list)
    for ctrl in access_form.get('controls', []):
        for event, value in ctrl.items():
            if event.startswith('On') and value:
                access_events[ctrl['Name'].lower()].append(event)
                result['access_events'] += 1

    # Sammle HTML Events - sind im 'events' Dict nach Control-ID
    html_events_dict = html_form.get('events', {})
    html_events = defaultdict(list)
    for ctrl_id, event_list in html_events_dict.items():
        html_events[ctrl_id.lower()] = event_list
        result['html_events'] += len(event_list)

    # Vergleiche
    for ctrl_name, events in access_events.items():
        html_ctrl_events = html_events.get(ctrl_name, [])

        for event in events:
            # Einfaches Mapping (kann erweitert werden)
            mapped = map_access_event_to_html(event)

            if any(m in html_ctrl_events for m in mapped):
                result['matched'] += 1
            else:
                result['missing_in_html'].append({
                    'control': ctrl_name,
                    'event': event,
                    'expected_html': mapped
                })

    # Berechne Percentage
    if result['access_events'] > 0:
        result['match_percentage'] = (result['matched'] / result['access_events']) * 100.0

    return result


def map_access_event_to_html(access_event):
    """
    Mapped Access Event zu HTML Event(s)
    """
    mapping = {
        'OnClick': ['onclick', 'click'],
        'OnDblClick': ['ondblclick', 'dblclick'],
        'OnChange': ['onchange', 'change'],
        'AfterUpdate': ['onchange', 'change', 'onblur', 'blur'],
        'BeforeUpdate': ['onchange', 'change'],
        'OnLoad': ['DOMContentLoaded', 'load'],
        'OnCurrent': ['load', 'custom'],
        'OnEnter': ['onfocus', 'focus'],
        'OnExit': ['onblur', 'blur'],
        'OnKeyPress': ['onkeypress', 'keypress'],
        'OnMouseMove': ['onmousemove', 'mousemove'],
    }
    return mapping.get(access_event, [])


def compare_validations(html_form, access_form):
    """
    Vergleicht Validierungen zwischen HTML und Access
    """
    result = {
        'access_rules': 0,
        'html_validations': 0,
        'matched': 0,
        'missing_in_html': [],
        'match_percentage': 0.0
    }

    # HTML Validations sind im 'validations' Dict
    html_validations_dict = html_form.get('validations', {})
    result['html_validations'] = len(html_validations_dict)

    # Sammle Access ValidationRules
    for ctrl in access_form.get('controls', []):
        if ctrl.get('ValidationRule'):
            result['access_rules'] += 1
            ctrl_name = ctrl['Name'].lower()

            # Prüfe ob HTML Control Validierung hat
            if ctrl_name in [k.lower() for k in html_validations_dict.keys()]:
                result['matched'] += 1
            else:
                result['missing_in_html'].append({
                    'control': ctrl['Name'],
                    'rule': ctrl['ValidationRule']
                })

    # Berechne Percentage
    if result['access_rules'] > 0:
        result['match_percentage'] = (result['matched'] / result['access_rules']) * 100.0

    return result


def categorize_severity(controls_comp, events_comp, validations_comp):
    """
    Kategorisiert Abweichungen nach Schweregrad
    Returns: dict {kritisch: [], wichtig: [], optional: []}
    """
    result = {
        'kritisch': [],
        'wichtig': [],
        'optional': []
    }

    # KRITISCH: Fehlende Controls mit Events
    for missing in controls_comp['missing_in_html']:
        # Prüfe ob Control Events hatte
        had_events = False  # TODO: Prüfung erweitern

        if had_events or missing['type'] == 104:  # CommandButton
            result['kritisch'].append({
                'type': 'Fehlendes Control',
                'detail': f"Control '{missing['name']}' fehlt in HTML"
            })

    # KRITISCH: Fehlende Validierungen
    for missing in validations_comp['missing_in_html']:
        result['kritisch'].append({
            'type': 'Fehlende Validierung',
            'detail': f"Control '{missing['control']}' hat keine Validierung: {missing['rule']}"
        })

    # WICHTIG: Fehlende Events
    for missing in events_comp['missing_in_html']:
        result['wichtig'].append({
            'type': 'Fehlendes Event',
            'detail': f"Control '{missing['control']}' fehlt Event '{missing['event']}'"
        })

    # WICHTIG: Typ-Mismatch
    for mismatch in controls_comp['type_mismatches']:
        result['wichtig'].append({
            'type': 'Typ-Mismatch',
            'detail': f"Control '{mismatch['name']}': Erwartet {mismatch['expected']}, ist {mismatch['html_type']}"
        })

    # OPTIONAL: Zusätzliche HTML-Controls
    for additional in controls_comp['additional_in_html']:
        result['optional'].append({
            'type': 'Zusätzliches Control',
            'detail': f"Control '{additional['name']}' nur in HTML vorhanden"
        })

    return result


def create_excel_report(comparison_results, output_path):
    """
    Erstellt Excel-Report mit mehreren Sheets
    """
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    kritisch_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    wichtig_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    optional_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Übersicht
    ws_overview = wb.active
    ws_overview.title = "Übersicht"

    headers = ["Formular", "Controls Match %", "Events Match %", "Validierung Match %", "Gesamt-Score", "Status"]
    ws_overview.append(headers)

    for cell in ws_overview[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for form_name, result in comparison_results.items():
        controls_pct = result['controls_comparison']['match_percentage']
        events_pct = result['events_comparison']['match_percentage']
        validations_pct = result['validations_comparison']['match_percentage']

        # Gewichteter Durchschnitt
        gesamt = (controls_pct * 0.4 + events_pct * 0.3 + validations_pct * 0.3)

        status = "🔴 KRITISCH" if gesamt < 70 else "🟡 WICHTIG" if gesamt < 90 else "🟢 OK"

        row = [
            form_name,
            f"{controls_pct:.1f}%",
            f"{events_pct:.1f}%",
            f"{validations_pct:.1f}%",
            f"{gesamt:.1f}%",
            status
        ]
        ws_overview.append(row)

        # Färbe Status-Zelle
        status_cell = ws_overview.cell(row=ws_overview.max_row, column=6)
        if gesamt < 70:
            status_cell.fill = kritisch_fill
        elif gesamt < 90:
            status_cell.fill = wichtig_fill
        else:
            status_cell.fill = optional_fill

    # Auto-width
    for column in ws_overview.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_overview.column_dimensions[column[0].column_letter].width = adjusted_width

    # Sheet 2: Kritische Abweichungen
    ws_kritisch = wb.create_sheet("Kritische Abweichungen")
    headers = ["Formular", "Typ", "Detail"]
    ws_kritisch.append(headers)

    for cell in ws_kritisch[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for form_name, result in comparison_results.items():
        for item in result['severity']['kritisch']:
            ws_kritisch.append([form_name, item['type'], item['detail']])

    # Sheet 3: Wichtige Abweichungen
    ws_wichtig = wb.create_sheet("Wichtige Abweichungen")
    ws_wichtig.append(headers)

    for cell in ws_wichtig[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for form_name, result in comparison_results.items():
        for item in result['severity']['wichtig']:
            ws_wichtig.append([form_name, item['type'], item['detail']])

    # Sheet 4: Detail-Vergleich
    ws_detail = wb.create_sheet("Detailvergleich")
    detail_headers = [
        "Formular",
        "HTML Controls",
        "Access Controls",
        "Matched Controls",
        "Fehlend in HTML",
        "Zusätzlich in HTML",
        "Access Events",
        "HTML Events",
        "Matched Events",
        "Fehlende Events"
    ]
    ws_detail.append(detail_headers)

    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for form_name, result in comparison_results.items():
        cc = result['controls_comparison']
        ec = result['events_comparison']

        row = [
            form_name,
            cc['html_controls'],
            cc['access_controls'],
            cc['matched'],
            len(cc['missing_in_html']),
            len(cc['additional_in_html']),
            ec['access_events'],
            ec['html_events'],
            ec['matched'],
            len(ec['missing_in_html'])
        ]
        ws_detail.append(row)

    wb.save(output_path)
    print(f">> Excel-Report erstellt: {output_path}")


def create_markdown_report(comparison_results, output_path):
    """
    Erstellt Markdown-Report
    """
    lines = []
    lines.append("# Funktionsabgleich HTML ↔ Access")
    lines.append(f"\n**Erstellt:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    # Executive Summary
    lines.append("## Executive Summary\n")

    total_forms = len(comparison_results)
    kritisch_count = sum(1 for r in comparison_results.values()
                        if (r['controls_comparison']['match_percentage'] * 0.4 +
                            r['events_comparison']['match_percentage'] * 0.3 +
                            r['validations_comparison']['match_percentage'] * 0.3) < 70)
    wichtig_count = sum(1 for r in comparison_results.values()
                       if 70 <= (r['controls_comparison']['match_percentage'] * 0.4 +
                                 r['events_comparison']['match_percentage'] * 0.3 +
                                 r['validations_comparison']['match_percentage'] * 0.3) < 90)
    ok_count = total_forms - kritisch_count - wichtig_count

    lines.append(f"- **Gesamt Formulare:** {total_forms}")
    lines.append(f"- **🔴 Kritisch:** {kritisch_count} ({kritisch_count/total_forms*100:.1f}%)")
    lines.append(f"- **🟡 Wichtig:** {wichtig_count} ({wichtig_count/total_forms*100:.1f}%)")
    lines.append(f"- **🟢 OK:** {ok_count} ({ok_count/total_forms*100:.1f}%)")

    # Top 10 Kritische Abweichungen
    lines.append("\n## Top 10 Kritische Abweichungen\n")

    all_kritisch = []
    for form_name, result in comparison_results.items():
        for item in result['severity']['kritisch']:
            all_kritisch.append({
                'form': form_name,
                'type': item['type'],
                'detail': item['detail']
            })

    for i, item in enumerate(all_kritisch[:10], 1):
        lines.append(f"{i}. **{item['form']}** - {item['type']}: {item['detail']}")

    # Formular-für-Formular
    lines.append("\n## Formular-für-Formular Vergleich\n")

    for form_name, result in sorted(comparison_results.items()):
        cc = result['controls_comparison']
        ec = result['events_comparison']
        vc = result['validations_comparison']

        gesamt = (cc['match_percentage'] * 0.4 +
                 ec['match_percentage'] * 0.3 +
                 vc['match_percentage'] * 0.3)

        status = "🔴" if gesamt < 70 else "🟡" if gesamt < 90 else "🟢"

        lines.append(f"### {status} {form_name}\n")
        lines.append(f"**Gesamt-Score:** {gesamt:.1f}%\n")
        lines.append("**Controls:**")
        lines.append(f"- HTML: {cc['html_controls']}, Access: {cc['access_controls']}, Match: {cc['matched']} ({cc['match_percentage']:.1f}%)")
        lines.append(f"- Fehlend in HTML: {len(cc['missing_in_html'])}")
        lines.append(f"- Zusätzlich in HTML: {len(cc['additional_in_html'])}")
        lines.append(f"- Typ-Mismatches: {len(cc['type_mismatches'])}\n")

        lines.append("**Events:**")
        lines.append(f"- Access: {ec['access_events']}, HTML: {ec['html_events']}, Match: {ec['matched']} ({ec['match_percentage']:.1f}%)")
        lines.append(f"- Fehlend in HTML: {len(ec['missing_in_html'])}\n")

        lines.append("**Validierungen:**")
        lines.append(f"- Access: {vc['access_rules']}, Match: {vc['matched']} ({vc['match_percentage']:.1f}%)")
        lines.append(f"- Fehlend in HTML: {len(vc['missing_in_html'])}\n")

        # Kritische Punkte
        if result['severity']['kritisch']:
            lines.append("**🔴 Kritische Punkte:**")
            for item in result['severity']['kritisch'][:5]:
                lines.append(f"- {item['type']}: {item['detail']}")
            lines.append("")

    # Handlungsempfehlungen
    lines.append("\n## Handlungsempfehlungen (Priorisiert)\n")
    lines.append("### Priorität 1: KRITISCH 🔴\n")
    lines.append("Formulare mit Gesamt-Score < 70%:\n")

    for form_name, result in sorted(comparison_results.items(),
                                   key=lambda x: x[1]['controls_comparison']['match_percentage'] * 0.4 +
                                                x[1]['events_comparison']['match_percentage'] * 0.3 +
                                                x[1]['validations_comparison']['match_percentage'] * 0.3):
        gesamt = (result['controls_comparison']['match_percentage'] * 0.4 +
                 result['events_comparison']['match_percentage'] * 0.3 +
                 result['validations_comparison']['match_percentage'] * 0.3)

        if gesamt < 70:
            lines.append(f"- **{form_name}** ({gesamt:.1f}%)")
            lines.append(f"  - {len(result['severity']['kritisch'])} kritische Abweichungen")
            lines.append(f"  - {len(result['controls_comparison']['missing_in_html'])} fehlende Controls")
            lines.append(f"  - {len(result['events_comparison']['missing_in_html'])} fehlende Events\n")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f">> Markdown-Report erstellt: {output_path}")


def create_fehlend_report(comparison_results, output_path):
    """
    Erstellt Checkliste für fehlende Funktionen
    """
    lines = []
    lines.append("# Fehlende Funktionen - Checkliste")
    lines.append(f"\n**Erstellt:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    # Was fehlt in HTML?
    lines.append("## Was fehlt in HTML?\n")

    for form_name, result in sorted(comparison_results.items()):
        has_missing = (result['controls_comparison']['missing_in_html'] or
                      result['events_comparison']['missing_in_html'] or
                      result['validations_comparison']['missing_in_html'])

        if has_missing:
            lines.append(f"### {form_name}\n")

            if result['controls_comparison']['missing_in_html']:
                lines.append("**Fehlende Controls:**")
                for item in result['controls_comparison']['missing_in_html']:
                    lines.append(f"- [ ] `{item['name']}` (Typ: {item['type']}, Caption: {item.get('caption', 'N/A')})")
                lines.append("")

            if result['events_comparison']['missing_in_html']:
                lines.append("**Fehlende Events:**")
                for item in result['events_comparison']['missing_in_html']:
                    lines.append(f"- [ ] `{item['control']}` → {item['event']} (Erwartet: {', '.join(item['expected_html'])})")
                lines.append("")

            if result['validations_comparison']['missing_in_html']:
                lines.append("**Fehlende Validierungen:**")
                for item in result['validations_comparison']['missing_in_html']:
                    lines.append(f"- [ ] `{item['control']}` → {item['rule']}")
                lines.append("")

    # Was ist besser in HTML?
    lines.append("\n## Was ist besser in HTML?\n")

    for form_name, result in sorted(comparison_results.items()):
        if result['controls_comparison']['additional_in_html']:
            lines.append(f"### {form_name}\n")
            lines.append("**Zusätzliche Controls:**")
            for item in result['controls_comparison']['additional_in_html']:
                lines.append(f"- `{item['name']}` (Typ: {item['type']})")
            lines.append("")

    # Was muss migriert werden?
    lines.append("\n## Was muss migriert werden?\n")
    lines.append("Basierend auf kritischen Abweichungen:\n")

    for form_name, result in sorted(comparison_results.items()):
        if result['severity']['kritisch']:
            lines.append(f"### {form_name}\n")
            for item in result['severity']['kritisch']:
                lines.append(f"- [ ] {item['type']}: {item['detail']}")
            lines.append("")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f">> Fehlende-Funktionen-Checkliste erstellt: {output_path}")


def main():
    print(">> Starte Funktionsabgleich HTML <-> Access...\n")

    # Lade JSON-Daten
    print(">> Lade Analyse-Daten...")
    html_data = load_json(HTML_JSON)
    access_data = load_json(ACCESS_JSON)

    # HTML-Formulare sind als Dict gespeichert
    html_forms_dict = html_data.get('formulare', {})
    html_forms = [{'name': k, **v} for k, v in html_forms_dict.items()]

    # Access-Formulare sind unter 'forms' als Dict
    access_forms_dict = access_data.get('forms', {})
    access_forms = [{'name': k, **v} for k, v in access_forms_dict.items()]

    print(f"   HTML-Formulare: {len(html_forms)}")
    print(f"   Access-Formulare: {len(access_forms)}\n")

    # Matching
    print(">> Matche Formulare...")
    matches = match_formulare(html_forms, access_forms)
    print(f"   Matches gefunden: {len(matches)}\n")

    # Vergleiche jedes Formular
    print(">> Vergleiche Formulare...")
    comparison_results = {}

    for html_name, access_name in matches.items():
        print(f"   Verarbeite: {html_name} <-> {access_name}")

        html_form = next(f for f in html_forms if f['name'] == html_name)
        access_form = next(f for f in access_forms if f['name'] == access_name)

        controls_comp = compare_controls(html_form, access_form)
        events_comp = compare_events(html_form, access_form)
        validations_comp = compare_validations(html_form, access_form)
        severity = categorize_severity(controls_comp, events_comp, validations_comp)

        comparison_results[html_name] = {
            'access_name': access_name,
            'controls_comparison': controls_comp,
            'events_comparison': events_comp,
            'validations_comparison': validations_comp,
            'severity': severity
        }

    print(f"\n>> Vergleich abgeschlossen: {len(comparison_results)} Formulare\n")

    # Erstelle Reports
    print(">> Erstelle Reports...\n")
    create_excel_report(comparison_results, EXCEL_OUT)
    create_markdown_report(comparison_results, MD_OUT)
    create_fehlend_report(comparison_results, FEHLEND_OUT)

    print("\n>> Funktionsabgleich abgeschlossen!")
    print(f"   Excel: {EXCEL_OUT}")
    print(f"   Markdown: {MD_OUT}")
    print(f"   Checkliste: {FEHLEND_OUT}")


if __name__ == '__main__':
    main()
