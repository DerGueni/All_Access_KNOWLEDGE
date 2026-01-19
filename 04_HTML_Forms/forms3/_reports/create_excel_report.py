"""
Erstelle Excel-Report mit mehreren Sheets aus CSV-Daten
"""

import csv
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Pfade
REPORT_PATH = Path(r"C:\Users\guenther.siegert\Documents\0006_All_Access_KNOWLEDGE\04_HTML_Forms\forms3\_reports")
CSV_FILE = REPORT_PATH / "BUTTON_ABWEICHUNGEN_MIT_FORMULAR_2026-01-15.csv"
EXCEL_FILE = REPORT_PATH / "BUTTON_ABWEICHUNGEN_MIT_FORMULAR_2026-01-15.xlsx"

# Farben
COLOR_OK = "C6EFCE"  # Grün
COLOR_MISS = "FFC7CE"  # Rot
COLOR_NEW = "FFEB9C"  # Gelb
COLOR_HEADER = "4472C4"  # Blau


def load_csv_data():
    """Lade CSV-Daten"""
    data = []
    with open(CSV_FILE, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    return data


def create_overview_sheet(wb, data):
    """Erstelle Übersichts-Sheet"""
    ws = wb.create_sheet("Übersicht", 0)

    # Header
    ws['A1'] = "Button-Abweichungsanalyse: HTML vs Access"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    ws['A2'] = "Datum: 15.01.2026"
    ws.merge_cells('A2:D2')

    # Statistik
    stats = {
        'OK': len([r for r in data if r['Status'] == 'OK']),
        'MISS': len([r for r in data if r['Status'] == 'MISS']),
        'NEW': len([r for r in data if r['Status'] == 'NEW']),
        'total': len(data)
    }

    ws['A4'] = "Gesamtstatistik"
    ws['A4'].font = Font(size=14, bold=True)

    headers = ['Kategorie', 'Anzahl', 'Prozent', 'Beschreibung']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(5, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Zeilen
    rows = [
        ('OK - Identisch', stats['OK'], f"{stats['OK']*100//stats['total']}%", 'Button in HTML und Access vorhanden'),
        ('MISS - Fehlt in HTML', stats['MISS'], f"{stats['MISS']*100//stats['total']}%", 'Button nur in Access'),
        ('NEW - Nur in HTML', stats['NEW'], f"{stats['NEW']*100//stats['total']}%", 'Button nur in HTML'),
        ('Gesamt', stats['total'], '100%', 'Alle Button-Einträge'),
    ]

    for row_idx, (cat, count, pct, desc) in enumerate(rows, 6):
        ws.cell(row_idx, 1, cat)
        ws.cell(row_idx, 2, count)
        ws.cell(row_idx, 3, pct)
        ws.cell(row_idx, 4, desc)

        # Farbe je nach Kategorie
        if 'OK' in cat:
            fill = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
        elif 'MISS' in cat:
            fill = PatternFill(start_color=COLOR_MISS, end_color=COLOR_MISS, fill_type="solid")
        elif 'NEW' in cat:
            fill = PatternFill(start_color=COLOR_NEW, end_color=COLOR_NEW, fill_type="solid")
        else:
            fill = None

        if fill:
            for col in range(1, 5):
                ws.cell(row_idx, col).fill = fill

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40


def create_by_html_form_sheet(wb, data):
    """Erstelle Sheet gruppiert nach HTML-Formular"""
    ws = wb.create_sheet("Nach HTML-Formular")

    # Header
    headers = ['Status', 'HTML Formular', 'Access Formular', 'Label', 'HTML ID', 'HTML Action', 'Access Name', 'Access OnClick']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Daten
    row = 2
    for entry in sorted(data, key=lambda x: (x['HTML_Formular'], x['Status'], x['Label'])):
        ws.cell(row, 1, entry['Status'])
        ws.cell(row, 2, entry['HTML_Formular'])
        ws.cell(row, 3, entry['Access_Formular'])
        ws.cell(row, 4, entry['Label'])
        ws.cell(row, 5, entry['HTML_ID'])
        ws.cell(row, 6, entry['HTML_Action'][:50])  # Kürzen
        ws.cell(row, 7, entry['Access_Name'])
        ws.cell(row, 8, entry['Access_OnClick'][:50])  # Kürzen

        # Farbe
        if entry['Status'] == 'OK':
            fill = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
        elif entry['Status'] == 'MISS':
            fill = PatternFill(start_color=COLOR_MISS, end_color=COLOR_MISS, fill_type="solid")
        else:  # NEW
            fill = PatternFill(start_color=COLOR_NEW, end_color=COLOR_NEW, fill_type="solid")

        ws.cell(row, 1).fill = fill

        row += 1

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 35

    # Filter
    ws.auto_filter.ref = f"A1:H{row-1}"


def create_by_status_sheet(wb, data):
    """Erstelle Sheet gruppiert nach Status"""
    ws = wb.create_sheet("Nach Status")

    # Header
    headers = ['Status', 'Label', 'HTML Formular', 'Access Formular', 'HTML ID', 'Access Name']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Daten
    row = 2
    for entry in sorted(data, key=lambda x: (x['Status'], x['HTML_Formular'], x['Label'])):
        ws.cell(row, 1, entry['Status'])
        ws.cell(row, 2, entry['Label'])
        ws.cell(row, 3, entry['HTML_Formular'])
        ws.cell(row, 4, entry['Access_Formular'])
        ws.cell(row, 5, entry['HTML_ID'])
        ws.cell(row, 6, entry['Access_Name'])

        # Farbe
        if entry['Status'] == 'OK':
            fill = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
        elif entry['Status'] == 'MISS':
            fill = PatternFill(start_color=COLOR_MISS, end_color=COLOR_MISS, fill_type="solid")
        else:  # NEW
            fill = PatternFill(start_color=COLOR_NEW, end_color=COLOR_NEW, fill_type="solid")

        ws.cell(row, 1).fill = fill

        row += 1

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    # Filter
    ws.auto_filter.ref = f"A1:F{row-1}"


def create_statistics_sheet(wb, data):
    """Erstelle Statistik-Sheet pro Formular"""
    ws = wb.create_sheet("Statistik")

    # Header
    ws['A1'] = "Statistik pro Formular"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')

    headers = ['HTML Formular', 'Gesamt', 'OK', 'MISS', 'NEW']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    # Gruppiere nach Formular
    by_form = defaultdict(lambda: {'OK': 0, 'MISS': 0, 'NEW': 0})
    for entry in data:
        form = entry['HTML_Formular']
        status = entry['Status']
        by_form[form][status] += 1

    # Daten
    row = 4
    for form in sorted(by_form.keys()):
        stats = by_form[form]
        total = stats['OK'] + stats['MISS'] + stats['NEW']

        ws.cell(row, 1, form)
        ws.cell(row, 2, total)
        ws.cell(row, 3, stats['OK'])
        ws.cell(row, 4, stats['MISS'])
        ws.cell(row, 5, stats['NEW'])

        # Farbe für Status-Zellen
        ws.cell(row, 3).fill = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
        ws.cell(row, 4).fill = PatternFill(start_color=COLOR_MISS, end_color=COLOR_MISS, fill_type="solid")
        ws.cell(row, 5).fill = PatternFill(start_color=COLOR_NEW, end_color=COLOR_NEW, fill_type="solid")

        row += 1

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10


def main():
    print("Erstelle Excel-Report mit mehreren Sheets...")

    # Lade Daten
    print("  [1/5] Lade CSV-Daten...")
    data = load_csv_data()
    print(f"    {len(data)} Einträge geladen")

    # Erstelle Workbook
    print("  [2/5] Erstelle Übersicht...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Entferne Standard-Sheet
    create_overview_sheet(wb, data)

    print("  [3/5] Erstelle Sheet 'Nach HTML-Formular'...")
    create_by_html_form_sheet(wb, data)

    print("  [4/5] Erstelle Sheet 'Nach Status'...")
    create_by_status_sheet(wb, data)

    print("  [5/5] Erstelle Statistik...")
    create_statistics_sheet(wb, data)

    # Speichern
    wb.save(EXCEL_FILE)
    print(f"\n[OK] Excel-Report erstellt: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
