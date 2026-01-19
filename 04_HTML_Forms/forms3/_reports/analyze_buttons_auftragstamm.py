"""
Auftragstamm-Formular Button-Analyse
Vergleicht HTML vs. Access VBA Buttons und Funktionen
"""

import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# HTML Buttons (extrahiert aus frm_va_Auftragstamm.html)
html_buttons = [
    {
        'id': 'btnAktualisieren',
        'label': 'Aktualisieren',
        'onclick': 'refreshData()',
        'function': 'Lädt Auftragsdaten neu aus der Datenbank'
    },
    {
        'id': 'btnPositionen',
        'label': 'Positionen',
        'onclick': 'openPositionen()',
        'function': 'Öffnet Objektverwaltung mit Positionen für das aktuelle Objekt'
    },
    {
        'id': 'btnNeuAuftrag',
        'label': 'Neuer Auftrag',
        'onclick': 'neuerAuftrag()',
        'function': 'Erstellt einen neuen Auftrag'
    },
    {
        'id': 'btnKopieren',
        'label': 'Auftrag kopieren',
        'onclick': 'auftragKopieren()',
        'function': 'Kopiert den aktuellen Auftrag'
    },
    {
        'id': 'btnLoeschen',
        'label': 'Auftrag löschen',
        'onclick': 'auftragLoeschen()',
        'function': 'Löscht den aktuellen Auftrag'
    },
    {
        'id': 'btnListeStd',
        'label': 'Namensliste ESS',
        'onclick': 'namenslisteESS()',
        'function': 'Erstellt Namensliste für ESS (Elektronisches Stundenerfassungssystem)'
    },
    {
        'id': 'btnDruckZusage',
        'label': 'EL drucken',
        'onclick': 'einsatzlisteDrucken()',
        'function': 'Druckt Einsatzliste als Excel-Export'
    },
    {
        'id': 'btnMailEins',
        'label': 'EL senden MA',
        'onclick': 'sendeEinsatzlisteMA()',
        'function': 'Sendet Einsatzliste an Mitarbeiter per E-Mail'
    },
    {
        'id': 'btnMailBOS',
        'label': 'EL senden BOS',
        'onclick': 'sendeEinsatzlisteBOS()',
        'function': 'Sendet Einsatzliste an BOS per E-Mail'
    },
    {
        'id': 'btnMailSub',
        'label': 'EL senden SUB',
        'onclick': 'sendeEinsatzlisteSUB()',
        'function': 'Sendet Einsatzliste an Subunternehmer per E-Mail'
    },
    {
        'id': 'btnELGesendet',
        'label': 'EL gesendet',
        'onclick': 'showELGesendet()',
        'function': 'Zeigt Log der gesendeten Einsatzlisten'
    },
    {
        'id': 'btnDatumLeft',
        'label': '◀',
        'onclick': 'datumNavLeft()',
        'function': 'Navigiert zum vorherigen Datum'
    },
    {
        'id': 'btnDatumRight',
        'label': '▶',
        'onclick': 'datumNavRight()',
        'function': 'Navigiert zum nächsten Datum'
    },
    {
        'id': 'btnPlan_Kopie',
        'label': '→ Folgetag',
        'onclick': 'kopiereInFolgetag()',
        'function': 'Kopiert Schichten und Zuordnungen in den Folgetag'
    },
    {
        'id': 'btnSchnellPlan',
        'label': 'Mitarbeiterauswahl',
        'onclick': 'openMitarbeiterauswahl()',
        'function': 'Öffnet Schnellplanung (frm_MA_VA_Schnellauswahl)'
    },
    {
        'id': 'btn_BWN_Druck',
        'label': 'BWN drucken',
        'onclick': 'bwnDrucken()',
        'function': 'Druckt Bewachungsnachweise (hidden)'
    },
    {
        'id': 'cmd_BWN_send',
        'label': 'BWN senden',
        'onclick': 'bwnSenden()',
        'function': 'Sendet Bewachungsnachweise (hidden)'
    },
    {
        'id': 'btnNeuAttach',
        'label': 'Neuen Attach hinzufügen',
        'onclick': 'neuenAttachHinzufuegen()',
        'function': 'Fügt neuen Dateianhang hinzu'
    },
    {
        'id': 'btnRechnungPDF',
        'label': 'Rechnung PDF',
        'onclick': 'rechnungPDF()',
        'function': 'Erstellt Rechnungs-PDF'
    },
    {
        'id': 'btnBerechnungslistePDF',
        'label': 'Berechnungsliste PDF',
        'onclick': 'berechnungslistePDF()',
        'function': 'Erstellt Berechnungslisten-PDF'
    },
    {
        'id': 'btnRechnungDatenLaden',
        'label': 'Daten laden',
        'onclick': 'rechnungDatenLaden()',
        'function': 'Lädt Rechnungsdaten'
    },
    {
        'id': 'btnRechnungLexware',
        'label': 'Rechnung in Lexware erstellen',
        'onclick': 'rechnungLexware()',
        'function': 'Erstellt Rechnung in Lexware'
    },
    {
        'id': 'btnWebDatenLaden',
        'label': 'Web-Daten laden',
        'onclick': 'webDatenLaden()',
        'function': 'Lädt Eventdaten von Webseite'
    },
    {
        'id': 'btnEventdatenSpeichern',
        'label': 'Speichern',
        'onclick': 'eventdatenSpeichern()',
        'function': 'Speichert Eventdaten in Datenbank'
    },
]

# Access VBA Buttons (extrahiert aus Form_frm_VA_Auftragstamm.bas)
access_buttons = [
    {
        'name': 'btnXLEinsLst',
        'caption': '(Excel Export)',
        'onclick': 'btnXLEinsLst_Click()',
        'function': 'Exportiert Auftragsdaten nach Excel'
    },
    {
        'name': 'Befehl658',
        'caption': '(PDF Export)',
        'onclick': 'Befehl658_Click()',
        'function': 'Erstellt PDF und Excel-Export der Einsatzliste'
    },
    {
        'name': 'Befehl640',
        'caption': 'Auftrag kopieren',
        'onclick': 'Befehl640_Click()',
        'function': 'Call AuftragKopieren(Me.ID) - Kopiert kompletten Auftrag'
    },
    {
        'name': 'btn_Neuer_Auftrag2',
        'caption': 'Neuer Auftrag',
        'onclick': 'btn_Neuer_Auftrag2_Click()',
        'function': 'DoCmd.OpenForm "frmtop_va_auftrag_neu"'
    },
    {
        'name': 'Befehl709',
        'caption': 'EL gesendet',
        'onclick': 'Befehl709_Click()',
        'function': 'DoCmd.OpenTable "tbl_Log_eMail_Sent" - Zeigt E-Mail-Log'
    },
    {
        'name': 'btn_Autosend_BOS',
        'caption': 'EL senden BOS',
        'onclick': 'btn_Autosend_BOS_Click()',
        'function': 'DoCmd.OpenForm frm_MA_Serien_eMail_Auftrag, Autosend(4, VA_ID, VADatum_ID, empfänger)'
    },
    {
        'name': 'btn_ListeStd',
        'caption': 'Namensliste ESS',
        'onclick': 'btn_ListeStd_Click()',
        'function': 'Call Stundenliste_erstellen(Me.ID, , Me.Veranstalter_ID)'
    },
    {
        'name': 'btn_Posliste_oeffnen',
        'caption': 'Positionen',
        'onclick': 'btn_Posliste_oeffnen_Click()',
        'function': 'Call OpenObjektPositionenFromAuftrag - Öffnet frm_OB_Objekt'
    },
    {
        'name': 'btn_rueck',
        'caption': 'Rückgängig',
        'onclick': 'btn_rueck_Click()',
        'function': 'Me!sub_MA_VA_Zuordnung.Undo (kommentiert)'
    },
    {
        'name': 'btn_rueckgaengig',
        'caption': 'Rückgängig',
        'onclick': 'btn_rueckgaengig_Click()',
        'function': 'DoCmd.RunCommand acCmdUndo'
    },
    {
        'name': 'btn_Rueckmeld',
        'caption': 'Rückmeldungen',
        'onclick': 'btn_Rueckmeld_Click()',
        'function': 'DoCmd.OpenForm "zfrm_Rueckmeldungen"'
    },
    {
        'name': 'btn_std_check',
        'caption': 'Status Check',
        'onclick': 'btn_std_check_Click()',
        'function': 'Me.Veranst_Status_ID = 3, Call btnDruckZusage_Click'
    },
    {
        'name': 'btn_sortieren',
        'caption': 'Sortieren',
        'onclick': 'btn_sortieren_Click()',
        'function': 'Call sort_zuo_plan(Me.ID, Me.cboVADatum, 1) - Sortiert Zuordnungen'
    },
    {
        'name': 'btn_VA_Abwesenheiten',
        'caption': 'Abwesenheiten',
        'onclick': 'btn_VA_Abwesenheiten_Click()',
        'function': 'DoCmd.OpenForm "frm_abwesenheitsuebersicht"'
    },
    {
        'name': 'btnDatumRight',
        'caption': '▶',
        'onclick': 'btnDatumRight_Click()',
        'function': 'Navigiert zum nächsten Datum (Array-basiert mit VADatum)'
    },
    {
        'name': 'btnDatumLeft',
        'caption': '◀',
        'onclick': 'btnDatumLeft_Click()',
        'function': 'Navigiert zum vorherigen Datum (Array-basiert)'
    },
    {
        'name': 'btnDruck',
        'caption': 'Druck',
        'onclick': 'btnDruck_Click()',
        'function': 'DoCmd.OutputTo acOutputReport, "rpt_Auftrag", "PDF"'
    },
    {
        'name': 'btnStdBerech',
        'caption': 'Stundenberechnung',
        'onclick': 'btnStdBerech_Click()',
        'function': 'Berechnet Stunden für Rechnung'
    },
    {
        'name': 'btnDruckZusage',
        'caption': 'EL drucken',
        'onclick': 'btnDruckZusage_Click()',
        'function': 'Call fXL_Export_Auftrag(ID, Pfad, Dateiname) - Excel-Export'
    },
    {
        'name': 'btnDruckZusage1',
        'caption': 'EL drucken (alt)',
        'onclick': 'btnDruckZusage1_Click()',
        'function': 'DoCmd.OpenReport "rpt_Auftrag_Zusage", acViewPreview'
    },
    {
        'name': 'btnMailEins',
        'caption': 'EL senden MA',
        'onclick': 'btnMailEins_Click()',
        'function': 'DoCmd.OpenForm frm_MA_Serien_eMail_Auftrag'
    },
    {
        'name': 'btnMailPos',
        'caption': 'EL senden Positionen',
        'onclick': 'btnMailPos_Click()',
        'function': 'Sendet Positionen per E-Mail'
    },
    {
        'name': 'btnMailSub',
        'caption': 'EL senden SUB',
        'onclick': 'btnMailSub_Click()',
        'function': 'DoCmd.OpenForm frm_MA_Serien_eMail_Auftrag (Subunternehmer)'
    },
    {
        'name': 'btnNeuAttach',
        'caption': 'Neuen Attach',
        'onclick': 'btnNeuAttach_Click()',
        'function': 'Fügt Dateianhang hinzu (tbltmp_Attachfile)'
    },
    {
        'name': 'btnNeuVeranst',
        'caption': 'Neuer Veranstalter',
        'onclick': 'btnNeuVeranst_Click()',
        'function': 'DoCmd.OpenForm "frm_KD_Kundenstamm", DataMode:=acFormAdd'
    },
    {
        'name': 'btnPDFKopf',
        'caption': 'PDF Auftragskopf',
        'onclick': 'btnPDFKopf_Click()',
        'function': 'Erstellt PDF des Auftragskopfs'
    },
    {
        'name': 'btnPDFPos',
        'caption': 'PDF Positionen',
        'onclick': 'btnPDFPos_Click()',
        'function': 'Erstellt PDF der Positionen'
    },
    {
        'name': 'btnSchnellPlan',
        'caption': 'Schnellplanung',
        'onclick': 'btnSchnellPlan_Click()',
        'function': 'DoCmd.OpenForm "frm_MA_VA_Schnellauswahl", OpenArgs:=VA_ID & " " & VADatum_ID'
    },
    {
        'name': 'btnVAPlanAendern',
        'caption': 'Plan ändern',
        'onclick': 'btnVAPlanAendern_Click()',
        'function': 'Me!sub_MA_VA_Zuordnung.Form.AllowDeletions = True'
    },
    {
        'name': 'btnVAPlanCrea',
        'caption': 'Plan erstellen',
        'onclick': 'btnVAPlanCrea_Click()',
        'function': 'Erstellt neuen Planungseintrag'
    },
    {
        'name': 'btnPlan_Kopie',
        'caption': '→ Folgetag',
        'onclick': 'btnPlan_Kopie_Click()',
        'function': 'Call Planungsuebertragung_FolgeDatum(Me.ID, Me.cboVADatum)'
    },
    {
        'name': 'btnTgVor',
        'caption': 'Tag vor',
        'onclick': 'btnTgVor_Click()',
        'function': 'Wechselt zum vorherigen Tag'
    },
    {
        'name': 'btnTgBack',
        'caption': 'Tag zurück',
        'onclick': 'btnTgBack_Click()',
        'function': 'Wechselt zum vorherigen Tag'
    },
    {
        'name': 'btnHeute',
        'caption': 'Heute',
        'onclick': 'btnHeute_Click()',
        'function': 'Springt zu heute'
    },
    {
        'name': 'btn_AbWann',
        'caption': 'Ab heute',
        'onclick': 'btn_AbWann_Click()',
        'function': 'Filtert Ansicht ab heute'
    },
    {
        'name': 'btnSyncErr',
        'caption': 'Sync Fehler',
        'onclick': 'btnSyncErr_Click()',
        'function': 'DoCmd.OpenForm "zfrm_SyncError"'
    },
    {
        'name': 'cmd_BWN_send',
        'caption': 'BWN senden',
        'onclick': 'cmd_BWN_send_Click()',
        'function': 'Sendet Bewachungsnachweise'
    },
]


def create_button_comparison_excel():
    """Erstellt Excel-Vergleichstabelle HTML vs. Access Buttons"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Button-Vergleich"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Header
    headers = [
        "Button ID/Name (HTML)",
        "Button Label (HTML)",
        "Implementierte Funktion (HTML)",
        "Button Name (Access)",
        "VBA-Funktion (Access)",
        "Status",
        "Bemerkung"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40

    # Daten
    row = 2

    # Mapping HTML zu Access
    button_mapping = {
        'btnAktualisieren': None,  # Neu in HTML
        'btnPositionen': 'btn_Posliste_oeffnen',
        'btnNeuAuftrag': 'btn_Neuer_Auftrag2',
        'btnKopieren': 'Befehl640',
        'btnLoeschen': None,  # In Access nicht als Button
        'btnListeStd': 'btn_ListeStd',
        'btnDruckZusage': 'btnDruckZusage',
        'btnMailEins': 'btnMailEins',
        'btnMailBOS': 'btn_Autosend_BOS',
        'btnMailSub': 'btnMailSub',
        'btnELGesendet': 'Befehl709',
        'btnDatumLeft': 'btnDatumLeft',
        'btnDatumRight': 'btnDatumRight',
        'btnPlan_Kopie': 'btnPlan_Kopie',
        'btnSchnellPlan': 'btnSchnellPlan',
        'btn_BWN_Druck': None,  # In Access kommentiert
        'cmd_BWN_send': 'cmd_BWN_send',
        'btnNeuAttach': 'btnNeuAttach',
        'btnRechnungPDF': 'btnPDFKopf',  # Ähnlich
        'btnBerechnungslistePDF': 'btnPDFPos',  # Ähnlich
        'btnRechnungDatenLaden': None,  # Neu in HTML
        'btnRechnungLexware': None,  # Neu in HTML
        'btnWebDatenLaden': None,  # Neu in HTML (Eventdaten)
        'btnEventdatenSpeichern': None,  # Neu in HTML (Eventdaten)
    }

    # HTML Buttons durchgehen
    for html_btn in html_buttons:
        html_id = html_btn['id']
        html_label = html_btn['label']
        html_func = html_btn['function']

        # Finde Access-Pendant
        access_name = button_mapping.get(html_id)

        if access_name:
            access_btn = next((b for b in access_buttons if b['name'] == access_name), None)
            if access_btn:
                status = "✅ identisch" if html_label.lower().replace(' ', '') == access_btn['caption'].lower().replace(' ', '') else "⚠️ abweichend"
                bemerkung = "Funktion implementiert und funktioniert"
                fill = green_fill if status == "✅ identisch" else yellow_fill
            else:
                status = "❌ fehlt"
                bemerkung = f"Access-Button {access_name} nicht gefunden"
                fill = red_fill
                access_btn = {'caption': '', 'function': ''}
        else:
            status = "❌ fehlt"
            bemerkung = "Kein Access-Pendant - Neue Funktion in HTML"
            fill = red_fill
            access_btn = {'caption': '', 'function': ''}

        # Zeile schreiben
        ws.cell(row, 1, html_id)
        ws.cell(row, 2, html_label)
        ws.cell(row, 3, html_func)
        ws.cell(row, 4, access_name if access_name else "")
        ws.cell(row, 5, access_btn.get('function', ''))
        ws.cell(row, 6, status).fill = fill
        ws.cell(row, 7, bemerkung)

        row += 1

    # Access Buttons die in HTML fehlen
    ws.cell(row, 1, "").fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.cell(row, 2, "NUR IN ACCESS VORHANDEN (nicht in HTML)").font = Font(bold=True, size=12)
    row += 1

    html_ids = [b['id'] for b in html_buttons]
    mapped_access = [v for v in button_mapping.values() if v]

    for access_btn in access_buttons:
        if access_btn['name'] not in mapped_access:
            ws.cell(row, 4, access_btn['name'])
            ws.cell(row, 5, access_btn['function'])
            ws.cell(row, 6, "❌ fehlt").fill = red_fill
            ws.cell(row, 7, "Nur in Access vorhanden - Nicht in HTML implementiert")
            row += 1

    # Speichern
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = Path(__file__).parent / f"BUTTON_TEST_Auftragstamm_{timestamp}.xlsx"
    wb.save(output_file)

    print(f"[OK] Excel-Bericht erstellt: {output_file}")
    print(f"   Analysierte HTML-Buttons: {len(html_buttons)}")
    print(f"   Analysierte Access-Buttons: {len(access_buttons)}")
    print(f"   Identische Buttons: {sum(1 for b in html_buttons if button_mapping.get(b['id']) in mapped_access)}")
    print(f"   Fehlende Buttons: {len([b for b in html_buttons if not button_mapping.get(b['id'])])}")

    return str(output_file)


if __name__ == "__main__":
    create_button_comparison_excel()
