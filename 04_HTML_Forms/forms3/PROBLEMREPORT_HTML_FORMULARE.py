#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Erstellt Excel-Report mit allen Problemen in HTML-Formularen
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Erstelle Workbook
wb = Workbook()

# ============================================================================
# SHEET 1: Fehlende Button-Funktionen
# ============================================================================
ws1 = wb.active
ws1.title = "Fehlende Funktionen"

# Header-Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

# Headers
headers1 = ["Nr", "Formular", "Button-ID", "Fehlende Funktion", "Priorität", "Beschreibung", "Status"]
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Daten: Fehlende Funktionen
fehlende_funktionen = [
    # frm_va_Auftragstamm.html
    (1, "frm_va_Auftragstamm.html", "title-btn", "toggleMaximize()", "KRITISCH", "Fenster maximieren - Fenster-Bedienung", "OFFEN"),
    (2, "frm_va_Auftragstamm.html", "btnELGesendet", "showELGesendet()", "HOCH", "Einsatzliste Gesendet-Status anzeigen", "OFFEN"),
    (3, "frm_va_Auftragstamm.html", "cmd_BWN_send", "bwnSenden()", "MITTEL", "Bestätigung/BWN senden", "OFFEN"),
    (4, "frm_va_Auftragstamm.html", "Datensatz-Btn", "neuenAttachHinzufuegen()", "MITTEL", "Neuen Anhang hinzufügen", "OFFEN"),
    (5, "frm_va_Auftragstamm.html", "Datensatz-Btn", "rechnungPDF()", "HOCH", "Rechnung als PDF erstellen", "OFFEN"),
    (6, "frm_va_Auftragstamm.html", "Datensatz-Btn", "berechnungslistePDF()", "HOCH", "Berechnungsliste als PDF", "OFFEN"),
    (7, "frm_va_Auftragstamm.html", "Datensatz-Btn", "rechnungDatenLaden()", "MITTEL", "Rechnungsdaten laden", "OFFEN"),
    (8, "frm_va_Auftragstamm.html", "Datensatz-Btn", "rechnungLexware()", "MITTEL", "Lexware-Export Rechnung", "OFFEN"),
    (9, "frm_va_Auftragstamm.html", "Datensatz-Btn", "webDatenLaden()", "MITTEL", "Web-Daten laden", "OFFEN"),
    (10, "frm_va_Auftragstamm.html", "Datensatz-Btn", "eventdatenSpeichern()", "MITTEL", "Eventdaten speichern", "OFFEN"),
    # frm_MA_Mitarbeiterstamm.html
    (11, "frm_MA_Mitarbeiterstamm.html", "btnEinsatzÜbersicht", "openEinsatzübersicht()", "KRITISCH", "Einsatzübersicht öffnen", "OFFEN"),
    (12, "frm_MA_Mitarbeiterstamm.html", "btnZKFest", "btnZKFest_Click()", "HOCH", "Zeitkonto Festanstellung", "OFFEN"),
    (13, "frm_MA_Mitarbeiterstamm.html", "btnZKMini", "btnZKMini_Click()", "HOCH", "Zeitkonto Minijob", "OFFEN"),
    (14, "frm_MA_Mitarbeiterstamm.html", "btnZKeinzel", "btnZKeinzel_Click()", "HOCH", "Zeitkonto Einzelabrechnung", "OFFEN"),
    # frm_KD_Kundenstamm.html
    (15, "frm_KD_Kundenstamm.html", "Suche-Btn", "sucheKundeNr()", "KRITISCH", "Kunde nach Nummer suchen", "OFFEN"),
    (16, "frm_KD_Kundenstamm.html", "Nav-Btn", "gotoFirstRecord()", "KRITISCH", "Navigation: Erster Datensatz", "OFFEN"),
    (17, "frm_KD_Kundenstamm.html", "Nav-Btn", "gotoPrevRecord()", "KRITISCH", "Navigation: Vorheriger Datensatz", "OFFEN"),
    (18, "frm_KD_Kundenstamm.html", "Nav-Btn", "gotoNextRecord()", "KRITISCH", "Navigation: Nächster Datensatz", "OFFEN"),
    (19, "frm_KD_Kundenstamm.html", "Nav-Btn", "gotoLastRecord()", "KRITISCH", "Navigation: Letzter Datensatz", "OFFEN"),
    # frm_Einsatzuebersicht.html
    (20, "frm_Einsatzuebersicht.html", "btnClose", "closeForm()", "MITTEL", "Formular schließen", "OFFEN"),
]

for row_data in fehlende_funktionen:
    row_num = row_data[0] + 1
    for col, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=col, value=value)
        # Priorität farbig markieren
        if col == 5:
            if value == "KRITISCH":
                cell.fill = critical_fill
            elif value == "HOCH":
                cell.fill = high_fill
            elif value == "MITTEL":
                cell.fill = medium_fill

# Spaltenbreiten
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 40
ws1.column_dimensions['G'].width = 10

# ============================================================================
# SHEET 2: Layout-Probleme
# ============================================================================
ws2 = wb.create_sheet("Layout-Probleme")

headers2 = ["Nr", "Formular", "Problem", "Beschreibung", "Lösung", "Status"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

layout_probleme = [
    (1, "frm_va_Auftragstamm.html", "Title-Bar versteckt", ".title-bar { display: none; } - Header ist komplett versteckt", "display: none zu display: flex ändern", "OFFEN"),
    (2, "frm_DP_Dienstplan_MA.html", "Titel-Größe falsch", "font-size: 14px statt 15px", "font-size auf 15px ändern", "OFFEN"),
    (3, "Alle Hauptformulare", "Kein Footer", "Keine einheitliche Footer-Zeile vorhanden", "Optional: Footer-Komponente hinzufügen", "OPTIONAL"),
]

for row_data in layout_probleme:
    row_num = row_data[0] + 1
    for col, value in enumerate(row_data, 1):
        ws2.cell(row=row_num, column=col, value=value)

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 50
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 10

# ============================================================================
# SHEET 3: Fehlende API-Endpoints
# ============================================================================
ws3 = wb.create_sheet("API-Endpoints")

headers3 = ["Nr", "Endpoint", "Methode", "Beschreibung", "Verwendet in", "Status"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

api_probleme = [
    (1, "/api/kundenpreise", "GET", "Verrechnungssätze Pivot-Tabelle", "frm_KD_Verrechnungssaetze.html", "NEU IMPLEMENTIERT"),
    (2, "/api/einsatzliste/pdf", "POST", "Einsatzliste als PDF generieren", "frm_va_Auftragstamm.html", "FEHLT"),
    (3, "/api/rechnung/pdf", "POST", "Rechnung als PDF generieren", "frm_va_Auftragstamm.html", "FEHLT"),
    (4, "/api/berechnungsliste/pdf", "POST", "Berechnungsliste als PDF", "frm_va_Auftragstamm.html", "FEHLT"),
    (5, "/api/lexware/export", "POST", "Lexware-Export", "frm_va_Auftragstamm.html", "FEHLT"),
    (6, "/api/eventdaten", "GET/POST", "Eventdaten laden/speichern", "frm_va_Auftragstamm.html", "FEHLT"),
    (7, "/api/zeitkonto/festanstellung", "POST", "Zeitkonto Festanstellung berechnen", "frm_MA_Mitarbeiterstamm.html", "FEHLT"),
    (8, "/api/zeitkonto/minijob", "POST", "Zeitkonto Minijob berechnen", "frm_MA_Mitarbeiterstamm.html", "FEHLT"),
    (9, "/api/zeitkonto/einzelabrechnung", "POST", "Zeitkonto Einzelabrechnung", "frm_MA_Mitarbeiterstamm.html", "FEHLT"),
]

for row_data in api_probleme:
    row_num = row_data[0] + 1
    for col, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_num, column=col, value=value)
        if col == 6 and value == "FEHLT":
            cell.fill = critical_fill

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 40
ws3.column_dimensions['E'].width = 35
ws3.column_dimensions['F'].width = 20

# ============================================================================
# SHEET 4: Formular-Übersicht
# ============================================================================
ws4 = wb.create_sheet("Formular-Übersicht")

headers4 = ["Nr", "Formular", "Typ", "Buttons OK", "Buttons Fehlt", "Layout OK", "Sidebar", "Status"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

formular_uebersicht = [
    (1, "frm_va_Auftragstamm.html", "Hauptformular", 20, 10, "NEIN", "JA", "PROBLEME"),
    (2, "frm_MA_Mitarbeiterstamm.html", "Hauptformular", 18, 4, "JA", "JA", "PROBLEME"),
    (3, "frm_KD_Kundenstamm.html", "Hauptformular", 8, 5, "JA", "JA", "PROBLEME"),
    (4, "frm_OB_Objekt.html", "Hauptformular", 25, 0, "JA", "JA", "OK"),
    (5, "frm_Menuefuehrung1.html", "Hauptmenü", 30, 0, "JA", "JA", "OK"),
    (6, "frm_MA_Abwesenheit.html", "Nebenformular", 5, 0, "JA", "JA", "OK"),
    (7, "frm_MA_Zeitkonten.html", "Nebenformular", 9, 0, "JA", "JA", "OK"),
    (8, "frm_DP_Dienstplan_MA.html", "Planungsformular", 2, 0, "NEIN", "NEIN", "LAYOUT"),
    (9, "frm_DP_Dienstplan_Objekt.html", "Planungsformular", 0, 0, "JA", "NEIN", "OK"),
    (10, "frm_Einsatzuebersicht.html", "Nebenformular", 5, 1, "JA", "JA", "PROBLEME"),
    (11, "frm_VA_Planungsuebersicht.html", "Planungsformular", 7, 0, "JA", "NEIN", "OK"),
    (12, "frm_N_Bewerber.html", "Nebenformular", 2, 0, "JA", "JA", "OK"),
    (13, "frm_MA_VA_Schnellauswahl.html", "Nebenformular", 17, 0, "JA", "JA", "OK"),
    (14, "frm_KD_Verrechnungssaetze.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (15, "frm_Abwesenheiten.html", "Nebenformular", 4, 0, "JA", "JA", "OK"),
    (16, "frm_abwesenheitsuebersicht.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (17, "frm_Angebot.html", "Placeholder", 0, 0, "-", "-", "IN ENTWICKLUNG"),
    (18, "frm_Ausweis_Create.html", "Nebenformular", 2, 0, "JA", "JA", "OK"),
    (19, "frm_DP_Einzeldienstplaene.html", "Planungsformular", 2, 0, "JA", "NEIN", "OK"),
    (20, "frm_KD_Umsatzauswertung.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (21, "frm_Kundenpreise_gueni.html", "Nebenformular", 2, 0, "JA", "JA", "OK"),
    (22, "frm_MA_Adressen.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (23, "frm_MA_Offene_Anfragen.html", "Nebenformular", 4, 0, "JA", "JA", "OK"),
    (24, "frm_MA_Serien_eMail_Auftrag.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (25, "frm_MA_Serien_eMail_dienstplan.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (26, "frm_MA_Tabelle.html", "Nebenformular", 2, 0, "JA", "JA", "OK"),
    (27, "frm_MA_VA_Positionszuordnung.html", "Nebenformular", 5, 0, "JA", "JA", "OK"),
    (28, "frm_Rechnung.html", "Placeholder", 0, 0, "-", "-", "IN ENTWICKLUNG"),
    (29, "frm_Rueckmeldestatistik.html", "Nebenformular", 3, 0, "JA", "JA", "OK"),
    (30, "frm_Systeminfo.html", "Info", 1, 0, "JA", "JA", "OK"),
]

for row_data in formular_uebersicht:
    row_num = row_data[0] + 1
    for col, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_num, column=col, value=value)
        if col == 8:
            if value == "PROBLEME":
                cell.fill = critical_fill
            elif value == "LAYOUT":
                cell.fill = medium_fill
            elif value == "IN ENTWICKLUNG":
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 10
ws4.column_dimensions['H'].width = 18

# ============================================================================
# SHEET 5: Zusammenfassung
# ============================================================================
ws5 = wb.create_sheet("Zusammenfassung")

ws5.cell(row=1, column=1, value="PROBLEMREPORT HTML-FORMULARE").font = Font(bold=True, size=16)
ws5.cell(row=2, column=1, value=f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
ws5.cell(row=4, column=1, value="STATISTIK").font = Font(bold=True, size=14)

stats = [
    ("Gesamtzahl Formulare:", 30),
    ("Formulare OK:", 22),
    ("Formulare mit Problemen:", 5),
    ("Formulare in Entwicklung:", 3),
    ("", ""),
    ("Fehlende Button-Funktionen:", 20),
    ("- Kritisch:", 7),
    ("- Hoch:", 7),
    ("- Mittel:", 6),
    ("", ""),
    ("Layout-Probleme:", 3),
    ("Fehlende API-Endpoints:", 8),
]

for i, (label, value) in enumerate(stats, 5):
    ws5.cell(row=i, column=1, value=label)
    ws5.cell(row=i, column=2, value=value)

ws5.cell(row=18, column=1, value="PRIORITÄTEN").font = Font(bold=True, size=14)
ws5.cell(row=19, column=1, value="KRITISCH = Kernfunktionalität blockiert")
ws5.cell(row=19, column=1).fill = critical_fill
ws5.cell(row=20, column=1, value="HOCH = Wichtige Geschäftsprozesse betroffen")
ws5.cell(row=20, column=1).fill = high_fill
ws5.cell(row=21, column=1, value="MITTEL = Nice-to-have Funktionen")
ws5.cell(row=21, column=1).fill = medium_fill

ws5.column_dimensions['A'].width = 40
ws5.column_dimensions['B'].width = 15

# Speichern
output_path = r"C:\Users\guenther.siegert\Documents\0006_All_Access_KNOWLEDGE\04_HTML_Forms\forms3\PROBLEMREPORT_HTML_FORMULARE.xlsx"
wb.save(output_path)
print(f"Excel-Report erstellt: {output_path}")
