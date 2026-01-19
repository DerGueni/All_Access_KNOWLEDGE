"""
Erstellt Excel-Datei mit Abweichungen zwischen Access und HTML
MIT ANMERKUNGEN aus Screenshots
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Excel Workbook erstellen
wb = openpyxl.Workbook()

# Sheet 1: Zusammenfassung
ws_summary = wb.active
ws_summary.title = "Zusammenfassung"

# Header-Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

# Titel
ws_summary['A1'] = "Abweichungen Access ↔ HTML - Zusammenfassung"
ws_summary['A1'].font = Font(bold=True, size=16, color="1F4E78")
ws_summary.merge_cells('A1:D1')
ws_summary['A1'].alignment = Alignment(horizontal="center")

# Erstellungsdatum
ws_summary['A2'] = f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws_summary.merge_cells('A2:D2')
ws_summary['A2'].alignment = Alignment(horizontal="center")
ws_summary['A2'].font = Font(italic=True, size=10)

# Statistiken
row = 4
ws_summary[f'A{row}'] = "Statistik nach Schweregrad"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
ws_summary[f'A{row}'] = "Schweregrad"
ws_summary[f'B{row}'] = "Anzahl"
ws_summary[f'A{row}'].fill = header_fill
ws_summary[f'B{row}'].fill = header_fill
ws_summary[f'A{row}'].font = header_font
ws_summary[f'B{row}'].font = header_font
ws_summary[f'A{row}'].alignment = header_alignment
ws_summary[f'B{row}'].alignment = header_alignment

# Daten
severity_counts = {
    'Critical': 3,
    'High': 6,
    'Medium': 3,
    'Low': 0
}

severity_colors = {
    'Critical': 'FF0000',
    'High': 'FFA500',
    'Medium': 'FFFF00',
    'Low': '00B050'
}

for severity, count in severity_counts.items():
    row += 1
    ws_summary[f'A{row}'] = severity
    ws_summary[f'B{row}'] = count
    ws_summary[f'A{row}'].fill = PatternFill(start_color=severity_colors[severity], end_color=severity_colors[severity], fill_type="solid")
    ws_summary[f'B{row}'].alignment = Alignment(horizontal="center")

row += 2
ws_summary[f'A{row}'] = "Statistik nach Kategorie"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
ws_summary[f'A{row}'] = "Kategorie"
ws_summary[f'B{row}'] = "Anzahl"
ws_summary[f'A{row}'].fill = header_fill
ws_summary[f'B{row}'].fill = header_fill
ws_summary[f'A{row}'].font = header_font
ws_summary[f'B{row}'].font = header_font
ws_summary[f'A{row}'].alignment = header_alignment
ws_summary[f'B{row}'].alignment = header_alignment

category_counts = {
    'Optik': 3,
    'Struktur': 5,
    'Funktionen': 3,
    'Integration': 1
}

for category, count in category_counts.items():
    row += 1
    ws_summary[f'A{row}'] = category
    ws_summary[f'B{row}'] = count
    ws_summary[f'B{row}'].alignment = Alignment(horizontal="center")

row += 2
ws_summary[f'A{row}'] = "Betroffene Formulare"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
forms_affected = [
    'frm_va_Auftragstamm',
    'frm_MA_Mitarbeiterstamm',
    'frm_KD_Kundenstamm',
    'frm_MA_VA_Schnellauswahl',
    'frm_N_Dienstplanuebersicht',
    'frm_VA_Planungsuebersicht',
    'Alle Formulare (Backend)'
]

for form in forms_affected:
    row += 1
    ws_summary[f'A{row}'] = f"• {form}"

# Spaltenbreiten
ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 15

# Sheet 2: Detaillierte Liste MIT ANMERKUNGEN
ws_details = wb.create_sheet("Detaillierte Abweichungen")

# Header
headers = ['Nr.', 'Formular', 'Kategorie', 'Schweregrad', 'Beschreibung', 'Evidenz', 'Priorität', 'Status', 'Anmerkungen']
for col, header in enumerate(headers, 1):
    cell = ws_details.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Daten aus GAP_LIST.csv MIT ANMERKUNGEN
gaps = [
    {
        'nr': 1,
        'form': 'frm_va_Auftragstamm',
        'category': 'Optik',
        'severity': 'High',
        'issue': 'Header/Toolbar violett und Button-Leiste fehlen',
        'evidence': 'Screenshots/frm_va_Auftragstamm.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': 'Die ist so korrekt und muss in den anderen Formularen ebenfalls so umgesetzt werden'
    },
    {
        'nr': 2,
        'form': 'frm_va_Auftragstamm',
        'category': 'Struktur',
        'severity': 'High',
        'issue': 'Tabs Einsatzliste/Antworten/Rechnung fehlen oder nicht sichtbar',
        'evidence': 'Screenshots/frm_va_Auftragstamm.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': 'Müssen vorhanden und sichtbar sein'
    },
    {
        'nr': 3,
        'form': 'frm_va_Auftragstamm',
        'category': 'Struktur',
        'severity': 'High',
        'issue': 'Auftragsliste rechts anders/fehlende Spalten',
        'evidence': 'Screenshots/frm_va_Auftragstamm.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': 'Die fehlenden Spalten Soll ist und Status müssen vorhanden und sichtbar sein'
    },
    {
        'nr': 4,
        'form': 'frm_MA_Mitarbeiterstamm',
        'category': 'Optik',
        'severity': 'High',
        'issue': 'Header-Buttons (Löschen/Transfer/Listen) fehlen',
        'evidence': 'Screenshots/frm_ma_Mitarbeiterstamm.html.jpg',
        'priority': 'B',
        'status': 'Offen',
        'note': 'Nicht erforderlich'
    },
    {
        'nr': 5,
        'form': 'frm_MA_Mitarbeiterstamm',
        'category': 'Struktur',
        'severity': 'Medium',
        'issue': 'Such-/Filterzeile (Suche/Filter/Zusatz) abweichend',
        'evidence': 'Screenshots/frm_ma_Mitarbeiterstamm.html.jpg',
        'priority': 'B',
        'status': 'Offen',
        'note': 'Muss alles so wie in Access vorhanden und sichtbar sein'
    },
    {
        'nr': 6,
        'form': 'frm_KD_Kundenstamm',
        'category': 'Struktur',
        'severity': 'Medium',
        'issue': 'Listen-Spalten Kontaktname/Vorname fehlen',
        'evidence': 'Screenshots/frm_KD_Kundenstamm.html.jpg',
        'priority': 'B',
        'status': 'Offen',
        'note': 'Müssen vorhanden und sichtbar sein'
    },
    {
        'nr': 7,
        'form': 'frm_KD_Kundenstamm',
        'category': 'Optik',
        'severity': 'Medium',
        'issue': 'Header-Buttons/Abstände weichen ab',
        'evidence': 'Screenshots/frm_KD_Kundenstamm.html.jpg',
        'priority': 'B',
        'status': 'Offen',
        'note': 'Bitte umsetzen'
    },
    {
        'nr': 8,
        'form': 'frm_MA_VA_Schnellauswahl',
        'category': 'Struktur',
        'severity': 'High',
        'issue': 'Linkes Hauptmenü ersetzt durch Funktionsliste',
        'evidence': 'Screenshots/frm_ma_va_Schnellauswah.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': ''
    },
    {
        'nr': 9,
        'form': 'frm_MA_VA_Schnellauswahl',
        'category': 'Funktionen',
        'severity': 'High',
        'issue': 'Listenbereiche leer oder nicht verifiziert (geplant/zugesagt)',
        'evidence': 'Screenshots/frm_ma_va_Schnellauswah.html.jpg',
        'priority': 'A',
        'status': 'In Bearbeitung',
        'note': 'Muss alles so sein wie in Access'
    },
    {
        'nr': 10,
        'form': 'frm_N_Dienstplanuebersicht',
        'category': 'Funktionen',
        'severity': 'Critical',
        'issue': 'Kalender-Tabelle/Daten fehlen komplett',
        'evidence': 'Screenshots/frm_DP_Dienstplan_MA.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': 'HTML Formular löschen und durch frm_DP_Dienstplan_MA.HTML ersetzen (ist bereits vorhanden)'
    },
    {
        'nr': 11,
        'form': 'frm_VA_Planungsuebersicht',
        'category': 'Funktionen',
        'severity': 'Critical',
        'issue': 'Planungstabelle bleibt leer',
        'evidence': 'Screenshots/frm_DP_Dienstplan_Objekt.html.jpg',
        'priority': 'A',
        'status': 'Offen',
        'note': 'HTML Formular löschen und durch frm_DP_Dienstplan_Objekt.HTML ersetzen (ist bereits vorhanden)'
    },
    {
        'nr': 12,
        'form': 'Alle Backend-Formulare',
        'category': 'Integration',
        'severity': 'Critical',
        'issue': 'Bridge-Import-Pfad war falsch (keine Datenbindung)',
        'evidence': 'forms/_Codes/logic/*.js',
        'priority': 'A',
        'status': 'Behoben',
        'note': ''
    }
]

# Daten einfügen
for row_idx, gap in enumerate(gaps, 2):
    ws_details.cell(row_idx, 1, gap['nr'])
    ws_details.cell(row_idx, 2, gap['form'])
    ws_details.cell(row_idx, 3, gap['category'])
    ws_details.cell(row_idx, 4, gap['severity'])
    ws_details.cell(row_idx, 5, gap['issue'])
    ws_details.cell(row_idx, 6, gap['evidence'])
    ws_details.cell(row_idx, 7, gap['priority'])
    ws_details.cell(row_idx, 8, gap['status'])
    ws_details.cell(row_idx, 9, gap['note'])

    # Farbe basierend auf Schweregrad
    severity_cell = ws_details.cell(row_idx, 4)
    severity_cell.fill = PatternFill(start_color=severity_colors[gap['severity']],
                                     end_color=severity_colors[gap['severity']],
                                     fill_type="solid")

    # Status-Farbe
    status_cell = ws_details.cell(row_idx, 8)
    if gap['status'] == 'Behoben':
        status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")
    elif gap['status'] == 'In Bearbeitung':
        status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type="solid")
    else:
        status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")

    # Anmerkung in ROT wenn vorhanden
    note_cell = ws_details.cell(row_idx, 9)
    if gap['note']:
        note_cell.font = Font(color='FF0000', bold=True)

    # Border
    for col in range(1, 10):
        cell = ws_details.cell(row_idx, col)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Spaltenbreiten
ws_details.column_dimensions['A'].width = 5
ws_details.column_dimensions['B'].width = 30
ws_details.column_dimensions['C'].width = 12
ws_details.column_dimensions['D'].width = 12
ws_details.column_dimensions['E'].width = 50
ws_details.column_dimensions['F'].width = 35
ws_details.column_dimensions['G'].width = 10
ws_details.column_dimensions['H'].width = 15
ws_details.column_dimensions['I'].width = 70  # Anmerkungen

# Zeile 1 fixieren (Freeze Panes)
ws_details.freeze_panes = 'A2'

# Auto-Filter
ws_details.auto_filter.ref = f"A1:I{len(gaps) + 1}"

# Sheet 3: Checkliste Auftragstamm MIT ANMERKUNGEN
ws_checklist = wb.create_sheet("Checkliste Auftragstamm")

ws_checklist['A1'] = "Checkliste: frm_va_Auftragstamm"
ws_checklist['A1'].font = Font(bold=True, size=14)
ws_checklist.merge_cells('A1:D1')

checklist_items = [
    ('Header-Bar', 'Gradient #5b3bd2 → #2a0d6e, Höhe 110px', 'Offen', 'Nicht umsetzen'),
    ('Toolbar Buttons', 'Schrift Segoe UI 9pt, 133×24 px', 'Offen', 'Nicht umsetzen'),
    ('Tab-Header', 'Einsatzliste/Antworten/Rechnung, Farbe #e8e8e8', 'Offen', 'Nicht umsetzen'),
    ('Sidebar Buttons', 'Farbe und Größe wie Access', 'Offen', 'Nicht umsetzen'),
    ('Einsatzliste Tabelle', 'Spaltenbreiten, Zeilen-Streifen', 'Offen', 'Nicht umsetzen'),
    ('Buttons unter Tabelle', 'Farbverlauf #95b3d7', 'Offen', 'Nicht umsetzen'),
    ('API Datenbindung', 'Bridge-Anfragen funktional', 'Teilweise', 'Muss funktionieren')
]

row = 3
ws_checklist.cell(row, 1, 'Element').font = header_font
ws_checklist.cell(row, 2, 'Anforderung').font = header_font
ws_checklist.cell(row, 3, 'Status').font = header_font
ws_checklist.cell(row, 4, 'Anmerkung').font = header_font

for col in range(1, 5):
    ws_checklist.cell(row, col).fill = header_fill
    ws_checklist.cell(row, col).alignment = header_alignment

for item in checklist_items:
    row += 1
    ws_checklist.cell(row, 1, item[0])
    ws_checklist.cell(row, 2, item[1])
    ws_checklist.cell(row, 3, item[2])
    ws_checklist.cell(row, 4, item[3])

    status = item[2]
    status_cell = ws_checklist.cell(row, 3)
    if status == 'Behoben':
        status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")
    elif status == 'Teilweise':
        status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type="solid")
    else:
        status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")

    # Anmerkung in ROT
    note_cell = ws_checklist.cell(row, 4)
    note_cell.font = Font(color='FF0000', bold=True)

ws_checklist.column_dimensions['A'].width = 30
ws_checklist.column_dimensions['B'].width = 50
ws_checklist.column_dimensions['C'].width = 15
ws_checklist.column_dimensions['D'].width = 25

# Speichern
output_path = r"C:\Users\guenther.siegert\Desktop\Abweichungen_Access_HTML.xlsx"
wb.save(output_path)
print(f"Excel-Datei erstellt: {output_path}")
