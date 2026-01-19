"""
Erstellt Excel-Liste mit allen Abweichungen aus der erweiterten Senior Master Agent Prüfung
Teil 2: Auftragstamm (detailliert), Dienstplanübersicht, Planungsübersicht, Verrechnungssätze, MA page subrch, Anfragen-Panel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Workbook erstellen
wb = Workbook()
ws = wb.active
ws.title = "Abweichungen Teil 2"

# Styles definieren
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
orange_fill = PatternFill(start_color="FF9F43", end_color="FF9F43", fill_type="solid")
yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
gray_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header
headers = [
    "Nr", "Formular/Bereich", "Problem-Kategorie", "Problem", "Schweregrad",
    "Datei", "Details", "Lösungsvorschlag", "Aktion", "Status"
]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Daten - Alle gefundenen Abweichungen aus Teil 2 Prüfung
abweichungen = [
    # === AUFTRAGSTAMM - Buttons und Bedingte Formatierungen ===
    (1, "frm_va_Auftragstamm", "Bedingte Formatierung",
     "Soll>Ist Rot-Markierung fehlt: Wenn MA_Anzahl_Ist < MA_Anzahl soll Zeile rot sein",
     "Hoch", "frm_va_Auftragstamm.html", "Zeile 4500-4600",
     "CSS-Klasse 'soll-nicht-erfuellt' bei Soll>Ist hinzufügen + rot einfärben", "BEHEBEN", "Offen"),

    (2, "frm_va_Auftragstamm", "Bedingte Formatierung",
     "Status-Farben in Auftragsliste fehlen: 'In Planung'=gelb, 'Beendet'=grün, etc.",
     "Mittel", "frm_va_Auftragstamm.html", "renderAuftragsliste()",
     "Status-abhängige Hintergrundfarben wie in Access", "BEHEBEN", "Offen"),

    (3, "frm_va_Auftragstamm", "Fehlender Button",
     "btnAuftrBerech (Berechnung) fehlt im HTML - VBA: Form_frm_VA_Auftragstamm.bas Zeile 892",
     "Mittel", "frm_va_Auftragstamm.html", "Button-Leiste",
     "Button hinzufügen mit onclick → VBA Bridge Call", "BEHEBEN", "Offen"),

    (4, "frm_va_Auftragstamm", "Fehlender Button",
     "btnExcel (Excel-Export) fehlt - VBA: btnExcel_Click exportiert Auftragsliste",
     "Niedrig", "frm_va_Auftragstamm.html", "Toolbar",
     "Button hinzufügen, SheetJS für Excel-Export nutzen", "IGNORIEREN", "Nice-to-have"),

    (5, "frm_va_Auftragstamm", "Tab Zusatzdateien",
     "Upload-Funktion fehlt - nur Download implementiert",
     "Mittel", "frm_va_Auftragstamm.html", "Tab Zusatzdateien",
     "REST-API POST /api/auftraege/<id>/dateien für Upload", "BEHEBEN", "Offen"),

    (6, "frm_va_Auftragstamm", "Tab Rechnung",
     "API-Endpoint /api/auftraege/<id>/rechnungen fehlt",
     "Hoch", "api_server.py", "Neuer Endpoint",
     "SELECT * FROM tbl_Rch_Kopf WHERE VA_ID=?", "BEHEBEN", "Offen"),

    (7, "frm_va_Auftragstamm", "Tab Kosten",
     "API-Endpoint /api/auftraege/<id>/kosten fehlt",
     "Hoch", "api_server.py", "Neuer Endpoint",
     "SELECT * FROM tbl_VA_Kosten WHERE VA_ID=?", "BEHEBEN", "Offen"),

    (8, "frm_va_Auftragstamm", "Feldabhängigkeit",
     "Objekt-Auswahl aktualisiert nicht automatisch Ansprechpartner-Dropdown",
     "Mittel", "frm_va_Auftragstamm.logic.js", "onchange Objekt",
     "Bei Objekt-Änderung Ansprechpartner neu laden", "BEHEBEN", "Offen"),

    (9, "frm_va_Auftragstamm", "Filter",
     "Datumsbereich-Filter (Von/Bis oben) filtert Auftragsliste nicht",
     "Mittel", "frm_va_Auftragstamm.logic.js", "applyDateFilter()",
     "Datum-Filter muss loadAuftragsliste() mit Params aufrufen", "BEHEBEN", "Offen"),

    (10, "frm_va_Auftragstamm", "Filter",
     "Textsuche durchsucht nicht alle Felder (nur Auftrag, nicht Veranstalter/Objekt)",
     "Niedrig", "frm_va_Auftragstamm.logic.js", "filterAuftraege()",
     "Suchtext auch in Veranstalter und Objekt suchen", "IGNORIEREN", "Nice-to-have"),

    (11, "frm_va_Auftragstamm", "Buttons",
     "cmdAuftragKopieren zeigt keinen Bestätigungsdialog wie in Access",
     "Niedrig", "frm_va_Auftragstamm.html", "onclick cmdAuftragKopieren",
     "confirm() Dialog vor Kopieren hinzufügen", "IGNORIEREN", "UX-Verbesserung"),

    (12, "frm_va_Auftragstamm", "Tabs",
     "Tab 'Eventdaten' Scraper-Fehlerbehandlung fehlt bei Timeout",
     "Niedrig", "frm_va_Auftragstamm.html", "loadEventdaten()",
     "try/catch mit Timeout-Meldung an User", "IGNORIEREN", "Robustheit"),

    (13, "frm_va_Auftragstamm", "Subform",
     "sub_VA_Tag Kommunikation: Ausgewählter Tag wird nicht an Parent gemeldet",
     "Hoch", "sub_VA_Tag.html", "postMessage",
     "Bei Tag-Klick: parent.postMessage({type:'TAG_SELECTED', datum: ...})", "BEHEBEN", "Offen"),

    (14, "frm_va_Auftragstamm", "Subform",
     "sub_MA_VA_Zuordnung: Doppelklick auf MA öffnet nicht Mitarbeiterstamm",
     "Mittel", "sub_MA_VA_Zuordnung.html", "dblclick Handler",
     "dblclick → Shell-Navigation zu frm_MA_Mitarbeiterstamm mit MA_ID", "BEHEBEN", "Offen"),

    (15, "frm_va_Auftragstamm", "Performance",
     "Auftragsliste lädt alle Aufträge (kein Lazy Loading bei >500 Einträgen)",
     "Niedrig", "frm_va_Auftragstamm.logic.js", "loadAuftragsliste()",
     "Virtual Scrolling oder Pagination bei großen Listen", "IGNORIEREN", "Performance"),

    # === DIENSTPLANÜBERSICHT ===
    (16, "frm_N_Dienstplanuebersicht", "KRITISCH - Datei fehlt",
     "HTML-Datei existiert NICHT in forms3 Ordner!",
     "Kritisch", "forms3/frm_N_Dienstplanuebersicht.html", "FEHLT KOMPLETT",
     "Formular aus Access exportieren und HTML erstellen", "BEHEBEN", "DRINGEND"),

    (17, "frm_N_Dienstplanuebersicht", "VBA Bridge",
     "Endpoints für Dienstplan-Operationen fehlen in vba_bridge.py",
     "Hoch", "vba_bridge.py", "allowed_functions",
     "DP_Anzeigen, DP_Drucken, DP_Senden zu Whitelist hinzufügen", "BEHEBEN", "Offen"),

    (18, "frm_N_Dienstplanuebersicht", "API",
     "/api/dienstplan/uebersicht Endpoint fehlt",
     "Hoch", "api_server.py", "Neuer Endpoint",
     "Aggregierte Dienstplan-Daten pro Objekt/Datum", "BEHEBEN", "Offen"),

    # === PLANUNGSÜBERSICHT ===
    (19, "frm_VA_Planungsuebersicht", "KRITISCH - Bridge.query fehlt",
     "Bridge.query() Methode existiert nicht - Formular lädt keine Daten",
     "Kritisch", "frm_VA_Planungsuebersicht.html", "Zeile ~1200",
     "Bridge.query → fetch('/api/...') ersetzen", "BEHEBEN", "DRINGEND"),

    (20, "frm_VA_Planungsuebersicht", "Duplikat",
     "2 Formulare existieren: frm_VA_Planungsuebersicht.html UND alternatives in forms/",
     "Mittel", "forms3/ und forms/", "Duplikat prüfen",
     "Eines als kanonisch definieren, anderes löschen", "PRÜFEN", "Aufräumen"),

    (21, "frm_VA_Planungsuebersicht", "Kalender",
     "Kalender-Widget verwendet veraltete jQuery-UI Datepicker",
     "Niedrig", "frm_VA_Planungsuebersicht.html", "CSS/JS Imports",
     "Native HTML5 date input oder moderne Lib verwenden", "IGNORIEREN", "Modernisierung"),

    (22, "frm_VA_Planungsuebersicht", "Filter",
     "Mitarbeiter-Filter lädt alle MA statt nur aktive",
     "Mittel", "frm_VA_Planungsuebersicht.logic.js", "loadMitarbeiter()",
     "?aktiv=true Parameter an API übergeben", "BEHEBEN", "Offen"),

    # === VERRECHNUNGSSÄTZE ===
    (23, "frm_KD_Verrechnungssaetze", "KRITISCH - Falsche Tabelle",
     "API verwendet tbl_KD_Kundenpreise - Tabelle existiert NICHT!",
     "Kritisch", "api_server.py", "/api/verrechnungssaetze",
     "Tabelle ist tbl_KD_Standardpreise oder tbl_KD_Verrechnungssaetze", "BEHEBEN", "DRINGEND"),

    (24, "frm_KD_Verrechnungssaetze", "Felder",
     "Feldnamen im HTML stimmen nicht mit DB überein",
     "Hoch", "frm_KD_Verrechnungssaetze.html", "data-field Attribute",
     "Feldnamen aus tbl_KD_Standardpreise Schema übernehmen", "BEHEBEN", "Offen"),

    (25, "frm_KD_Verrechnungssaetze", "CRUD",
     "Nur Lesen implementiert - Bearbeiten/Löschen fehlt",
     "Mittel", "frm_KD_Verrechnungssaetze.html", "Buttons",
     "PUT/DELETE Endpoints und onclick Handler hinzufügen", "BEHEBEN", "Offen"),

    (26, "frm_KD_Verrechnungssaetze", "Kundenfilter",
     "Dropdown 'Kunde' filtert Liste nicht",
     "Mittel", "frm_KD_Verrechnungssaetze.logic.js", "onchange cboKunde",
     "Bei Kunde-Auswahl Liste neu laden mit KD_ID Filter", "BEHEBEN", "Offen"),

    # === MA PAGE SUBRCH (Separates Formular) ===
    (27, "sub_Rch_Kopf (MA)", "NEUES FORMULAR ERFORDERLICH",
     "Page subrch muss als separates HTML-Formular extrahiert werden",
     "Hoch", "NEUES FORMULAR", "frm_MA_Subrch.html erstellen",
     "Alle Controls, Events, Eigenschaften aus Access übernehmen", "BEHEBEN", "Neu"),

    (28, "sub_Rch_Kopf (MA)", "API Endpoints",
     "7 neue API-Endpoints für MA-Rechnungs-Subform benötigt",
     "Hoch", "api_server.py", "Neue Endpoints",
     "/api/ma/<id>/rechnungen, /api/ma/<id>/rechnungen/<rid>/positionen, etc.", "BEHEBEN", "Offen"),

    (29, "sub_Rch_Kopf (MA)", "VBA Events",
     "~15 VBA Events müssen nach JavaScript portiert werden",
     "Hoch", "Form_sub_Rch_Kopf.bas", "Event-Handler",
     "AfterUpdate, OnClick, OnDblClick Events in JS implementieren", "BEHEBEN", "Offen"),

    (30, "sub_Rch_Kopf (MA)", "Subform Einbettung",
     "frm_MA_Mitarbeiterstamm muss iframe für sub_Rch_Kopf einbetten",
     "Mittel", "frm_MA_Mitarbeiterstamm.html", "Tab Rechnungen",
     "<iframe src='sub_Rch_Kopf.html?ma_id=...'> einbetten", "BEHEBEN", "Offen"),

    (31, "sub_Rch_Kopf (MA)", "PostMessage",
     "Kommunikation Parent<->Subform via postMessage definieren",
     "Mittel", "sub_Rch_Kopf.html", "message Handler",
     "MA_SELECTED, REFRESH_DATA, DATA_CHANGED Events", "BEHEBEN", "Offen"),

    # === ANFRAGEN-PANEL (VBA Bridge Integration) ===
    (32, "Anfragen-Panel", "KRITISCH - Falsche API",
     "Panel verwendet REST-API statt VBA Bridge - kann keine echten E-Mails senden!",
     "Kritisch", "frm_va_Auftragstamm.html", "sendAnfrage()",
     "Von fetch('/api/anfragen') auf VBA Bridge Port 5002 umstellen", "BEHEBEN", "DRINGEND"),

    (33, "Anfragen-Panel", "VBA Funktion",
     "zmd_Mail.Anfragen() muss über VBA Bridge aufgerufen werden",
     "Kritisch", "frm_va_Auftragstamm.logic.js", "sendAnfrage()",
     "fetch('http://localhost:5002/api/vba/anfragen') wie in Schnellauswahl", "BEHEBEN", "DRINGEND"),

    (34, "Anfragen-Panel", "Parameter",
     "ma_id, va_id, vadatum_id, vastart_id müssen korrekt übergeben werden",
     "Hoch", "frm_va_Auftragstamm.logic.js", "sendAnfrage()",
     "Alle 4 Parameter aus state.current* extrahieren", "BEHEBEN", "Offen"),

    (35, "Anfragen-Panel", "Erfolgs-Feedback",
     "Keine Rückmeldung ob E-Mail erfolgreich gesendet",
     "Mittel", "frm_va_Auftragstamm.html", "sendAnfrage callback",
     "Bei response.success → Status in Zeile aktualisieren + Toast", "BEHEBEN", "Offen"),

    (36, "Anfragen-Panel", "Fehler-Handling",
     "Bei VBA Bridge Fehler keine Fehlermeldung an User",
     "Mittel", "frm_va_Auftragstamm.logic.js", "catch Block",
     "try/catch mit User-freundlicher Fehlermeldung", "BEHEBEN", "Offen"),

    (37, "Anfragen-Panel", "Batch-Anfragen",
     "Mehrere MA gleichzeitig anfragen funktioniert nicht",
     "Niedrig", "frm_va_Auftragstamm.html", "btnAnfragenAlle",
     "Loop über ausgewählte MA, sequentiell VBA Bridge aufrufen", "IGNORIEREN", "Enhancement"),
]

# Daten einfügen
for row_idx, row_data in enumerate(abweichungen, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Schweregrad farbig markieren
        if col_idx == 5:  # Schweregrad
            if value == "Kritisch":
                cell.fill = orange_fill
                cell.font = Font(bold=True)
            elif value == "Hoch":
                cell.fill = red_fill
            elif value == "Mittel":
                cell.fill = yellow_fill
            elif value == "Niedrig":
                cell.fill = green_fill

        # Aktion farbig markieren
        if col_idx == 9:  # Aktion
            if value == "BEHEBEN":
                cell.fill = red_fill
                cell.font = Font(bold=True)
            elif value == "PRÜFEN":
                cell.fill = yellow_fill
            elif value == "IGNORIEREN":
                cell.fill = gray_fill
            elif value == "ERLEDIGT":
                cell.fill = green_fill

# Spaltenbreiten anpassen
column_widths = [5, 28, 22, 55, 12, 30, 25, 55, 12, 15]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Zeile 1 fixieren (Header)
ws.freeze_panes = "A2"

# Autofilter aktivieren
ws.auto_filter.ref = f"A1:J{len(abweichungen)+1}"

# Zusammenfassung Sheet
ws2 = wb.create_sheet("Zusammenfassung")
ws2["A1"] = "ZUSAMMENFASSUNG - Erweiterte Prüfung Teil 2"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws2["A4"] = "Prüfbereiche: Auftragstamm (detailliert), Dienstplanübersicht, Planungsübersicht, Verrechnungssätze, MA page subrch, Anfragen-Panel"

ws2["A6"] = "Bereich"
ws2["B6"] = "Gesamt"
ws2["C6"] = "Kritisch"
ws2["D6"] = "Hoch"
ws2["E6"] = "Mittel"
ws2["F6"] = "Niedrig"
ws2["G6"] = "BEHEBEN"
ws2["H6"] = "PRÜFEN"
ws2["I6"] = "IGNORIEREN"

for col in range(1, 10):
    ws2.cell(row=6, column=col).font = Font(bold=True)
    ws2.cell(row=6, column=col).fill = header_fill
    ws2.cell(row=6, column=col).font = Font(bold=True, color="FFFFFF")

summary_data = [
    ("Auftragstamm (Buttons/Format)", 15, 0, 3, 6, 6, 9, 0, 6),
    ("Dienstplanübersicht", 3, 1, 2, 0, 0, 3, 0, 0),
    ("Planungsübersicht", 4, 1, 0, 2, 1, 2, 1, 1),
    ("Verrechnungssätze", 4, 1, 1, 2, 0, 4, 0, 0),
    ("MA page subrch", 5, 0, 3, 2, 0, 5, 0, 0),
    ("Anfragen-Panel", 6, 2, 1, 2, 1, 5, 0, 1),
    ("GESAMT", 37, 5, 10, 14, 8, 28, 1, 8),
]

for row_idx, data in enumerate(summary_data, 7):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 13:  # Gesamt-Zeile
            cell.font = Font(bold=True)
            if col_idx > 1:
                cell.fill = gray_fill

# Kritisch-Zeile hervorheben
for row in range(7, 14):
    kritisch_cell = ws2.cell(row=row, column=3)
    if kritisch_cell.value and kritisch_cell.value > 0:
        kritisch_cell.fill = orange_fill
        kritisch_cell.font = Font(bold=True)

ws2.column_dimensions["A"].width = 30
for col in "BCDEFGHI":
    ws2.column_dimensions[col].width = 12

# Prioritäten Sheet
ws3 = wb.create_sheet("Prioritäten")
ws3["A1"] = "PRIORITÄTEN - Was zuerst beheben?"
ws3["A1"].font = Font(bold=True, size=14)

ws3["A3"] = "DRINGEND (Kritisch - System funktioniert nicht):"
ws3["A3"].font = Font(bold=True)
ws3["A3"].fill = orange_fill

dringend = [
    "1. #16: Dienstplanübersicht HTML-Datei fehlt komplett",
    "2. #19: Planungsübersicht Bridge.query() existiert nicht",
    "3. #23: Verrechnungssätze falsche Tabelle (existiert nicht)",
    "4. #32-33: Anfragen-Panel nutzt nicht VBA Bridge (keine echten E-Mails!)",
]
for i, item in enumerate(dringend, 4):
    ws3[f"A{i}"] = item

ws3["A9"] = "HOCH (Wichtige Funktionen fehlen):"
ws3["A9"].font = Font(bold=True)
ws3["A9"].fill = red_fill

hoch = [
    "5. #1: Soll>Ist Rot-Markierung",
    "6. #6-7: Tab Rechnung/Kosten API-Endpoints",
    "7. #13: sub_VA_Tag Kommunikation",
    "8. #27-31: MA page subrch als neues Formular",
    "9. #17-18: Dienstplan VBA Bridge + API",
]
for i, item in enumerate(hoch, 10):
    ws3[f"A{i}"] = item

ws3["A16"] = "MITTEL (Funktionalität eingeschränkt):"
ws3["A16"].font = Font(bold=True)
ws3["A16"].fill = yellow_fill

mittel = [
    "10. #2: Status-Farben Auftragsliste",
    "11. #8: Objekt→Ansprechpartner Abhängigkeit",
    "12. #9: Datumsbereich-Filter",
    "13. #14: Doppelklick öffnet MA-Stamm",
    "14. #35-36: Anfragen Feedback/Fehler-Handling",
]
for i, item in enumerate(mittel, 17):
    ws3[f"A{i}"] = item

ws3.column_dimensions["A"].width = 80

# Speichern
output_path = r"C:\Users\guenther.siegert\Desktop\ABWEICHUNGEN_HTML_FORMS_TEIL2_18012026.xlsx"
wb.save(output_path)
print(f"Excel-Datei erstellt: {output_path}")
print(f"Gesamt: 37 Abweichungen")
print(f"- Kritisch: 5")
print(f"- Hoch: 10")
print(f"- Mittel: 14")
print(f"- Niedrig: 8")
print(f"- BEHEBEN: 28")
print(f"- PRÜFEN: 1")
print(f"- IGNORIEREN: 8")
