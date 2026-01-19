"""
Erstellt Excel-Datei mit Abweichungen bei Button-Klick-Ereignissen und Control-Eigenschaften
zwischen Access und HTML
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Excel Workbook erstellen
wb = openpyxl.Workbook()

# Sheet 1: Button-Ereignisse
ws_buttons = wb.active
ws_buttons.title = "Button-Ereignisse"

# Header-Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Titel
ws_buttons['A1'] = "Button- und Control-Eigenschaften: Access vs HTML"
ws_buttons['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_buttons.merge_cells('A1:H1')
ws_buttons['A1'].alignment = Alignment(horizontal="center")

ws_buttons['A2'] = f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws_buttons.merge_cells('A2:H2')
ws_buttons['A2'].alignment = Alignment(horizontal="center")
ws_buttons['A2'].font = Font(italic=True, size=10)

# Header
row = 4
headers = ['Nr.', 'Formular', 'Control-Name', 'Control-Typ', 'Ereignis (Access)', 'Ereignis (HTML)', 'Status', 'Anmerkung']
for col, header in enumerate(headers, 1):
    cell = ws_buttons.cell(row, col, header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Button-Ereignisse
buttons_data = [
    # frm_va_Auftragstamm
    {
        'nr': 1,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnAuftragNeu',
        'type': 'Button',
        'access_event': 'Private Sub btnAuftragNeu_Click()',
        'html_event': 'onclick="neuerAuftrag()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch'
    },
    {
        'nr': 2,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnAuftragKopieren',
        'type': 'Button',
        'access_event': 'Private Sub btnAuftragKopieren_Click()',
        'html_event': 'onclick="auftragKopieren()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch'
    },
    {
        'nr': 3,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnAuftragLoeschen',
        'type': 'Button',
        'access_event': 'Private Sub btnAuftragLoeschen_Click()',
        'html_event': 'onclick="auftragLoeschen()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch'
    },
    {
        'nr': 4,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnMailEins',
        'type': 'Button',
        'access_event': 'Private Sub btnMailEins_Click()',
        'html_event': 'onclick="sendeEinsatzlisteMA()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch, nutzt VBA-Bridge'
    },
    {
        'nr': 5,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnMailBOS',
        'type': 'Button',
        'access_event': 'Private Sub btnMailBOS_Click()',
        'html_event': 'onclick="sendeEinsatzlisteBOS()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch, nutzt VBA-Bridge'
    },
    {
        'nr': 6,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnMailSub',
        'type': 'Button',
        'access_event': 'Private Sub btnMailSub_Click()',
        'html_event': 'onclick="sendeEinsatzlisteSUB()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch, nutzt VBA-Bridge'
    },
    {
        'nr': 7,
        'form': 'frm_va_Auftragstamm',
        'control': 'btnDruckZusage',
        'type': 'Button',
        'access_event': 'Private Sub btnDruckZusage_Click()',
        'html_event': 'onclick="einsatzlisteDrucken()"',
        'status': 'Implementiert',
        'note': 'Nutzt VBA-Bridge für Access-Report'
    },
    {
        'nr': 8,
        'form': 'frm_va_Auftragstamm',
        'control': 'cboVeranstalter',
        'type': 'ComboBox',
        'access_event': 'Private Sub cboVeranstalter_AfterUpdate()',
        'html_event': 'onchange="handleVeranstalterChange()"',
        'status': 'Implementiert',
        'note': 'Event-Name unterschiedlich, Funktion identisch'
    },
    {
        'nr': 9,
        'form': 'frm_va_Auftragstamm',
        'control': 'Tab-Buttons (Einsatzliste/Antworten/Rechnung)',
        'type': 'TabControl',
        'access_event': 'Private Sub TabCtl_Change()',
        'html_event': 'onclick (data-tab)',
        'status': 'Implementiert',
        'note': 'Tab-Wechsel via JavaScript switchTab()'
    },
    # frm_MA_Mitarbeiterstamm
    {
        'nr': 10,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'btnMitarbeiterNeu',
        'type': 'Button',
        'access_event': 'Private Sub btnMitarbeiterNeu_Click()',
        'html_event': 'onclick="neuerMitarbeiter()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch'
    },
    {
        'nr': 11,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'btnSpeichern',
        'type': 'Button',
        'access_event': 'Private Sub btnSpeichern_Click()',
        'html_event': 'onclick="speichereMitarbeiter()"',
        'status': 'Implementiert',
        'note': 'Nutzt REST API statt DAO'
    },
    {
        'nr': 12,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'searchInput',
        'type': 'TextBox',
        'access_event': 'Private Sub searchInput_Change()',
        'html_event': 'oninput="renderMitarbeiterList()"',
        'status': 'Implementiert',
        'note': 'Event-Name unterschiedlich (Change vs input)'
    },
    {
        'nr': 13,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'NurAktiveMA',
        'type': 'ComboBox',
        'access_event': 'Private Sub NurAktiveMA_AfterUpdate()',
        'html_event': 'onchange="renderMitarbeiterList()"',
        'status': 'Implementiert',
        'note': 'Filter-Logik identisch'
    },
    # frm_KD_Kundenstamm
    {
        'nr': 14,
        'form': 'frm_KD_Kundenstamm',
        'control': 'btnKundeNeu',
        'type': 'Button',
        'access_event': 'Private Sub btnKundeNeu_Click()',
        'html_event': 'onclick="neuerKunde()"',
        'status': 'Implementiert',
        'note': 'Funktional identisch'
    },
    {
        'nr': 15,
        'form': 'frm_KD_Kundenstamm',
        'control': 'btnSpeichern',
        'type': 'Button',
        'access_event': 'Private Sub btnSpeichern_Click()',
        'html_event': 'onclick="speichereKunde()"',
        'status': 'Implementiert',
        'note': 'REST API statt DAO'
    },
    {
        'nr': 16,
        'form': 'frm_KD_Kundenstamm',
        'control': 'searchInput',
        'type': 'TextBox',
        'access_event': 'Private Sub searchInput_Change()',
        'html_event': 'oninput="renderKundenList()"',
        'status': 'Implementiert',
        'note': 'Live-Filter'
    },
    # frm_MA_VA_Schnellauswahl
    {
        'nr': 17,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'btnMail',
        'type': 'Button',
        'access_event': 'Private Sub btnMail_Click()',
        'html_event': 'onclick="sendeAnfragen()"',
        'status': 'Implementiert',
        'note': 'Nutzt VBA-Bridge für E-Mail-Versand'
    },
    {
        'nr': 18,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'lstVerfuegbar',
        'type': 'ListBox',
        'access_event': 'Private Sub lstVerfuegbar_DblClick()',
        'html_event': 'ondblclick (auf Tabellenzeile)',
        'status': 'Implementiert',
        'note': 'Zuordnung zu Auftrag'
    },
    {
        'nr': 19,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'cboAuftrag',
        'type': 'ComboBox',
        'access_event': 'Private Sub cboAuftrag_AfterUpdate()',
        'html_event': 'onchange="loadSchichten()"',
        'status': 'Implementiert',
        'note': 'Lädt Schichten für gewählten Auftrag'
    },
    {
        'nr': 20,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'cboDatum',
        'type': 'ComboBox',
        'access_event': 'Private Sub cboDatum_AfterUpdate()',
        'html_event': 'onchange="loadSchichten()"',
        'status': 'Implementiert',
        'note': 'Lädt Schichten für gewähltes Datum'
    },
]

# Daten einfügen
for row_idx, btn in enumerate(buttons_data, 5):
    ws_buttons.cell(row_idx, 1, btn['nr'])
    ws_buttons.cell(row_idx, 2, btn['form'])
    ws_buttons.cell(row_idx, 3, btn['control'])
    ws_buttons.cell(row_idx, 4, btn['type'])
    ws_buttons.cell(row_idx, 5, btn['access_event'])
    ws_buttons.cell(row_idx, 6, btn['html_event'])
    ws_buttons.cell(row_idx, 7, btn['status'])
    ws_buttons.cell(row_idx, 8, btn['note'])

    # Status-Farbe
    status_cell = ws_buttons.cell(row_idx, 7)
    if btn['status'] == 'Implementiert':
        status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif btn['status'] == 'Teilweise':
        status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Border
    for col in range(1, 9):
        cell = ws_buttons.cell(row_idx, col)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Spaltenbreiten
ws_buttons.column_dimensions['A'].width = 5
ws_buttons.column_dimensions['B'].width = 28
ws_buttons.column_dimensions['C'].width = 35
ws_buttons.column_dimensions['D'].width = 12
ws_buttons.column_dimensions['E'].width = 35
ws_buttons.column_dimensions['F'].width = 35
ws_buttons.column_dimensions['G'].width = 15
ws_buttons.column_dimensions['H'].width = 45

# Freeze Panes
ws_buttons.freeze_panes = 'A5'

# Auto-Filter
ws_buttons.auto_filter.ref = f"A4:H{len(buttons_data) + 4}"

# Sheet 2: Control-Eigenschaften
ws_props = wb.create_sheet("Control-Eigenschaften")

ws_props['A1'] = "Control-Eigenschaften: Access vs HTML"
ws_props['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_props.merge_cells('A1:G1')
ws_props['A1'].alignment = Alignment(horizontal="center")

ws_props['A2'] = f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
ws_props.merge_cells('A2:G2')
ws_props['A2'].alignment = Alignment(horizontal="center")
ws_props['A2'].font = Font(italic=True, size=10)

# Header
row = 4
headers_props = ['Nr.', 'Formular', 'Control', 'Eigenschaft', 'Access-Wert', 'HTML-Wert', 'Anmerkung']
for col, header in enumerate(headers_props, 1):
    cell = ws_props.cell(row, col, header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Eigenschaften-Daten
props_data = [
    {
        'nr': 1,
        'form': 'frm_va_Auftragstamm',
        'control': 'Auftrag (TextBox)',
        'property': 'Enabled',
        'access_value': 'True/False (je nach Modus)',
        'html_value': 'disabled-Attribut dynamisch',
        'note': 'Funktional identisch'
    },
    {
        'nr': 2,
        'form': 'frm_va_Auftragstamm',
        'control': 'cboVeranstalter',
        'property': 'RowSource',
        'access_value': 'SELECT * FROM tbl_KD_Kundenstamm',
        'html_value': 'Dynamisch via REST API geladen',
        'note': 'Datenquelle unterschiedlich, Ergebnis gleich'
    },
    {
        'nr': 3,
        'form': 'frm_va_Auftragstamm',
        'control': 'auftraegeTable',
        'property': 'RecordSource',
        'access_value': 'qry_VA_Auftragsliste',
        'html_value': 'GET /api/auftraege',
        'note': 'REST API statt Query'
    },
    {
        'nr': 4,
        'form': 'frm_va_Auftragstamm',
        'control': 'gridSchichten',
        'property': 'RecordSource',
        'access_value': 'qry_VA_Schichten',
        'html_value': 'GET /api/auftraege/:id/schichten',
        'note': 'REST API mit VA_ID Parameter'
    },
    {
        'nr': 5,
        'form': 'frm_va_Auftragstamm',
        'control': 'iframe_Einsatzliste',
        'property': 'SourceObject',
        'access_value': 'sub_MA_VA_Zuordnung',
        'html_value': '<iframe src="sub_MA_VA_Zuordnung.html">',
        'note': 'Subform als iframe, Kommunikation via postMessage'
    },
    {
        'nr': 6,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'mitarbeiterTable',
        'property': 'RecordSource',
        'access_value': 'tbl_MA_Mitarbeiterstamm',
        'html_value': 'GET /api/mitarbeiter',
        'note': 'REST API'
    },
    {
        'nr': 7,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'Foto',
        'property': 'Picture',
        'access_value': 'OLE-Object (BMP)',
        'html_value': '<img src="data:image/jpeg;base64,...">',
        'note': 'Base64-encodiert via API'
    },
    {
        'nr': 8,
        'form': 'frm_MA_Mitarbeiterstamm',
        'control': 'IstAktiv',
        'property': 'DefaultValue',
        'access_value': 'True (-1)',
        'html_value': 'checked="checked"',
        'note': 'Boolean-Werte unterschiedlich repräsentiert'
    },
    {
        'nr': 9,
        'form': 'frm_KD_Kundenstamm',
        'control': 'kundenTable',
        'property': 'RecordSource',
        'access_value': 'tbl_KD_Kundenstamm',
        'html_value': 'GET /api/kunden',
        'note': 'REST API mit Filter-Parametern'
    },
    {
        'nr': 10,
        'form': 'frm_KD_Kundenstamm',
        'control': 'kundenTable',
        'property': 'Columns (Kontaktname/Vorname)',
        'access_value': 'Spalten in Datasheet-View',
        'html_value': '<th>Kontaktname</th><th>Vorname</th>',
        'note': 'Jetzt implementiert (vorher fehlend)'
    },
    {
        'nr': 11,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'lstVerfuegbar',
        'property': 'RowSource',
        'access_value': 'qry_MA_Verfuegbar',
        'html_value': 'GET /api/mitarbeiter/verfuegbar',
        'note': 'Filter-Parameter: auftrag_id, datum_id'
    },
    {
        'nr': 12,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'lstGeplant',
        'property': 'RowSource',
        'access_value': 'qry_MA_Geplant',
        'html_value': 'GET /api/planungen?status=1',
        'note': 'Status_ID=1 für "Geplant"'
    },
    {
        'nr': 13,
        'form': 'frm_MA_VA_Schnellauswahl',
        'control': 'lstZugesagt',
        'property': 'RowSource',
        'access_value': 'qry_MA_Zugesagt',
        'html_value': 'GET /api/planungen?status=3',
        'note': 'Status_ID=3 für "Zugesagt"'
    },
    {
        'nr': 14,
        'form': 'Alle Formulare',
        'control': 'Navigationselemente',
        'property': 'RecordsetType',
        'access_value': 'Dynaset/Snapshot',
        'html_value': 'Array aus REST API',
        'note': 'Keine direkte Datenbindung, manuelles Rendering'
    },
    {
        'nr': 15,
        'form': 'Alle Formulare',
        'control': 'Form.Dirty',
        'property': 'Dirty-Flag',
        'access_value': 'Automatisch bei Änderungen',
        'html_value': 'Manuell via JavaScript state.isDirty',
        'note': 'Muss manuell getrackt werden'
    },
]

# Daten einfügen
for row_idx, prop in enumerate(props_data, 5):
    ws_props.cell(row_idx, 1, prop['nr'])
    ws_props.cell(row_idx, 2, prop['form'])
    ws_props.cell(row_idx, 3, prop['control'])
    ws_props.cell(row_idx, 4, prop['property'])
    ws_props.cell(row_idx, 5, prop['access_value'])
    ws_props.cell(row_idx, 6, prop['html_value'])
    ws_props.cell(row_idx, 7, prop['note'])

    # Border
    for col in range(1, 8):
        cell = ws_props.cell(row_idx, col)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Spaltenbreiten
ws_props.column_dimensions['A'].width = 5
ws_props.column_dimensions['B'].width = 28
ws_props.column_dimensions['C'].width = 30
ws_props.column_dimensions['D'].width = 20
ws_props.column_dimensions['E'].width = 35
ws_props.column_dimensions['F'].width = 35
ws_props.column_dimensions['G'].width = 40

# Freeze Panes
ws_props.freeze_panes = 'A5'

# Auto-Filter
ws_props.auto_filter.ref = f"A4:G{len(props_data) + 4}"

# Speichern
output_path = r"C:\Users\guenther.siegert\Desktop\Button_Control_Eigenschaften.xlsx"
wb.save(output_path)
print(f"Excel-Datei erstellt: {output_path}")
